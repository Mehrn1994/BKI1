"""Service reporting routes - filtered reports with Excel/PDF export."""
import io
from datetime import datetime
from flask import Blueprint, jsonify, request, send_file

from app.database import get_db_readonly
from app.security import is_api_rate_limited

reports_bp = Blueprint('reports', __name__)

# Province English-to-Farsi mapping
PROVINCE_MAP = {
    'East Azerbaijan': 'آذربایجان شرقی', 'West Azerbaijan': 'آذربایجان غربی',
    'Ardabil': 'اردبیل', 'Isfahan': 'اصفهان', 'Alborz': 'البرز', 'Ilam': 'ایلام',
    'Bushehr': 'بوشهر', 'Tehran': 'تهران', 'South Khorasan': 'خراسان جنوبی',
    'Razavi Khorasan': 'خراسان رضوی', 'North Khorasan': 'خراسان شمالی',
    'Khuzestan': 'خوزستان', 'Zanjan': 'زنجان', 'Semnan': 'سمنان',
    'Sistan and Baluchestan': 'سیستان و بلوچستان', 'Fars': 'فارس',
    'Qazvin': 'قزوین', 'Qom': 'قم', 'Lorestan': 'لرستان',
    'Mazandaran': 'مازندران', 'Markazi': 'مرکزی', 'Hormozgan': 'هرمزگان',
    'Hamadan': 'همدان', 'Chaharmahal and Bakhtiari': 'چهارمحال و بختیاری',
    'Kurdistan': 'کردستان', 'Kerman': 'کرمان', 'Kermanshah': 'کرمانشاه',
    'Kohgiluyeh and Boyer-Ahmad': 'کهگیلویه و بویراحمد', 'Golestan': 'گلستان',
    'Gilan': 'گیلان', 'Yazd': 'یزد',
}
PROVINCE_MAP_REVERSE = {v: k for k, v in PROVINCE_MAP.items()}


def _detect_point_type(name):
    """Detect point type from branch name/description."""
    if not name:
        return 'نامشخص'
    nl = name.lower()
    if 'atm' in nl or 'خودپرداز' in nl:
        return 'ATM'
    if 'kiosk' in nl or 'کیوسک' in nl or 'cashless' in nl:
        return 'کیوسک'
    if 'bj' in nl or 'bajeh' in nl or 'باجه' in nl:
        return 'باجه'
    if '24' in nl and ('ساعته' in nl or 'saate' in nl):
        return '24 ساعته'
    if 'vsat' in nl:
        return 'VSAT'
    return 'شعبه'


def _province_fa(en_name):
    """Convert English province name to Farsi."""
    return PROVINCE_MAP.get(en_name, en_name or '')


@reports_bp.route('/api/reports/provinces', methods=['GET'])
def get_provinces():
    """Get all unique provinces across all tables."""
    provinces = set()
    try:
        with get_db_readonly() as conn:
            cursor = conn.cursor()
            for table, col in [('lan_ips', 'province'), ('vpls_tunnels', 'province'),
                               ('apn_mali', 'province'), ('apn_ips', 'province'),
                               ('intranet_tunnels', 'province'), ('ptmp_connections', 'province')]:
                try:
                    cursor.execute(f"SELECT DISTINCT {col} FROM {table} WHERE {col} IS NOT NULL AND {col} != ''")
                    for r in cursor.fetchall():
                        val = r[0]
                        fa = PROVINCE_MAP.get(val, val)
                        provinces.add(fa)
                except Exception:
                    pass
        return jsonify(sorted(provinces, key=lambda x: x))
    except Exception:
        return jsonify([])


@reports_bp.route('/api/reports/query', methods=['GET'])
def query_report():
    """Query services by province, service type, and point type."""
    province = request.args.get('province', '').strip()
    service_type = request.args.get('service_type', '').strip()
    point_type = request.args.get('point_type', '').strip()

    results = []
    try:
        with get_db_readonly() as conn:
            cursor = conn.cursor()
            # Get English province name for tables that use English
            province_en = PROVINCE_MAP_REVERSE.get(province, province)

            if service_type in ('', 'all', 'MPLS', 'VPLS', 'MPLS/VPLS'):
                _query_vpls(cursor, province, province_en, point_type, results)

            if service_type in ('', 'all', 'Intranet'):
                _query_intranet(cursor, province, province_en, point_type, results)

            if service_type in ('', 'all', 'APN'):
                _query_apn_mali(cursor, province, province_en, point_type, results)
                _query_apn_int(cursor, province, province_en, point_type, results)

            if service_type in ('', 'all', 'PTMP'):
                _query_ptmp(cursor, province, province_en, point_type, results)

        return jsonify({'status': 'ok', 'count': len(results), 'results': results})
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e), 'results': []}), 500


def _query_vpls(cursor, province_fa, province_en, point_type, results):
    sql = """SELECT branch_name, description, province, ip_address, wan_ip, tunnel_dest,
                    tunnel_name, username, reservation_date, status
             FROM vpls_tunnels WHERE LOWER(status) IN ('reserved','used')"""
    params = []
    if province_en:
        sql += " AND province = ?"
        params.append(province_en)
    cursor.execute(sql, params)
    for r in cursor.fetchall():
        name = r['branch_name'] or (r['description'] or '').replace('** ', '').replace(' **', '').strip()
        pt = _detect_point_type(name)
        if point_type and pt != point_type:
            continue
        results.append({
            'service': 'MPLS/VPLS',
            'branch_name': name,
            'branch_name_fa': name,
            'province': _province_fa(r['province']),
            'point_type': pt,
            'ip': r['ip_address'] or '',
            'wan_ip': r['wan_ip'] or '',
            'tunnel_dest': r['tunnel_dest'] or '',
            'tunnel_name': r['tunnel_name'] or '',
            'username': r['username'] or '',
            'date': r['reservation_date'] or '',
            'status': r['status'] or ''
        })


def _query_intranet(cursor, province_fa, province_en, point_type, results):
    sql = """SELECT tunnel_name, description, province, ip_address, ip_lan, ip_intranet,
                    reserved_by, reserved_at, status
             FROM intranet_tunnels WHERE LOWER(status) = 'reserved'"""
    params = []
    if province_fa:
        sql += " AND province = ?"
        params.append(province_fa)
    if not params and province_en:
        # Try English too
        cursor.execute(sql.replace("province = ?", "(province = ? OR province = ?)"), [province_fa, province_en])
    else:
        cursor.execute(sql, params)
    for r in cursor.fetchall():
        name = (r['description'] or r['tunnel_name'] or '').replace('** ', '').replace(' **', '').strip()
        pt = _detect_point_type(name)
        if point_type and pt != point_type:
            continue
        results.append({
            'service': 'Intranet',
            'branch_name': name,
            'branch_name_fa': name,
            'province': r['province'] or '',
            'point_type': pt,
            'ip': r['ip_address'] or '',
            'wan_ip': r['ip_lan'] or '',
            'tunnel_dest': '',
            'tunnel_name': r['tunnel_name'] or '',
            'username': r['reserved_by'] or '',
            'date': r['reserved_at'] or '',
            'status': r['status'] or ''
        })


def _query_apn_mali(cursor, province_fa, province_en, point_type, results):
    sql = """SELECT branch_name, province, ip_wan, lan_ip, username, reservation_date, type
             FROM apn_mali WHERE branch_name IS NOT NULL AND branch_name != ''"""
    params = []
    if province_fa:
        sql += " AND province = ?"
        params.append(province_fa)
    cursor.execute(sql, params)
    for r in cursor.fetchall():
        name = r['branch_name'] or ''
        pt = _detect_point_type(name)
        if point_type and pt != point_type:
            continue
        results.append({
            'service': 'APN Mali',
            'branch_name': name,
            'branch_name_fa': name,
            'province': r['province'] or '',
            'point_type': pt,
            'ip': r['ip_wan'] or '',
            'wan_ip': r['lan_ip'] or '',
            'tunnel_dest': '',
            'tunnel_name': '',
            'username': r['username'] or '',
            'date': r['reservation_date'] or '',
            'status': 'Active'
        })


def _query_apn_int(cursor, province_fa, province_en, point_type, results):
    sql = """SELECT branch_name, province, ip_wan_apn, lan_ip, username, reservation_date, type
             FROM apn_ips WHERE branch_name IS NOT NULL AND branch_name != ''"""
    params = []
    if province_fa:
        sql += " AND province = ?"
        params.append(province_fa)
    cursor.execute(sql, params)
    for r in cursor.fetchall():
        name = r['branch_name'] or ''
        pt = _detect_point_type(name)
        if point_type and pt != point_type:
            continue
        results.append({
            'service': 'APN INT',
            'branch_name': name,
            'branch_name_fa': name,
            'province': r['province'] or '',
            'point_type': pt,
            'ip': r['ip_wan_apn'] or '',
            'wan_ip': r['lan_ip'] or '',
            'tunnel_dest': '',
            'tunnel_name': '',
            'username': r['username'] or '',
            'date': r['reservation_date'] or '',
            'status': 'Active'
        })


def _query_ptmp(cursor, province_fa, province_en, point_type, results):
    sql = """SELECT branch_name, branch_name_en, province, interface_name, lan_ip,
                    description, username, reservation_date, status
             FROM ptmp_connections WHERE (branch_name IS NOT NULL OR branch_name_en IS NOT NULL)"""
    params = []
    if province_fa:
        sql += " AND (province = ? OR province = ?)"
        params.extend([province_fa, province_en])
    cursor.execute(sql, params)
    for r in cursor.fetchall():
        name = r['branch_name'] or r['branch_name_en'] or ''
        pt = _detect_point_type(name)
        if point_type and pt != point_type:
            continue
        results.append({
            'service': 'PTMP',
            'branch_name': name,
            'branch_name_fa': name,
            'province': r['province'] or '',
            'point_type': pt,
            'ip': r['interface_name'] or '',
            'wan_ip': r['lan_ip'] or '',
            'tunnel_dest': '',
            'tunnel_name': r['description'] or '',
            'username': r['username'] or '',
            'date': r['reservation_date'] or '',
            'status': r['status'] or ''
        })


@reports_bp.route('/api/reports/export/excel', methods=['GET'])
def export_excel():
    """Export filtered report as Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Get same query params
        resp = query_report()
        data = resp.get_json()
        rows = data.get('results', [])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Report'
        ws.sheet_view.rightToLeft = True

        # Header
        headers = ['ردیف', 'نوع سرویس', 'استان', 'نام نقطه', 'نوع نقطه', 'IP', 'WAN IP', 'Tunnel', 'کاربر', 'تاریخ']
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Data rows
        for idx, r in enumerate(rows, 1):
            row_data = [idx, r['service'], r['province'], r['branch_name'], r['point_type'],
                        r['ip'], r['wan_ip'], r['tunnel_name'], r['username'], r['date']]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=idx+1, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center' if col <= 2 else 'right', vertical='center')

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        province = request.args.get('province', 'all')
        svc = request.args.get('service_type', 'all')
        filename = f'report_{province}_{svc}_{datetime.now().strftime("%Y%m%d")}.xlsx'

        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except ImportError:
        return jsonify({'status': 'error', 'error': 'openpyxl not installed'}), 500
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500
