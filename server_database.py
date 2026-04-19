"""
Network Config Portal Server - COMPLETE FIXED VERSION
All APIs fixed + DB Manager only for Sahebdel
"""

from flask import Flask, jsonify, request, render_template, Response, send_from_directory, send_file, stream_with_context
from flask_cors import CORS
import sqlite3
import os
import subprocess
import platform
import shutil
from datetime import datetime, timedelta
import json
import hashlib
import secrets
import pandas as pd
import time
import threading
import ipaddress as _ipaddr
# Email service removed - ticketing disabled

app = Flask(__name__)
CORS(app, origins=["http://localhost:5000", "http://127.0.0.1:5000"])
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload size

# Optional: Remote connection module (SSH/Telnet/RDP)
# Requires: pip install flask-socketio paramiko eventlet
socketio = None
REMOTE_ENABLED = False
try:
    from flask_socketio import SocketIO
    socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')
    from remote_connect import remote_bp, register_socketio_handlers
    app.register_blueprint(remote_bp)
    register_socketio_handlers(socketio)
    REMOTE_ENABLED = True
    print("✅ Remote Connection module loaded (SSH/Telnet/RDP)")
except ImportError as e:
    print(f"⚠️  Remote Connection module disabled - missing package: {e}")
    print("   Install with: pip install flask-socketio paramiko eventlet")

# ==================== RATE LIMITING ====================
_rate_lock = threading.Lock()
login_attempts = {}  # {ip: [timestamp, ...]}
LOGIN_MAX_ATTEMPTS = 5
LOGIN_WINDOW_SECONDS = 300  # 5 minutes

def is_rate_limited(ip):
    """Check if an IP has exceeded login attempt limit (thread-safe)"""
    with _rate_lock:
        now = time.time()
        if ip not in login_attempts:
            login_attempts[ip] = []
        login_attempts[ip] = [t for t in login_attempts[ip] if now - t < LOGIN_WINDOW_SECONDS]
        return len(login_attempts[ip]) >= LOGIN_MAX_ATTEMPTS

def record_login_attempt(ip):
    """Record a failed login attempt (thread-safe)"""
    with _rate_lock:
        if ip not in login_attempts:
            login_attempts[ip] = []
        login_attempts[ip].append(time.time())

# General API rate limiting for sensitive write endpoints
_api_rate = {}  # {ip+endpoint: [timestamps]}
API_RATE_LIMIT = 30  # max requests per window
API_RATE_WINDOW = 60  # seconds

def is_api_rate_limited(ip, endpoint):
    """Check if an IP has exceeded API rate limit (thread-safe)"""
    with _rate_lock:
        key = f"{ip}:{endpoint}"
        now = time.time()
        if key not in _api_rate:
            _api_rate[key] = []
        _api_rate[key] = [t for t in _api_rate[key] if now - t < API_RATE_WINDOW]
        if len(_api_rate[key]) >= API_RATE_LIMIT:
            return True
        _api_rate[key].append(now)
        return False

def validate_octet(value):
    """Validate that value is a valid IP octet (0-255)"""
    try:
        n = int(value)
        return 0 <= n <= 255
    except (ValueError, TypeError):
        return False

def validate_ip(ip):
    """Validate IPv4 address format"""
    import re
    if not ip or not isinstance(ip, str):
        return False
    pattern = r'^(\d{1,3})\.(\d{1,3})\.(\d{1,3})\.(\d{1,3})$'
    match = re.match(pattern, ip)
    if not match:
        return False
    return all(0 <= int(g) <= 255 for g in match.groups())

# ==================== AUTO-RELEASE EXPIRED RESERVATIONS ====================
AUTO_RELEASE_INTERVAL = 3600 * 6  # Check every 6 hours
auto_release_active = True

def auto_release_expired_reservations():
    """
    Automatically release LAN IPs that:
    - Have status = 'Reserved' (not 'Used' or 'activated')
    - Have expiry_date < today (60 days passed)
    
    This runs in a background thread every 6 hours.
    """
    global auto_release_active
    
    print("🔄 Auto-release checker started!")
    
    while auto_release_active:
        try:
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            today = datetime.now().strftime('%Y-%m-%d')
            
            # Find expired reservations that are still 'reserved' (not activated/used)
            cursor.execute("""
                SELECT id, octet2, octet3, branch_name, username, reservation_date, expiry_date
                FROM reserved_ips 
                WHERE expiry_date < ? 
                AND (status = 'reserved' OR status IS NULL)
            """, (today,))
            
            expired = cursor.fetchall()
            
            if expired:
                print(f"🗑️ Found {len(expired)} expired reservations to release")
                
                for row in expired:
                    octet2 = row['octet2']
                    octet3 = row['octet3']
                    branch_name = row['branch_name']
                    expiry_date = row['expiry_date']
                    
                    # Release from lan_ips
                    cursor.execute("""
                        UPDATE lan_ips 
                        SET username = NULL, reservation_date = NULL, branch_name = NULL, status = 'Free'
                        WHERE octet2 = ? AND octet3 = ? AND status = 'Reserved'
                    """, (octet2, octet3))
                    
                    # Delete from reserved_ips
                    cursor.execute("DELETE FROM reserved_ips WHERE id = ?", (row['id'],))
                    
                    print(f"   ✅ Released: 10.{octet2}.{octet3}.0/24 ({branch_name}) - expired on {expiry_date}")
                
                conn.commit()
                
                # Log the auto-release activity
                try:
                    log_activity('info', 'آزادسازی خودکار', f'{len(expired)} IP منقضی شده آزاد شد', 'System')
                except Exception:
                    pass
            else:
                print(f"✓ No expired reservations found (checked at {datetime.now().strftime('%H:%M:%S')})")
            
            conn.close()
            
        except Exception as e:
            print(f"❌ Auto-release error: {e}")
        
        # Sleep for 6 hours
        time.sleep(AUTO_RELEASE_INTERVAL)

def start_auto_release_thread():
    """Start the auto-release background thread"""
    global auto_release_active
    auto_release_active = True
    thread = threading.Thread(target=auto_release_expired_reservations, daemon=True)
    thread.start()
    print("✅ Auto-release thread started (checks every 6 hours)")

# Add response headers for caching
@app.after_request
def add_cache_headers(response):
    # Cache static assets only
    if request.path.endswith(('.css', '.js', '.png', '.jpg', '.ico')):
        response.headers['Cache-Control'] = 'public, max-age=3600'
    # API responses must never be cached publicly (contain user-specific data)
    elif request.path.startswith('/api/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    # Add security headers
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    return response

DB_PATH = os.path.join(os.path.dirname(__file__), 'data', 'network_ipam.db')
LIVE_DB_PATH = os.path.join(os.path.dirname(__file__), 'data', 'live.db')
BACKUP_DIR = os.path.join(os.path.dirname(__file__), 'data', 'backups')
ACTIVITY_LOG = os.path.join(os.path.dirname(__file__), 'data', 'activity.json')
CHAT_UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'data', 'chat_files')
os.makedirs(CHAT_UPLOAD_DIR, exist_ok=True)

os.makedirs(BACKUP_DIR, exist_ok=True)
os.makedirs(os.path.dirname(ACTIVITY_LOG), exist_ok=True)

# Simple cache for stats (60 seconds - much longer)
_stats_cache = {'data': None, 'time': 0}
STATS_CACHE_SECONDS = 60  # Cache for 1 minute

ALLOWED_USERS = ["Yarian", "Sattari", "Barari", "Sahebdel", "Vahedi", "Aghajani", "Hossein", "Rezaei", "Bagheri"]
DB_ADMIN_USER = "Sahebdel"

# ==================== LIVE DB MIRROR ====================
# live.db = دیتابیس فعال سیستم (تیم روی این کار می‌کند)
# network_ipam.db = دیتابیس اصلی (هر شب از live.db آپدیت می‌شود)

_MIRROR_TABLES = [
    'lan_ips', 'apn_ips', 'apn_mali', 'intranet_tunnels',
    'ptmp_connections', 'reserved_ips', 'tunnel200_ips', 'tunnel_mali',
    'vpls_tunnels', 'user_passwords', 'custom_translations',
]
_mirror_ready = False
_mirror_lock = threading.Lock()


def _fix_main_db_broken_triggers():
    """
    تریگرهای شکسته را از network_ipam.db پاک می‌کند.

    مشکل: تریگرهای _trg_* با NEW.id/OLD.id ساخته شده‌اند ولی جداولی مثل
    user_passwords ستون id ندارند. این باعث خطای login می‌شود:
      sqlite3.OperationalError: no such column: NEW.id

    راه‌حل: تریگرهای _trg_* را از network_ipam.db حذف می‌کنیم.
    (change tracking فقط باید در live.db باشد، نه در network_ipam.db)
    """
    if not os.path.exists(DB_PATH):
        return
    try:
        conn = sqlite3.connect(DB_PATH)
        broken = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='trigger' AND name LIKE '_trg_%'"
        ).fetchall()
        for (tname,) in broken:
            conn.execute(f"DROP TRIGGER IF EXISTS {tname}")
        if broken:
            conn.commit()
            print(f"✅ [Mirror] {len(broken)} تریگر شکسته از network_ipam.db حذف شد")
        conn.close()
    except Exception as e:
        print(f"⚠️ [Mirror] _fix_main_db_broken_triggers error: {e}")


def _init_live_db():
    """live.db را راه‌اندازی می‌کند؛ اگر نبود از network_ipam.db کپی می‌کند."""
    global _mirror_ready
    with _mirror_lock:
        if _mirror_ready:
            return
        # ابتدا تریگرهای شکسته network_ipam.db را پاک می‌کنیم (رفع خطای login)
        _fix_main_db_broken_triggers()
        if not os.path.exists(LIVE_DB_PATH):
            if os.path.exists(DB_PATH):
                shutil.copy2(DB_PATH, LIVE_DB_PATH)
                print(f"✅ [Mirror] live.db ایجاد شد (کپی از network_ipam.db)")
        _setup_mirror_triggers()
        _mirror_ready = True


def _setup_mirror_triggers():
    """جدول _change_log و تریگرهای خودکار را در live.db می‌سازد."""
    conn = sqlite3.connect(LIVE_DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS _change_log (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            ts         TEXT    NOT NULL DEFAULT CURRENT_TIMESTAMP,
            table_name TEXT    NOT NULL,
            op         TEXT    NOT NULL,
            row_id     INTEGER NOT NULL,
            row_json   TEXT,
            merged     INTEGER DEFAULT 0,
            merged_at  TEXT
        )""")
    conn.execute("CREATE INDEX IF NOT EXISTS _idx_cl ON _change_log(merged, ts)")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS _merge_log (
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            ts      TEXT DEFAULT CURRENT_TIMESTAMP,
            applied INTEGER, skipped INTEGER, errors INTEGER, dry_run INTEGER,
            detail  TEXT
        )""")
    conn.commit()

    active = 0
    for table in _MIRROR_TABLES:
        try:
            if not conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)
            ).fetchone():
                continue
            # Drop and recreate to ensure correct rowid-based triggers
            conn.execute(f"DROP TRIGGER IF EXISTS _trg_{table}_i")
            conn.execute(f"DROP TRIGGER IF EXISTS _trg_{table}_u")
            conn.execute(f"DROP TRIGGER IF EXISTS _trg_{table}_d")
            conn.execute(
                f"CREATE TRIGGER _trg_{table}_i "
                f"AFTER INSERT ON {table} BEGIN "
                f"INSERT INTO _change_log(table_name,op,row_id) "
                f"VALUES('{table}','I',NEW.rowid); END"
            )
            conn.execute(
                f"CREATE TRIGGER _trg_{table}_u "
                f"AFTER UPDATE ON {table} BEGIN "
                f"INSERT INTO _change_log(table_name,op,row_id) "
                f"VALUES('{table}','U',NEW.rowid); END"
            )
            cols = [r[1] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]
            jp = ",".join(f"'{c}',OLD.{c}" for c in cols)
            conn.execute(
                f"CREATE TRIGGER _trg_{table}_d "
                f"AFTER DELETE ON {table} BEGIN "
                f"INSERT INTO _change_log(table_name,op,row_id,row_json) "
                f"VALUES('{table}','D',OLD.rowid,json_object({jp})); END"
            )
            active += 1
        except Exception as e:
            print(f"[Mirror] Trigger {table}: {e}")

    conn.commit()
    conn.close()
    print(f"✅ [Mirror] Triggers فعال روی {active} جدول — live.db آماده")


def _mirror_merge_to_main(dry_run=False):
    """تغییرات live.db را به network_ipam.db اعمال می‌کند."""
    _init_live_db()
    live = sqlite3.connect(LIVE_DB_PATH)
    live.row_factory = sqlite3.Row
    main = sqlite3.connect(DB_PATH)
    main.row_factory = sqlite3.Row

    pending = live.execute(
        "SELECT * FROM _change_log WHERE merged=0 ORDER BY id ASC"
    ).fetchall()

    stats = {'total': len(pending), 'applied': 0, 'skipped': 0, 'errors': 0,
             'dry_run': dry_run, 'detail': [],
             'ts': datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    merged_ids = []

    for ch in pending:
        table, op, rid = ch['table_name'], ch['op'], ch['row_id']
        try:
            if op == 'D':
                if not dry_run:
                    main.execute(f"DELETE FROM {table} WHERE id=?", (rid,))
                stats['applied'] += 1
                stats['detail'].append(f"DELETE {table} id={rid}")
            elif op in ('I', 'U'):
                row = live.execute(f"SELECT * FROM {table} WHERE id=?", (rid,)).fetchone()
                if row is None:
                    stats['skipped'] += 1
                    continue
                data = dict(row)
                cols = list(data.keys())
                exists = main.execute(f"SELECT id FROM {table} WHERE id=?", (rid,)).fetchone()
                if not dry_run:
                    if exists:
                        non_id = [c for c in cols if c != 'id']
                        sql = f"UPDATE {table} SET {','.join(c+'=?' for c in non_id)} WHERE id=?"
                        main.execute(sql, [data[c] for c in non_id] + [rid])
                    else:
                        ph = ','.join('?' * len(cols))
                        main.execute(
                            f"INSERT OR REPLACE INTO {table}({','.join(cols)}) VALUES({ph})",
                            [data[c] for c in cols]
                        )
                stats['applied'] += 1
                stats['detail'].append(f"{op} {table} id={rid}")
            merged_ids.append(ch['id'])
        except Exception as e:
            stats['errors'] += 1
            stats['detail'].append(f"ERROR {table} id={rid}: {e}")

    if not dry_run:
        if stats['errors'] == 0:
            main.commit()
        else:
            main.rollback()
        now_s = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if merged_ids:
            live.execute(
                f"UPDATE _change_log SET merged=1,merged_at=? "
                f"WHERE id IN ({','.join('?'*len(merged_ids))})",
                [now_s] + merged_ids
            )
        live.execute(
            "DELETE FROM _change_log WHERE merged=1 "
            "AND ts < datetime('now','-24 hours')"
        )
        live.execute(
            "INSERT INTO _merge_log(applied,skipped,errors,dry_run,detail) VALUES(?,?,?,?,?)",
            (stats['applied'], stats['skipped'], stats['errors'],
             1 if dry_run else 0, json.dumps(stats['detail'], ensure_ascii=False))
        )
        live.execute(
            "DELETE FROM _merge_log WHERE id NOT IN "
            "(SELECT id FROM _merge_log ORDER BY id DESC LIMIT 100)"
        )
        live.commit()

    main.close()
    live.close()
    return stats


def _nightly_mirror_thread():
    """هر شب 23:30 merge خودکار اجرا می‌کند."""
    print("[Mirror] Nightly merge scheduler started (23:30 daily)")
    while True:
        now = datetime.now()
        target = now.replace(hour=23, minute=30, second=0, microsecond=0)
        if now >= target:
            target += timedelta(days=1)
        time.sleep((target - now).total_seconds())
        try:
            stats = _mirror_merge_to_main(dry_run=False)
            print(f"[Mirror] Nightly merge done — "
                  f"applied={stats['applied']} skipped={stats['skipped']} errors={stats['errors']}")
        except Exception as e:
            print(f"[Mirror] Nightly merge error: {e}")


def get_db():
    """همیشه به live.db وصل می‌شود (دیتابیس فعال سیستم)."""
    _init_live_db()
    conn = sqlite3.connect(LIVE_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

SALT_KEY = os.environ.get('BKI_SALT_KEY', 'BKI-Network-Portal-2026')

def hash_password(password, use_salt=True):
    if use_salt:
        salted = f"{SALT_KEY}:{password}:{SALT_KEY}"
        return hashlib.sha256(salted.encode()).hexdigest()
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(stored_hash, password):
    """Verify password with auto-migration: try salted first, then legacy"""
    if stored_hash == hash_password(password, use_salt=True):
        return True
    # Fallback: check legacy (unsalted) hash for existing users
    if stored_hash == hash_password(password, use_salt=False):
        return True
    return False

def init_tables():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""CREATE TABLE IF NOT EXISTS user_passwords (
        username TEXT PRIMARY KEY, password_hash TEXT NOT NULL, created_at TEXT, last_login TEXT)""")
    cursor.execute("""CREATE TABLE IF NOT EXISTS reserved_ips (
        id INTEGER PRIMARY KEY AUTOINCREMENT, province TEXT, octet2 INTEGER, octet3 INTEGER,
        branch_name TEXT, username TEXT, reservation_date TEXT, expiry_date TEXT,
        request_number TEXT, point_type TEXT, mehregostar_code TEXT,
        status TEXT DEFAULT 'reserved', activated_at TEXT, config_type TEXT)""")
    
    # Add status column to existing reserved_ips table if not exists
    try:
        cursor.execute("ALTER TABLE reserved_ips ADD COLUMN status TEXT DEFAULT 'reserved'")
    except Exception:
        pass  # Column already exists
    try:
        cursor.execute("ALTER TABLE reserved_ips ADD COLUMN activated_at TEXT")
    except Exception:
        pass
    try:
        cursor.execute("ALTER TABLE reserved_ips ADD COLUMN config_type TEXT")
    except Exception:
        pass
    
    # Create indexes for faster COUNT queries (HUGE performance improvement)
    try:
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_lan_ips_username ON lan_ips(username)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_lan_ips_branch ON lan_ips(branch_name)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_lan_ips_province ON lan_ips(province)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_apn_ips_username ON apn_ips(username)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_apn_mali_username ON apn_mali(username)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_intranet_tunnels_status ON intranet_tunnels(status)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_tunnel200_status ON tunnel200_ips(status)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_tunnel_mali_status ON tunnel_mali(status)")
        # Additional indexes for frequently queried columns
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_lan_ips_status ON lan_ips(status)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_lan_ips_octet2 ON lan_ips(octet2)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_lan_ips_octet2_octet3 ON lan_ips(octet2, octet3)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_reserved_ips_expiry ON reserved_ips(expiry_date)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_reserved_ips_status ON reserved_ips(status)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_vpls_tunnels_province ON vpls_tunnels(province)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_vpls_tunnels_branch ON vpls_tunnels(branch_name)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_apn_ips_branch ON apn_ips(branch_name)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_apn_mali_branch ON apn_mali(branch_name)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_tunnel_mali_branch ON tunnel_mali(branch_name)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_tunnel200_branch ON tunnel200_ips(branch_name)")
        print("✓ Database indexes created")
    except Exception as e:
        print(f"⚠️ Index creation: {e}")

    # Tickets and Email tables removed - ticketing system disabled

    # PTMP Serial connections table
    cursor.execute("""CREATE TABLE IF NOT EXISTS ptmp_connections (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        interface_name TEXT NOT NULL,
        description TEXT,
        branch_name TEXT,
        branch_name_en TEXT,
        bandwidth TEXT,
        ip_type TEXT,
        ip_address TEXT,
        ip_mask TEXT,
        encapsulation TEXT,
        province TEXT,
        province_abbr TEXT,
        router_hostname TEXT,
        router_file TEXT,
        status TEXT DEFAULT 'Used',
        username TEXT,
        reservation_date TEXT,
        lan_ip TEXT)""")
    try:
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_ptmp_branch ON ptmp_connections(branch_name)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_ptmp_branch_en ON ptmp_connections(branch_name_en)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_ptmp_province ON ptmp_connections(province)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_ptmp_status ON ptmp_connections(status)")
    except Exception as e:
        print(f"⚠️ PTMP index creation: {e}")

    # Custom translations table (user-added Finglish→Persian)
    cursor.execute("""CREATE TABLE IF NOT EXISTS custom_translations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name_en TEXT UNIQUE NOT NULL,
        name_fa TEXT NOT NULL,
        added_by TEXT,
        added_at TEXT)""")

    # VPLS/MPLS tunnels table
    cursor.execute("""CREATE TABLE IF NOT EXISTS vpls_tunnels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ip_address TEXT,
        hub_ip TEXT,
        branch_ip TEXT,
        tunnel_name TEXT,
        description TEXT,
        province TEXT,
        branch_name TEXT,
        wan_ip TEXT,
        tunnel_dest TEXT,
        status TEXT DEFAULT 'Free',
        username TEXT,
        reservation_date TEXT)""")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_vpls_tunnels_status ON vpls_tunnels(status)")

    # Add lan_ip column to vpls_tunnels if not exists
    try:
        cursor.execute("ALTER TABLE vpls_tunnels ADD COLUMN lan_ip TEXT")
    except Exception:
        pass  # Column already exists

    # Add branch_name column to intranet_tunnels if not exists
    try:
        cursor.execute("ALTER TABLE intranet_tunnels ADD COLUMN branch_name TEXT")
    except Exception:
        pass  # Column already exists

    # Chat messages table
    cursor.execute("""CREATE TABLE IF NOT EXISTS chat_messages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sender TEXT NOT NULL,
        room TEXT NOT NULL DEFAULT 'general',
        message TEXT,
        file_name TEXT,
        file_path TEXT,
        timestamp TEXT NOT NULL)""")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_chat_room ON chat_messages(room)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_chat_timestamp ON chat_messages(timestamp)")

    # Sessions table for proper auth
    cursor.execute("""CREATE TABLE IF NOT EXISTS sessions (
        token TEXT PRIMARY KEY,
        username TEXT NOT NULL,
        created_at TEXT NOT NULL,
        expires_at TEXT NOT NULL,
        ip_address TEXT)""")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_sessions_username ON sessions(username)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_sessions_expires ON sessions(expires_at)")

    conn.commit()
    conn.close()

init_tables()

# ==================== SESSION MANAGEMENT ====================
SESSION_LIFETIME_HOURS = 8

def create_session(username, ip_address=None):
    """Create a new session token for a user."""
    token = secrets.token_urlsafe(48)
    now = datetime.now()
    expires = now + timedelta(hours=SESSION_LIFETIME_HOURS)
    conn = get_db()
    # Remove expired sessions for this user
    conn.execute("DELETE FROM sessions WHERE username=? AND expires_at<?",
                 (username, now.strftime('%Y-%m-%d %H:%M:%S')))
    conn.execute("INSERT INTO sessions (token, username, created_at, expires_at, ip_address) VALUES (?,?,?,?,?)",
                 (token, username, now.strftime('%Y-%m-%d %H:%M:%S'),
                  expires.strftime('%Y-%m-%d %H:%M:%S'), ip_address))
    conn.commit()
    conn.close()
    return token

def validate_session(token):
    """Validate a session token and return username if valid."""
    if not token or len(token) < 10:
        return None
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT username FROM sessions WHERE token=? AND expires_at>?",
                   (token, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    row = cursor.fetchone()
    conn.close()
    return row['username'] if row else None

def get_current_user():
    """Get authenticated user from session token in request headers/body/args."""
    # Check Authorization header: Bearer <token>
    auth = request.headers.get('Authorization', '')
    if auth.startswith('Bearer '):
        return validate_session(auth[7:])
    # Check X-Session-Token header
    token = request.headers.get('X-Session-Token', '')
    if token:
        return validate_session(token)
    # Check JSON body
    if request.is_json and request.json:
        token = request.json.get('session_token', '')
        if token:
            return validate_session(token)
    # Check query parameter
    token = request.args.get('session_token', '')
    if token:
        return validate_session(token)
    return None

# Auto-import PTMP from router configs if table is empty (first run only)
def _check_ptmp_import():
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM ptmp_connections")
        count = cursor.fetchone()[0]
        conn.close()
        if count == 0:
            print("PTMP table empty, running initial import from router configs...")
            from parse_router_configs import import_serial_to_db
            import_serial_to_db()
    except Exception as e:
        print(f"PTMP auto-import check: {e}")

_check_ptmp_import()

def log_activity(atype, title, desc, user="System"):
    try:
        activities = []
        if os.path.exists(ACTIVITY_LOG):
            with open(ACTIVITY_LOG, 'r', encoding='utf-8') as f:
                activities = json.load(f)
        activities.insert(0, {'type': atype, 'title': title, 'description': desc, 'user': user, 'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')})
        activities = activities[:100]
        with open(ACTIVITY_LOG, 'w', encoding='utf-8') as f:
            json.dump(activities, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Log error: {e}")

# ==================== CHAT SYSTEM ====================
from werkzeug.utils import secure_filename
import uuid

# Track online users via heartbeat: {username: last_seen_timestamp}
chat_online_heartbeats = {}

CHAT_ALLOWED_EXT = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'xlsx', 'xls', 'doc', 'docx', 'txt', 'zip', 'rar', 'csv'}

def allowed_chat_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in CHAT_ALLOWED_EXT

@app.route('/api/chat/history', methods=['GET'])
def chat_history():
    """Get chat message history for a room"""
    room = request.args.get('room', 'general')
    limit = min(int(request.args.get('limit', 50)), 200)
    after_id = int(request.args.get('after_id', 0))
    conn = get_db()
    cursor = conn.cursor()
    if after_id > 0:
        cursor.execute("""
            SELECT id, sender, room, message, file_name, file_path, timestamp
            FROM chat_messages WHERE room = ? AND id > ?
            ORDER BY id ASC LIMIT ?
        """, (room, after_id, limit))
    else:
        cursor.execute("""
            SELECT id, sender, room, message, file_name, file_path, timestamp
            FROM chat_messages WHERE room = ?
            ORDER BY id DESC LIMIT ?
        """, (room, limit))
    rows = cursor.fetchall()
    conn.close()
    messages = [dict(r) for r in rows]
    if after_id == 0:
        messages.reverse()
    return jsonify({'messages': messages})

@app.route('/api/chat/send', methods=['POST'])
def chat_send_message():
    """Send a chat message via HTTP POST"""
    try:
        data = request.json
        sender = data.get('sender', '')
        room = data.get('room', 'general')
        message = data.get('message', '')
        file_name = data.get('file_name', '')
        file_path_val = data.get('file_path', '')

        if not sender or (not message and not file_name):
            return jsonify({'status': 'error', 'error': 'Empty message'}), 400

        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO chat_messages (sender, room, message, file_name, file_path, timestamp)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (sender, room, message, file_name, file_path_val, now))
        msg_id = cursor.lastrowid
        conn.commit()
        conn.close()

        return jsonify({
            'status': 'ok',
            'message': {
                'id': msg_id,
                'sender': sender,
                'room': room,
                'message': message,
                'file_name': file_name,
                'file_path': file_path_val,
                'timestamp': now
            }
        })
    except Exception as e:
        print(f"❌ Chat send error: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/chat/heartbeat', methods=['POST'])
def chat_heartbeat():
    """Update user online status and return online users list"""
    data = request.json or {}
    username = data.get('username', '')
    if username and username != 'unknown':
        chat_online_heartbeats[username] = time.time()
    # Remove users not seen in last 15 seconds
    now = time.time()
    offline = [u for u, t in chat_online_heartbeats.items() if now - t > 15]
    for u in offline:
        del chat_online_heartbeats[u]
    return jsonify({'users': list(chat_online_heartbeats.keys())})

@app.route('/api/chat/poll', methods=['GET'])
def chat_poll():
    """Poll for new messages across rooms user cares about"""
    username = request.args.get('username', '')
    after_id = int(request.args.get('after_id', 0))
    conn = get_db()
    cursor = conn.cursor()
    # Get new messages in general + any DM rooms involving this user
    cursor.execute("""
        SELECT id, sender, room, message, file_name, file_path, timestamp
        FROM chat_messages
        WHERE id > ? AND (room = 'general' OR room LIKE ? OR room LIKE ?)
        ORDER BY id ASC LIMIT 50
    """, (after_id, f'dm_{username}_%', f'dm_%_{username}'))
    rows = cursor.fetchall()
    conn.close()
    messages = [dict(r) for r in rows]
    return jsonify({'messages': messages})

@app.route('/api/chat/rooms', methods=['GET'])
def chat_rooms():
    """Get list of rooms user has messages in"""
    username = request.args.get('username', '')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT room FROM chat_messages
        WHERE room = 'general' OR room LIKE ? OR room LIKE ?
        ORDER BY room
    """, (f'{username}_%', f'%_{username}'))
    rooms = [row['room'] for row in cursor.fetchall()]
    conn.close()
    if 'general' not in rooms:
        rooms.insert(0, 'general')
    return jsonify({'rooms': rooms})

@app.route('/api/chat/upload', methods=['POST'])
def chat_upload():
    """Upload a file for chat"""
    if 'file' not in request.files:
        return jsonify({'status': 'error', 'error': 'No file'}), 400
    f = request.files['file']
    if f.filename == '' or not allowed_chat_file(f.filename):
        return jsonify({'status': 'error', 'error': 'Invalid file type'}), 400
    ext = f.filename.rsplit('.', 1)[1].lower()
    unique_name = f"{uuid.uuid4().hex[:12]}.{ext}"
    f.save(os.path.join(CHAT_UPLOAD_DIR, unique_name))
    return jsonify({'status': 'ok', 'file_name': f.filename, 'file_path': unique_name})

@app.route('/data/chat_files/<path:filename>')
def chat_file_serve(filename):
    """Serve uploaded chat files"""
    return send_from_directory(CHAT_UPLOAD_DIR, filename)

@app.route('/api/chat/online', methods=['GET'])
def chat_online():
    """Get list of online users"""
    now = time.time()
    offline = [u for u, t in chat_online_heartbeats.items() if now - t > 15]
    for u in offline:
        del chat_online_heartbeats[u]
    return jsonify({'users': list(chat_online_heartbeats.keys())})

print("✅ Chat system ready (HTTP polling mode)")

# ==================== PAGE ROUTES ====================
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login')
def login_page():
    return render_template('login.html')

@app.route('/intranet')
def intranet_page():
    return render_template('intranet.html')

@app.route('/apn-int')
def apn_int_page():
    return render_template('apn_int.html')

@app.route('/apn-mali')
def apn_mali_page():
    return render_template('apn_mali.html')

@app.route('/ptmp')
def ptmp_page():
    return render_template('ptmp.html')

@app.route('/mpls-vpls')
def mpls_vpls_page():
    return render_template('mpls_vpls.html')

@app.route('/config-wizard')
def config_wizard_page():
    return render_template('config_wizard.html')

@app.route('/config-both')
def config_both_page():
    return render_template('config_both.html')

@app.route('/reserve-lan')
def reserve_lan_page():
    return render_template('reserve_lan.html')

@app.route('/db-manager')
def db_manager_page():
    return render_template('db_manager.html')

# Tickets page removed

# ==================== AUTH APIs ====================
@app.route('/api/users', methods=['GET'])
def get_users():
    conn = get_db()
    cursor = conn.cursor()
    users_info = []
    for username in ALLOWED_USERS:
        cursor.execute("SELECT username FROM user_passwords WHERE username = ?", (username,))
        users_info.append({"name": username, "registered": cursor.fetchone() is not None})
    conn.close()
    return jsonify({"users": users_info})

@app.route('/api/check-user', methods=['POST'])
def check_user():
    data = request.json
    username = data.get('username')
    if username not in ALLOWED_USERS:
        return jsonify({"error": "کاربر مجاز نیست"}), 403
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT username FROM user_passwords WHERE username = ?", (username,))
    has_password = cursor.fetchone() is not None
    conn.close()
    return jsonify({"username": username, "has_password": has_password})

@app.route('/api/register', methods=['POST'])
def register_user():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    if username not in ALLOWED_USERS:
        return jsonify({"error": "کاربر مجاز نیست"}), 403
    if not password or len(password) < 8:
        return jsonify({"error": "رمز باید حداقل ۸ کاراکتر باشد"}), 400
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT username FROM user_passwords WHERE username = ?", (username,))
    if cursor.fetchone():
        conn.close()
        return jsonify({"error": "قبلا ثبت نام کرده"}), 400
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor.execute("INSERT INTO user_passwords VALUES (?, ?, ?, ?)", (username, hash_password(password), now, now))
    conn.commit()
    conn.close()
    log_activity('success', 'ثبت نام', username, username)
    return jsonify({"success": True})

@app.route('/api/change-password', methods=['POST'])
def change_password():
    data = request.json
    username = data.get('username')
    old_password = data.get('old_password')
    new_password = data.get('new_password')
    if username not in ALLOWED_USERS:
        return jsonify({"error": "کاربر مجاز نیست"}), 403
    if not new_password or len(new_password) < 8:
        return jsonify({"error": "رمز جدید باید حداقل ۸ کاراکتر باشد"}), 400
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT password_hash FROM user_passwords WHERE username = ?", (username,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "کاربر یافت نشد"}), 404
    if not verify_password(row['password_hash'], old_password):
        conn.close()
        return jsonify({"error": "رمز فعلی اشتباه است"}), 401
    cursor.execute("UPDATE user_passwords SET password_hash = ? WHERE username = ?",
                   (hash_password(new_password), username))
    conn.commit()
    conn.close()
    log_activity('success', 'تغییر رمز', username, username)
    return jsonify({"success": True, "message": "رمز با موفقیت تغییر کرد"})

@app.route('/api/login', methods=['POST'])
def login():
    client_ip = request.remote_addr
    if is_rate_limited(client_ip):
        return jsonify({"success": False, "message": "تعداد تلاش بیش از حد مجاز. لطفا ۵ دقیقه صبر کنید."}), 429
    data = request.json
    username = data.get('username')
    password = data.get('password')
    if username not in ALLOWED_USERS:
        record_login_attempt(client_ip)
        return jsonify({"success": False, "message": "کاربر مجاز نیست"}), 403
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT password_hash FROM user_passwords WHERE username = ?", (username,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({"success": False, "message": "ابتدا رمز تعیین کنید", "need_register": True}), 401
    if not verify_password(row['password_hash'], password):
        record_login_attempt(client_ip)
        conn.close()
        return jsonify({"success": False, "message": "رمز اشتباه"}), 401
    # Auto-migrate to salted hash if using legacy
    if row['password_hash'] == hash_password(password, use_salt=False) and row['password_hash'] != hash_password(password, use_salt=True):
        cursor.execute("UPDATE user_passwords SET password_hash = ? WHERE username = ?", (hash_password(password), username))
        conn.commit()
    cursor.execute("UPDATE user_passwords SET last_login = ? WHERE username = ?", (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), username))
    conn.commit()
    conn.close()
    # Create session token
    token = create_session(username, request.remote_addr)
    return jsonify({"success": True, "is_admin": username == DB_ADMIN_USER,
                    "session_token": token, "username": username})

@app.route('/api/check-admin', methods=['GET'])
def check_admin():
    username = get_current_user()
    # Also accept legacy username param for backward compat (read-only info)
    if not username:
        username = request.args.get('username', '')
        if username not in ALLOWED_USERS:
            return jsonify({"is_admin": False, "authenticated": False})
    return jsonify({"is_admin": username == DB_ADMIN_USER, "username": username, "authenticated": True})

# ==================== STATS API ====================
@app.route('/api/debug/tables', methods=['GET'])
def debug_tables():
    """Debug endpoint to check table structure - admin only"""
    username = get_current_user()
    if not username or username != DB_ADMIN_USER:
        return jsonify({'error': 'دسترسی فقط برای مدیر سیستم'}), 403
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        tables = {}
        
        # Check each table
        for table in ['apn_ips', 'apn_mali', 'tunnel200_ips', 'tunnel_mali', 'lan_ips', 'intranet_tunnels']:
            try:
                cursor.execute(f"SELECT COUNT(*) as count FROM {table}")
                count = cursor.fetchone()[0]
                
                cursor.execute(f"PRAGMA table_info({table})")
                columns = [row[1] for row in cursor.fetchall()]
                
                tables[table] = {
                    'exists': True,
                    'count': count,
                    'columns': columns
                }
            except Exception as e:
                tables[table] = {
                    'exists': False,
                    'error': str(e)
                }
        
        conn.close()
        return jsonify(tables)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== SERVICE MANAGEMENT ====================
@app.route('/service-management')
def service_management():
    return render_template('service_management.html')

# ==================== NEW PAGES ====================
@app.route('/shared-files')
def shared_files_page():
    return render_template('shared_files.html')

@app.route('/reports')
def reports_page():
    return render_template('reports.html')

@app.route('/network-map')
def network_map_page():
    return render_template('network_map.html')

@app.route('/nat-diagram')
def nat_diagram_page():
    return render_template('nat_diagram.html')

@app.route('/fiber')
def fiber_page():
    return render_template('fiber.html')

# ==================== FINGLISH TO PERSIAN TRANSLATION ====================
FINGLISH_DICT = {
    # Common branch/place name words
    'imam': 'امام', 'emam': 'امام', 'Imam': 'امام',
    'beheshti': 'بهشتی', 'Beheshti': 'بهشتی',
    'motahari': 'مطهری', 'Motahari': 'مطهری',
    'modares': 'مدرس', 'Modares': 'مدرس',
    'shariati': 'شریعتی', 'Shariati': 'شریعتی',
    'azadi': 'آزادی', 'Azadi': 'آزادی',
    'enghelab': 'انقلاب', 'Enghelab': 'انقلاب',
    'taleghani': 'طالقانی', 'Taleghani': 'طالقانی', 'taleqani': 'طالقانی',
    'fatemi': 'فاطمی', 'Fatemi': 'فاطمی',
    'keshavarz': 'کشاورز', 'Keshavarz': 'کشاورز',
    'valiasr': 'ولیعصر', 'ValiAsr': 'ولیعصر',
    'jomhuri': 'جمهوری', 'Jomhuri': 'جمهوری', 'jomhori': 'جمهوری',
    'jahad': 'جهاد', 'Jahad': 'جهاد',
    'shohada': 'شهدا', 'Shohada': 'شهدا',
    'bahonar': 'باهنر', 'Bahonar': 'باهنر',
    'rajaei': 'رجایی', 'Rajaei': 'رجایی',
    'saduqi': 'صدوقی', 'Saduqi': 'صدوقی',
    'abuzar': 'ابوذر', 'Abuzar': 'ابوذر',
    'salman': 'سلمان', 'SalmanFarsi': 'سلمان فارسی',
    'golestan': 'گلستان', 'Golestan': 'گلستان',
    'hafez': 'حافظ', 'Hafez': 'حافظ',
    'ferdosi': 'فردوسی', 'Ferdosi': 'فردوسی', 'ferdowsi': 'فردوسی',
    'isargaran': 'ایثارگران', 'Isargaran': 'ایثارگران',
    'mohajerin': 'مهاجرین', 'Mohajerin': 'مهاجرین',
    'AmmarYaser': 'عمار یاسر',
    'SahebolAmr': 'صاحب الامر',
    'TareBar': 'تره‌بار', 'tarebar': 'تره‌بار',
    'taavon': 'تعاون', 'Taavon': 'تعاون',
    'beladiyan': 'بلدیه', 'Beladiyan': 'بلدیه',
    'bakeri': 'باکری', 'Bakeri': 'باکری',
    'rahnamaiy': 'راهنمایی', 'Rahnamaiy': 'راهنمایی',
    'mahdiyeh': 'مهدیه', 'Mahdiyeh': 'مهدیه',
    # Cities
    'tehran': 'تهران', 'Tehran': 'تهران', 'TEH': 'تهران',
    'tabriz': 'تبریز', 'Tabriz': 'تبریز',
    'mashhad': 'مشهد', 'Mashhad': 'مشهد',
    'isfahan': 'اصفهان', 'Isfahan': 'اصفهان', 'Isf': 'اصفهان',
    'shiraz': 'شیراز', 'Shiraz': 'شیراز',
    'rasht': 'رشت', 'Rasht': 'رشت',
    'gorgan': 'گرگان', 'Gorgan': 'گرگان',
    'semnan': 'سمنان', 'Semnan': 'سمنان', 'Smn': 'سمنان',
    'yazd': 'یزد', 'Yazd': 'یزد',
    'yasouj': 'یاسوج', 'Yasouj': 'یاسوج',
    'zanjan': 'زنجان', 'Zanjan': 'زنجان',
    'amol': 'آمل', 'Amol': 'آمل', 'Aml': 'آمل',
    'qom': 'قم', 'Qom': 'قم',
    'anzali': 'انزلی', 'Anzali': 'انزلی',
    'qeshm': 'قشم', 'Qeshm': 'قشم',
    'lengeh': 'لنگه', 'Lengeh': 'لنگه',
    'qaen': 'قائن', 'Qaen': 'قائن',
    'qorveh': 'قروه', 'Qorveh': 'قروه',
    'saqez': 'سقز', 'Saqez': 'سقز',
    'sonqor': 'سنقر', 'Sonqor': 'سنقر',
    'abhar': 'ابهر', 'Abhar': 'ابهر',
    'abyek': 'آبیک', 'Abyek': 'آبیک',
    'alvand': 'الوند', 'Alvand': 'الوند',
    'alborz': 'البرز', 'Alborz': 'البرز',
    'ahar': 'اهر', 'Ahar': 'اهر',
    'likak': 'لیکک', 'Likak': 'لیکک',
    'dayyer': 'دیّر', 'Dayyer': 'دیّر',
    'beyza': 'بیضا', 'Beyza': 'بیضا',
    'dehaj': 'دهج', 'Dehaj': 'دهج',
    'kahnuj': 'کهنوج', 'Kahnuj': 'کهنوج',
    'dorud': 'دورود', 'Dorud': 'دورود',
    'esfarayin': 'اسفراین', 'Esfrn': 'اسفراین',
    'gonabad': 'گناباد', 'Gonabad': 'گناباد',
    'garmsar': 'گرمسار', 'Grmsr': 'گرمسار',
    'damghan': 'دامغان', 'Dmqn': 'دامغان',
    'orumiyeh': 'ارومیه', 'Ormiye': 'ارومیه',
    'oshnaviye': 'اشنویه', 'Oshnaviye': 'اشنویه',
    # Types
    'ATM': 'خودپرداز', 'atm': 'خودپرداز',
    'kiosk': 'کیوسک', 'Kiosk': 'کیوسک',
    'BJ': 'باجه', 'Bj': 'باجه',
    'JKRM': 'جهاد کرمان',
    'blv': 'بلوار', 'Blv': 'بلوار', 'bolvar': 'بلوار',
    'hsptl': 'بیمارستان', 'Hsptl': 'بیمارستان',
    'psg': 'پاساژ', 'Psg': 'پاساژ',
    'clinic': 'کلینیک', 'Clinic': 'کلینیک',
    # Common words
    'manabe': 'منابع', 'Manabe': 'منابع', 'MTabiei': 'منابع طبیعی', 'ManabeTabiei': 'منابع طبیعی',
    'dampezeshki': 'دامپزشکی', 'Dampezeshki': 'دامپزشکی',
    'shahrivar': 'شهریور', 'Shahrivar': 'شهریور',
    'bahman': 'بهمن', 'Bahman': 'بهمن',
    'khordad': 'خرداد', 'Khordad': 'خرداد',
    'reza': 'رضا', 'Reza': 'رضا',
    'hossein': 'حسین', 'Hossein': 'حسین',
    'ali': 'علی', 'Ali': 'علی',
    'amuzesh': 'آموزش', 'Amuzesh': 'آموزش',
    'ArioBarzan': 'آریوبرزن', 'Dehdasht': 'دهدشت',
    'NakhlTaqi': 'نخل تقی',
    'DowlatAbad': 'دولت‌آباد',
    'IslamAbad': 'اسلام‌آباد', 'IslamAbd': 'اسلام‌آباد',
    'AzadShahr': 'آزادشهر',
    'HezarJarib': 'هزار جریب',
    'SiahKal': 'سیاهکل',
    'NoBahar': 'نوبهار',
    'TaqBostan': 'طاق بستان',
    'QaraZiyaDin': 'قره‌ضیاءالدین',
    'PolDasht': 'پلدشت',
    'GolTape': 'گل تپه',
    'AbGarm': 'آبگرم',
    'ImamReza': 'امام رضا', 'ImamRez': 'امام رضا',
    'ImamHossein': 'امام حسین',
    'ImamAli': 'امام علی',
    'AzarShahr': 'آذرشهر',
    'EynOlQozat': 'عین القضات',
    'QabusGonbad': 'قابوس گنبد',
    'KhajehNasir': 'خواجه نصیر',
    'BaqeSafa': 'باغ صفا',
    'RahAhan': 'راه‌آهن',
    # VPLS-specific names
    'JebalBarez': 'جبال بارز', 'Family': 'فمیلی', 'Market': 'مارکت',
    'Mahmoudi': 'محمودی', 'Mobile': 'موبایل', 'Lazemi': 'لازمی',
    'TalasaziAzimi': 'طلاسازی عظیمی',
    # Intranet description prefixes
    'MO': 'مرکز استان', 'Bazar': 'بازار', 'Baneh': 'بانه',
    'Bijar': 'بیجار', 'Dezaj': 'دزج', 'Vinsar': 'وینسار',
    'Sanandaj': 'سنندج', 'Shabestar': 'شبستر',
    'HosseinAbad': 'حسین‌آباد', 'HasanAbad': 'حسن‌آباد',
    'ShahrakSan': 'شهرک صنعتی', 'BabaRashani': 'بابارشانی',
    'DehGolan': 'دهگلان', 'Pataveh': 'پاتاوه', 'Pishin': 'پیشین',
    'EnqelabSnj': 'انقلاب سنندج', 'BazarBjr': 'بازار بیجار',
    # ── Additional cities ─────────────────────────────────────────────────────
    'khorramabad': 'خرم‌آباد', 'Khorramabad': 'خرم‌آباد', 'Khrbd': 'خرم‌آباد',
    'ilam': 'ایلام', 'Ilam': 'ایلام',
    'zahedan': 'زاهدان', 'Zahedan': 'زاهدان', 'ZHD': 'زاهدان',
    'birjand': 'بیرجند', 'Birjand': 'بیرجند',
    'bandarabbas': 'بندرعباس', 'BandarAbbas': 'بندرعباس', 'BND': 'بندرعباس',
    'bandar': 'بندر', 'Bandar': 'بندر',
    'bushehr': 'بوشهر', 'Bushehr': 'بوشهر', 'BSH': 'بوشهر',
    'shahrekord': 'شهرکرد', 'Shahrekord': 'شهرکرد', 'Shkrd': 'شهرکرد',
    'hamadan': 'همدان', 'Hamadan': 'همدان', 'HMD': 'همدان',
    'kermanshah': 'کرمانشاه', 'Kermanshah': 'کرمانشاه', 'KRN': 'کرمانشاه',
    'arak': 'اراک', 'Arak': 'اراک',
    'qazvin': 'قزوین', 'Qazvin': 'قزوین', 'Qzv': 'قزوین',
    'ardabil': 'اردبیل', 'Ardabil': 'اردبیل', 'ARD': 'اردبیل',
    'karaj': 'کرج', 'Karaj': 'کرج', 'KRJ': 'کرج',
    'saveh': 'ساوه', 'Saveh': 'ساوه',
    'boroujerd': 'بروجرد', 'Boroujerd': 'بروجرد',
    'kuhdasht': 'کوهدشت', 'Kuhdasht': 'کوهدشت',
    'azna': 'ازنا', 'Azna': 'ازنا',
    'aligoudarz': 'الیگودرز', 'Aligoudarz': 'الیگودرز',
    'poldokhtar': 'پل‌دختر', 'PolDokhtar': 'پل‌دختر',
    'mahabad': 'مهاباد', 'Mahabad': 'مهاباد',
    'bukan': 'بوکان', 'Bukan': 'بوکان',
    'marivan': 'مریوان', 'Marivan': 'مریوان',
    'kamyaran': 'کامیاران', 'Kamyaran': 'کامیاران',
    'divandarreh': 'دیواندره', 'Divandarreh': 'دیواندره',
    'paveh': 'پاوه', 'Paveh': 'پاوه',
    'kangavar': 'کنگاور', 'Kangavar': 'کنگاور',
    'sahneh': 'صحنه', 'Sahneh': 'صحنه',
    'harsin': 'هرسین', 'Harsin': 'هرسین',
    'nehavand': 'نهاوند', 'Nehavand': 'نهاوند',
    'malayer': 'ملایر', 'Malayer': 'ملایر',
    'asadabad': 'اسدآباد', 'Asadabad': 'اسدآباد',
    'tuyserkan': 'تویسرکان', 'Tuyserkan': 'تویسرکان',
    'bahar': 'بهار', 'Bahar': 'بهار',
    'razan': 'رزن', 'Razan': 'رزن',
    'kashan': 'کاشان', 'Kashan': 'کاشان', 'KSH': 'کاشان',
    'najafabad': 'نجف‌آباد', 'Najafabad': 'نجف‌آباد',
    'mobarakeh': 'مبارکه', 'Mobarakeh': 'مبارکه',
    'shahreza': 'شهرضا', 'Shahreza': 'شهرضا',
    'golpayegan': 'گلپایگان', 'Golpayegan': 'گلپایگان',
    'khomeynishahr': 'خمینی‌شهر', 'KhomeyniShahr': 'خمینی‌شهر',
    'kazerun': 'کازرون', 'Kazerun': 'کازرون',
    'fasa': 'فسا', 'Fasa': 'فسا',
    'neyriz': 'نی‌ریز', 'Neyriz': 'نی‌ریز',
    'larestan': 'لارستان', 'Larestan': 'لارستان',
    'jahrom': 'جهرم', 'Jahrom': 'جهرم',
    'firoozabad': 'فیروزآباد', 'Firoozabad': 'فیروزآباد',
    'gachsaran': 'گچساران', 'Gachsaran': 'گچساران',
    'maragheh': 'مراغه', 'Maragheh': 'مراغه',
    'mianeh': 'میانه', 'Mianeh': 'میانه',
    'marand': 'مرند', 'Marand': 'مرند',
    'khoy': 'خوی', 'Khoy': 'خوی',
    'salmas': 'سلماس', 'Salmas': 'سلماس',
    'piranshahr': 'پیران‌شهر', 'Piranshahr': 'پیران‌شهر',
    'sardasht': 'سردشت', 'Sardasht': 'سردشت',
    'naghadeh': 'نقده', 'Naghadeh': 'نقده',
    'miandoab': 'میاندوآب', 'Miandoab': 'میاندوآب',
    'tekab': 'تکاب', 'Tekab': 'تکاب',
    'jolfa': 'جلفا', 'Jolfa': 'جلفا',
    'meshginshahr': 'مشگین‌شهر', 'MeshginShahr': 'مشگین‌شهر',
    'parsabad': 'پارس‌آباد', 'Parsabad': 'پارس‌آباد',
    'germi': 'گرمی', 'Germi': 'گرمی',
    'khalkhal': 'خلخال', 'Khalkhal': 'خلخال',
    'takestan': 'تاکستان', 'Takestan': 'تاکستان',
    'khodabandeh': 'خدابنده', 'Khodabandeh': 'خدابنده',
    'boroujen': 'بروجن', 'Boroujen': 'بروجن',
    'natanz': 'نطنز', 'Natanz': 'نطنز',
    'ardestan': 'اردستان', 'Ardestan': 'اردستان',
    'zarinshahr': 'زرین‌شهر', 'ZarinShahr': 'زرین‌شهر',
    'lenjan': 'لنجان', 'Lenjan': 'لنجان',
    'fereydunshahr': 'فریدون‌شهر', 'FereydunShahr': 'فریدون‌شهر',
    'khansari': 'خوانسار', 'Khansari': 'خوانسار',
    'khomein': 'خمین', 'Khomein': 'خمین',
    'delijan': 'دلیجان', 'Delijan': 'دلیجان',
    'mahalat': 'محلات', 'Mahalat': 'محلات',
    'zarand': 'زرند', 'Zarand': 'زرند',
    'bam': 'بم', 'Bam': 'بم',
    'jiroft': 'جیرفت', 'Jiroft': 'جیرفت',
    'sirjan': 'سیرجان', 'Sirjan': 'سیرجان',
    'rafsanjan': 'رفسنجان', 'Rafsanjan': 'رفسنجان',
    'kerman': 'کرمان', 'Kerman': 'کرمان', 'KRM': 'کرمان',
    'ahvaz': 'اهواز', 'Ahvaz': 'اهواز', 'AHV': 'اهواز', 'AHZ': 'اهواز',
    'sabzevar': 'سبزوار', 'Sabzevar': 'سبزوار',
    'neyshabur': 'نیشابور', 'Neyshabur': 'نیشابور',
    'torbatheydariyeh': 'تربت‌حیدریه', 'TorbatHeydariyeh': 'تربت‌حیدریه',
    'kashmar': 'کاشمر', 'Kashmar': 'کاشمر',
    'quchan': 'قوچان', 'Quchan': 'قوچان',
    'bojnurd': 'بجنورد', 'Bojnurd': 'بجنورد',
    'shirvan': 'شیروان', 'Shirvan': 'شیروان',
    'zabol': 'زابل', 'Zabol': 'زابل',
    'iranshahr': 'ایرانشهر', 'Iranshahr': 'ایرانشهر',
    'chabahar': 'چابهار', 'Chabahar': 'چابهار',
    'khash': 'خاش', 'Khash': 'خاش',
    'saravan': 'سراوان', 'Saravan': 'سراوان',
    'dezful': 'دزفول', 'Dezful': 'دزفول',
    'andimeshk': 'اندیمشک', 'Andimeshk': 'اندیمشک',
    'shushtar': 'شوشتر', 'Shushtar': 'شوشتر',
    'behbahan': 'بهبهان', 'Behbahan': 'بهبهان',
    'omidiyeh': 'امیدیه', 'Omidiyeh': 'امیدیه',
    'izeh': 'ایذه', 'Izeh': 'ایذه',
    'abadan': 'آبادان', 'Abadan': 'آبادان',
    'khoramshahr': 'خرمشهر', 'Khoramshahr': 'خرمشهر',
    'masjedsoleyman': 'مسجدسلیمان', 'MasjedSoleyman': 'مسجدسلیمان',
    'masjed': 'مسجد', 'Masjed': 'مسجد',
    'ramhormoz': 'رامهرمز', 'Ramhormoz': 'رامهرمز',
    'shadegan': 'شادگان', 'Shadegan': 'شادگان',
    # ── Banking / financial terms ─────────────────────────────────────────────
    'markazi': 'مرکزی', 'Markazi': 'مرکزی',
    'meli': 'ملی', 'Meli': 'ملی',
    'mellat': 'ملت', 'Mellat': 'ملت',
    'tejarat': 'تجارت', 'Tejarat': 'تجارت',
    'saderat': 'صادرات', 'Saderat': 'صادرات',
    'sepah': 'سپه', 'Sepah': 'سپه',
    'keshavarzi': 'کشاورزی', 'Keshavarzi': 'کشاورزی',
    'maskan': 'مسکن', 'Maskan': 'مسکن',
    'refah': 'رفاه', 'Refah': 'رفاه',
    'pasargad': 'پاسارگاد', 'Pasargad': 'پاسارگاد',
    'parsian': 'پارسیان', 'Parsian': 'پارسیان',
    'saman': 'سامان', 'Saman': 'سامان',
    'sina': 'سینا', 'Sina': 'سینا',
    'ayandeh': 'آینده', 'Ayandeh': 'آینده',
    'karafarin': 'کارآفرین', 'Karafarin': 'کارآفرین',
    'eghtesad': 'اقتصاد', 'Eghtesad': 'اقتصاد',
    'ansar': 'انصار', 'Ansar': 'انصار',
    'bazargani': 'بازرگانی', 'Bazargani': 'بازرگانی',
    'sherkat': 'شرکت', 'Sherkat': 'شرکت',
    # ── Place structure words ─────────────────────────────────────────────────
    'meydan': 'میدان', 'Meydan': 'میدان',
    'kheyaban': 'خیابان', 'Kheyaban': 'خیابان',
    'shahrak': 'شهرک', 'Shahrak': 'شهرک',
    'sanati': 'صنعتی', 'Sanati': 'صنعتی',
    'tejari': 'تجاری', 'Tejari': 'تجاری',
    'markaz': 'مرکز', 'Markaz': 'مرکز',
    'shamal': 'شمال', 'Shamal': 'شمال',
    'jonoob': 'جنوب', 'Jonoob': 'جنوب',
    'shargh': 'شرق', 'Shargh': 'شرق',
    'gharb': 'غرب', 'Gharb': 'غرب',
    'kooy': 'کوی', 'Kooy': 'کوی',
    'nabsh': 'نبش', 'Nabsh': 'نبش',
    'bagh': 'باغ', 'Bagh': 'باغ',
    'abad': 'آباد', 'Abad': 'آباد',
    'shahr': 'شهر', 'Shahr': 'شهر',
    'roosta': 'روستا', 'Roosta': 'روستا',
    'shahrestan': 'شهرستان', 'Shahrestan': 'شهرستان',
    'SP': 'سرپرستی', 'Sp': 'سرپرستی',
    # ── Common personal names / titles ───────────────────────────────────────
    'mahdi': 'مهدی', 'Mahdi': 'مهدی',
    'javad': 'جواد', 'Javad': 'جواد',
    'mousa': 'موسی', 'Mousa': 'موسی',
    'karim': 'کریم', 'Karim': 'کریم',
    'hakim': 'حکیم', 'Hakim': 'حکیم',
    'moradi': 'مرادی', 'Moradi': 'مرادی',
    'hosseini': 'حسینی', 'Hosseini': 'حسینی',
    'ahmadi': 'احمدی', 'Ahmadi': 'احمدی',
    'rezaei': 'رضایی', 'Rezaei': 'رضایی',
    'mohammadi': 'محمدی', 'Mohammadi': 'محمدی',
    'akbar': 'اکبر', 'Akbar': 'اکبر',
    'asghar': 'اصغر', 'Asghar': 'اصغر',
    'ahmad': 'احمد', 'Ahmad': 'احمد',
    'hassan': 'حسن', 'Hassan': 'حسن',
    'mostafa': 'مصطفی', 'Mostafa': 'مصطفی',
    'sadegh': 'صادق', 'Sadegh': 'صادق',
    'zeinab': 'زینب', 'Zeinab': 'زینب',
    'fatemeh': 'فاطمه', 'Fatemeh': 'فاطمه',
    'karbala': 'کربلا', 'Karbala': 'کربلا',
    'najaf': 'نجف', 'Najaf': 'نجف',
    'nouri': 'نوری', 'Nouri': 'نوری',
    'kashani': 'کاشانی', 'Kashani': 'کاشانی',
    'chamran': 'چمران', 'Chamran': 'چمران',
    'resalat': 'رسالت', 'Resalat': 'رسالت',
    'sattari': 'ستاری', 'Sattari': 'ستاری',
    'shahid': 'شهید', 'Shahid': 'شهید',
    'doktor': 'دکتر', 'Dr': 'دکتر',
    'mohandas': 'مهندس', 'Mohandas': 'مهندس',
    'ostade': 'استاد', 'Ostade': 'استاد',
}

# Load custom translations from DB at startup
# NOTE: Called BEFORE PERSIAN_TO_FINGLISH is built, so we only update
# FINGLISH_DICT here. The reverse mapping is built right after this call
# from the fully-populated FINGLISH_DICT (which then includes custom entries).
def _load_custom_translations():
    loaded = 0
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT name_en, name_fa FROM custom_translations")
        for row in cursor.fetchall():
            FINGLISH_DICT[row[0]] = row[1]
            loaded += 1
        conn.close()
        if loaded:
            print(f"Loaded {loaded} custom translations from DB")
    except Exception as e:
        print(f"Warning: Failed to load custom translations: {e}")

_load_custom_translations()
# fuzzy-cache is built lazily on first call (after FINGLISH_DICT is final)

# Build reverse dictionary for Persian-to-Finglish search
PERSIAN_TO_FINGLISH = {}
for en, fa in FINGLISH_DICT.items():
    if fa not in PERSIAN_TO_FINGLISH:
        PERSIAN_TO_FINGLISH[fa] = []
    PERSIAN_TO_FINGLISH[fa].append(en)

import difflib as _difflib

# Thread lock for all FINGLISH_DICT / PERSIAN_TO_FINGLISH mutations
_translation_lock = threading.RLock()

# Flat list of all keys for fuzzy matching — rebuilt when dict changes
_FINGLISH_KEYS_CACHE = []
_FINGLISH_KEYS_LOWER_MAP = {}   # {lowercase_key: original_key} — cached permanently

def _rebuild_fuzzy_cache():
    """Rebuild the key lists used by difflib matching. Call under _translation_lock."""
    global _FINGLISH_KEYS_CACHE, _FINGLISH_KEYS_LOWER_MAP
    _FINGLISH_KEYS_CACHE = list(FINGLISH_DICT.keys())
    _FINGLISH_KEYS_LOWER_MAP = {k.lower(): k for k in _FINGLISH_KEYS_CACHE}

def _fuzzy_lookup(token, cutoff=0.82):
    """Find the closest FINGLISH_DICT key for *token* using difflib.

    Returns the Persian value for the best match, or None if no match
    is above *cutoff* similarity.  Higher cutoff = stricter matching.
    Strategy:
      1. Try exact-case match first (already done by caller).
      2. Try close matches against the full key list (case-sensitive).
      3. Retry case-insensitively using the pre-built lowercase key map.
    A minimum token length of 4 is enforced to avoid false positives on
    short abbreviations.
    """
    if len(token) < 4:
        return None
    if not _FINGLISH_KEYS_CACHE:
        _rebuild_fuzzy_cache()

    # Case-sensitive pass
    matches = _difflib.get_close_matches(token, _FINGLISH_KEYS_CACHE,
                                          n=1, cutoff=cutoff)
    if matches:
        return FINGLISH_DICT[matches[0]]

    # Case-insensitive pass — use pre-built (cached) lowercase map
    lower_keys = list(_FINGLISH_KEYS_LOWER_MAP.keys())
    ci_matches = _difflib.get_close_matches(token.lower(), lower_keys,
                                             n=1, cutoff=cutoff)
    if ci_matches:
        original_key = _FINGLISH_KEYS_LOWER_MAP[ci_matches[0]]
        return FINGLISH_DICT.get(original_key)

    return None


def translate_finglish(name):
    """Translate a Finglish/English branch name to Persian.

    Strategy (fastest-first):
      1. Direct dict lookup (exact).
      2. Persian-chars guard — already Persian, return as-is.
      3. Strip bandwidth/numeric suffix then direct lookup.
      4. Split on hyphens / underscores / spaces → word-by-word dict lookup.
         • Also tries combining adjacent tokens (CamelCase combos).
         • Variants: exact / lowercase / capitalize / UPPER.
      5. Fuzzy matching via difflib for any token that still didn't match
         (only for tokens ≥ 4 chars and similarity ≥ 0.78).

    Returns:
      - Persian translation string if at least one token was translated.
      - Empty string if nothing could be translated.
    """
    if not name:
        return ''

    import re as _re

    # ── 1. Direct exact match ────────────────────────────────────────────────
    if name in FINGLISH_DICT:
        return FINGLISH_DICT[name]

    # ── 2. Already Persian? ──────────────────────────────────────────────────
    if _re.search(r'[\u0600-\u06FF]', name):
        return name

    # ── 3. Strip numeric/bandwidth suffix ───────────────────────────────────
    clean = _re.sub(r'[-_ ]?\d+[KkMm]?(bps)?$', '', name).strip()
    if clean and clean != name and clean in FINGLISH_DICT:
        return FINGLISH_DICT[clean]

    # ── 4 & 5. Token-by-token + fuzzy ───────────────────────────────────────
    parts = _re.split(r'[-_ ]', clean or name)
    if len(parts) == 1:
        # Single token: try CamelCase split  "ImamReza" → ["Imam","Reza"]
        parts = _re.findall(r'[A-Z][a-z]*|[a-z]{2,}|[A-Z]{2,}', clean or name)
    parts = [p for p in parts if p]
    if not parts:
        return ''

    translated = []
    any_translated = False
    i = 0
    while i < len(parts):
        p = parts[i]
        matched = False

        # ── 4a. Try combining current + next (e.g. "Imam"+"Reza"="ImamReza") ─
        if i + 1 < len(parts):
            for combo in (p + parts[i + 1],
                          p.capitalize() + parts[i + 1].capitalize()):
                if combo in FINGLISH_DICT:
                    translated.append(FINGLISH_DICT[combo])
                    any_translated = True
                    i += 2
                    matched = True
                    break
            if matched:
                continue

        # ── 4b. Single token — exact / lowercase / capitalize / UPPER ────────
        for variant in (p, p.lower(), p.capitalize(), p.upper()):
            if variant in FINGLISH_DICT:
                translated.append(FINGLISH_DICT[variant])
                any_translated = True
                matched = True
                break

        # ── 5. Fuzzy fallback (difflib) ───────────────────────────────────────
        if not matched:
            fuzzy_result = _fuzzy_lookup(p)
            if fuzzy_result:
                translated.append(fuzzy_result)
                any_translated = True
                matched = True

        if not matched:
            translated.append(p)   # keep original token unchanged
        i += 1

    if not any_translated:
        return name   # Return original if no token could be translated
    return ' '.join(translated)

def get_persian_search_variants(query):
    """Get Finglish variants of a Persian search query for bidirectional search."""
    variants = set()
    q = query.strip()
    if q in PERSIAN_TO_FINGLISH:
        variants.update(PERSIAN_TO_FINGLISH[q])
    # Also check substrings
    for fa, en_list in PERSIAN_TO_FINGLISH.items():
        if fa in q or q in fa:
            variants.update(en_list)
    return list(variants)


@app.route('/api/search-services', methods=['GET'])
def search_services():
    """Search all services for a branch/IP across ALL tables including lan_ips"""
    try:
        query = request.args.get('q', '').strip()
        search_type = request.args.get('type', 'branch_name')

        if not query or len(query) < 2:
            return jsonify([])

        conn = get_db()
        cursor = conn.cursor()
        results = []

        like_q = f'%{query}%'

        def add_result(row_tuple, table, service, branch_name_fa=''):
            name = row_tuple[1] or ''
            # Try to translate English names for ALL tables
            fa_name = branch_name_fa or ''
            if not fa_name and name:
                import re as _re2
                # Check if name has English characters (not already Persian)
                if _re2.search(r'[A-Za-z]', name):
                    fa_name = translate_finglish(name)
            results.append({
                'id': row_tuple[0],
                'table': table,
                'service': service,
                'branch_name': name,
                'branch_name_fa': fa_name,
                'province': row_tuple[2] or '',
                'ip': row_tuple[3] or '',
                'lan_ip': row_tuple[4] or '',
                'username': row_tuple[5] or '',
                'date': row_tuple[6] or ''
            })

        if search_type == 'branch_name':
            # 1. LAN IPs (main branch table) - search by branch_name where active/reserved
            cursor.execute("""
                SELECT id, branch_name, province,
                       '10.' || octet2 || '.' || octet3 || '.0/24' as lan_ip_full,
                       wan_ip, username, reservation_date
                FROM lan_ips
                WHERE branch_name LIKE ?
                AND branch_name IS NOT NULL AND branch_name != ''
                AND status != 'Free'
            """, (like_q,))
            for r in cursor.fetchall():
                add_result(r, 'lan_ips', 'IP LAN')

            # 2. APN Mali - search by branch_name (any record with branch_name set)
            cursor.execute("""
                SELECT id, branch_name, province, ip_wan, lan_ip, username, reservation_date
                FROM apn_mali WHERE branch_name LIKE ?
                AND branch_name IS NOT NULL AND branch_name != ''
            """, (like_q,))
            for r in cursor.fetchall():
                add_result(r, 'apn_mali', 'APN مالی')

            # 3. APN Int
            cursor.execute("""
                SELECT id, branch_name, province, ip_wan_apn, lan_ip, username, reservation_date
                FROM apn_ips WHERE branch_name LIKE ?
                AND branch_name IS NOT NULL AND branch_name != ''
            """, (like_q,))
            for r in cursor.fetchall():
                add_result(r, 'apn_ips', 'APN غیرمالی')

            # 4. Intranet tunnels
            cursor.execute("""
                SELECT id, COALESCE(branch_name, description, tunnel_name), province, ip_address, ip_lan, reserved_by, reserved_at
                FROM intranet_tunnels
                WHERE (branch_name LIKE ? OR tunnel_name LIKE ? OR description LIKE ?)
                AND LOWER(status) = 'reserved'
            """, (like_q, like_q, like_q))
            for r in cursor.fetchall():
                add_result(r, 'intranet_tunnels', 'Intranet')

            # 5. VPLS/MPLS tunnels
            cursor.execute("""
                SELECT id, COALESCE(branch_name, description, tunnel_name), province, ip_address, wan_ip, username, reservation_date
                FROM vpls_tunnels
                WHERE (branch_name LIKE ? OR description LIKE ? OR tunnel_name LIKE ?)
                AND LOWER(status) = 'reserved'
            """, (like_q, like_q, like_q))
            for r in cursor.fetchall():
                add_result(r, 'vpls_tunnels', 'MPLS/VPLS')

            # 6. Tunnel Mali
            cursor.execute("""
                SELECT id, branch_name, '', ip_address, '', username, reservation_date
                FROM tunnel_mali WHERE branch_name LIKE ?
                AND branch_name IS NOT NULL AND branch_name != ''
            """, (like_q,))
            for r in cursor.fetchall():
                add_result(r, 'tunnel_mali', 'Tunnel مالی')

            # 7. Tunnel200
            cursor.execute("""
                SELECT id, branch_name, '', ip_address, '', username, reservation_date
                FROM tunnel200_ips WHERE branch_name LIKE ?
                AND branch_name IS NOT NULL AND branch_name != ''
            """, (like_q,))
            for r in cursor.fetchall():
                add_result(r, 'tunnel200_ips', 'Tunnel200')

            # 8. PTMP Serial connections (search both Persian and English names + bidirectional)
            try:
                # Build search conditions - include Finglish variants of Persian query
                search_params = [like_q, like_q]
                extra_conditions = ""
                persian_variants = get_persian_search_variants(query)
                if persian_variants:
                    for v in persian_variants[:10]:  # Limit to 10 variants
                        extra_conditions += " OR branch_name_en LIKE ?"
                        search_params.append(f'%{v}%')

                cursor.execute(f"""
                    SELECT id, COALESCE(branch_name, branch_name_en), province,
                           interface_name, lan_ip, username, reservation_date
                    FROM ptmp_connections
                    WHERE (branch_name LIKE ? OR branch_name_en LIKE ? {extra_conditions})
                    AND (branch_name IS NOT NULL OR branch_name_en IS NOT NULL)
                """, search_params)
                for r in cursor.fetchall():
                    add_result(r, 'ptmp_connections', 'PTMP سریال')
            except Exception as e:
                print(f"PTMP search error: {e}")

        elif search_type == 'ip_apn_mali':
            cursor.execute("""
                SELECT id, branch_name, province, ip_wan, lan_ip, username, reservation_date
                FROM apn_mali WHERE ip_wan LIKE ?
                AND branch_name IS NOT NULL AND branch_name != ''
            """, (like_q,))
            for r in cursor.fetchall():
                add_result(r, 'apn_mali', 'APN مالی')

        elif search_type == 'ip_apn_int':
            cursor.execute("""
                SELECT id, branch_name, province, ip_wan_apn, lan_ip, username, reservation_date
                FROM apn_ips WHERE ip_wan_apn LIKE ?
                AND branch_name IS NOT NULL AND branch_name != ''
            """, (like_q,))
            for r in cursor.fetchall():
                add_result(r, 'apn_ips', 'APN غیرمالی')

        elif search_type == 'ip_lan':
            # Search by LAN IP in lan_ips table and all service tables
            cursor.execute("""
                SELECT id, branch_name, province,
                       '10.' || octet2 || '.' || octet3 || '.0/24' as lan_ip_full,
                       wan_ip, username, reservation_date
                FROM lan_ips
                WHERE ('10.' || octet2 || '.' || octet3 || '.0/24') LIKE ?
                AND branch_name IS NOT NULL AND branch_name != ''
                AND status != 'Free'
            """, (like_q,))
            for r in cursor.fetchall():
                add_result(r, 'lan_ips', 'IP LAN')

            cursor.execute("""
                SELECT id, branch_name, province, ip_wan, lan_ip, username, reservation_date
                FROM apn_mali WHERE lan_ip LIKE ?
                AND branch_name IS NOT NULL AND branch_name != ''
            """, (like_q,))
            for r in cursor.fetchall():
                add_result(r, 'apn_mali', 'APN مالی')

            cursor.execute("""
                SELECT id, branch_name, province, ip_wan_apn, lan_ip, username, reservation_date
                FROM apn_ips WHERE lan_ip LIKE ?
                AND branch_name IS NOT NULL AND branch_name != ''
            """, (like_q,))
            for r in cursor.fetchall():
                add_result(r, 'apn_ips', 'APN غیرمالی')

            cursor.execute("""
                SELECT id, COALESCE(branch_name, description, tunnel_name), province, ip_address, ip_lan, reserved_by, reserved_at
                FROM intranet_tunnels WHERE ip_lan LIKE ?
                AND LOWER(status) = 'reserved'
            """, (like_q,))
            for r in cursor.fetchall():
                add_result(r, 'intranet_tunnels', 'Intranet')

            # VPLS/MPLS by LAN IP
            try:
                cursor.execute("""
                    SELECT id, COALESCE(branch_name, description, tunnel_name), province,
                           ip_address, lan_ip, username, reservation_date
                    FROM vpls_tunnels WHERE lan_ip LIKE ?
                    AND LOWER(status) IN ('reserved', 'used')
                """, (like_q,))
                for r in cursor.fetchall():
                    add_result(r, 'vpls_tunnels', 'MPLS/VPLS')
            except Exception:
                pass

            # PTMP by LAN IP
            try:
                cursor.execute("""
                    SELECT id, COALESCE(branch_name, branch_name_en), province,
                           interface_name, lan_ip, username, reservation_date
                    FROM ptmp_connections WHERE lan_ip LIKE ?
                    AND (branch_name IS NOT NULL OR branch_name_en IS NOT NULL)
                """, (like_q,))
                for r in cursor.fetchall():
                    add_result(r, 'ptmp_connections', 'PTMP سریال')
            except Exception:
                pass

        elif search_type == 'ip_intranet':
            cursor.execute("""
                SELECT id, COALESCE(branch_name, description, tunnel_name), province, ip_address, ip_lan, reserved_by, reserved_at
                FROM intranet_tunnels
                WHERE (ip_address LIKE ? OR ip_intranet LIKE ?)
                AND LOWER(status) = 'reserved'
            """, (like_q, like_q))
            for r in cursor.fetchall():
                add_result(r, 'intranet_tunnels', 'Intranet')

        elif search_type == 'ip_vpls':
            cursor.execute("""
                SELECT id, COALESCE(branch_name, description, tunnel_name), province, ip_address, wan_ip, username, reservation_date
                FROM vpls_tunnels
                WHERE (ip_address LIKE ? OR wan_ip LIKE ?)
                AND LOWER(status) = 'reserved'
            """, (like_q, like_q))
            for r in cursor.fetchall():
                add_result(r, 'vpls_tunnels', 'MPLS/VPLS')

        elif search_type == 'ip_ptmp':
            try:
                cursor.execute("""
                    SELECT id, COALESCE(branch_name, branch_name_en), province,
                           interface_name, lan_ip, username, reservation_date
                    FROM ptmp_connections
                    WHERE (interface_name LIKE ? OR description LIKE ?
                           OR branch_name LIKE ? OR branch_name_en LIKE ?)
                    AND (branch_name IS NOT NULL OR branch_name_en IS NOT NULL)
                """, (like_q, like_q, like_q, like_q))
                for r in cursor.fetchall():
                    add_result(r, 'ptmp_connections', 'PTMP سریال')
            except Exception:
                pass

        conn.close()
        return jsonify(results)
    except Exception as e:
        print(f"Search services error: {e}")
        return jsonify([])


@app.route('/api/delete-service', methods=['POST'])
def delete_service():
    """Delete/free a specific service from any table"""
    try:
        if is_api_rate_limited(request.remote_addr, 'delete-service'):
            return jsonify({'status': 'error', 'error': 'Too many requests. Please wait.'}), 429
        data = request.json
        table = data.get('table', '')
        record_id = data.get('id')
        username = data.get('username', '')

        if not table or not record_id:
            return jsonify({'status': 'error', 'error': 'پارامترهای ناقص'}), 400

        allowed_tables = ['lan_ips', 'apn_mali', 'apn_ips', 'intranet_tunnels', 'vpls_tunnels', 'tunnel_mali', 'tunnel200_ips', 'ptmp_connections']
        if table not in allowed_tables:
            return jsonify({'status': 'error', 'error': 'جدول نامعتبر'}), 400

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE")

        try:
            # Get info before deleting for logging
            cursor.execute(f"SELECT * FROM {table} WHERE id = ?", (record_id,))
            row = cursor.fetchone()
            if not row:
                conn.rollback()
                conn.close()
                return jsonify({'status': 'error', 'error': 'رکورد پیدا نشد'}), 404

            # Free the record (set fields to NULL/Free) based on table type
            if table == 'lan_ips':
                branch = row['branch_name'] or ''
                octet2 = row['octet2']
                octet3 = row['octet3']
                ip = f"10.{octet2}.{octet3}.0/24"
                cursor.execute("""
                    UPDATE lan_ips SET username = NULL, reservation_date = NULL,
                    branch_name = NULL, status = 'Free', notes = NULL, wan_ip = NULL
                    WHERE id = ?
                """, (record_id,))
                # Also delete from reserved_ips if exists
                cursor.execute("""
                    DELETE FROM reserved_ips WHERE octet2 = ? AND octet3 = ?
                """, (octet2, octet3))
                log_activity('warning', 'حذف IP LAN', f'{branch}: {ip}', username)

            elif table == 'apn_mali':
                branch = row['branch_name'] or ''
                ip = row['ip_wan'] or ''
                cursor.execute("""
                    UPDATE apn_mali SET username = NULL, branch_name = NULL, province = NULL,
                    type = NULL, lan_ip = NULL, reservation_date = NULL WHERE id = ?
                """, (record_id,))
                # Also free associated tunnel_mali
                if ip:
                    cursor.execute("""
                        UPDATE tunnel_mali SET status = NULL, username = NULL, branch_name = NULL,
                        reservation_date = NULL, description = NULL, destination_ip = NULL
                        WHERE destination_ip = ?
                    """, (ip,))
                log_activity('warning', 'حذف سرویس APN مالی', f'{branch}: {ip}', username)

            elif table == 'apn_ips':
                branch = row['branch_name'] or ''
                ip = row['ip_wan_apn'] or ''
                cursor.execute("""
                    UPDATE apn_ips SET username = NULL, branch_name = NULL, province = NULL,
                    type = NULL, lan_ip = NULL, reservation_date = NULL WHERE id = ?
                """, (record_id,))
                # Also free associated tunnel200
                if branch:
                    cursor.execute("""
                        UPDATE tunnel200_ips SET status = NULL, username = NULL, branch_name = NULL,
                        reservation_date = NULL, description = NULL
                        WHERE branch_name = ?
                    """, (branch,))
                log_activity('warning', 'حذف سرویس APN غیرمالی', f'{branch}: {ip}', username)

            elif table == 'intranet_tunnels':
                name = row['tunnel_name'] or ''
                ip = row['ip_address'] or ''
                cursor.execute("""
                    UPDATE intranet_tunnels SET status = 'Free', reserved_by = NULL, reserved_at = NULL,
                    tunnel_name = NULL, description = NULL, ip_lan = NULL, ip_intranet = NULL,
                    branch_name = NULL, province = NULL
                    WHERE id = ?
                """, (record_id,))
                log_activity('warning', 'حذف سرویس Intranet', f'{name}: {ip}', username)

            elif table == 'vpls_tunnels':
                branch = row['branch_name'] or ''
                ip = row['ip_address'] or ''
                cursor.execute("""
                    UPDATE vpls_tunnels SET status = 'Free', username = NULL, branch_name = NULL,
                    tunnel_name = NULL, description = NULL, wan_ip = NULL, tunnel_dest = NULL,
                    lan_ip = NULL, reservation_date = NULL
                    WHERE id = ?
                """, (record_id,))
                log_activity('warning', 'حذف سرویس MPLS/VPLS', f'{branch}: {ip}', username)

            elif table == 'tunnel_mali':
                branch = row['branch_name'] or ''
                ip = row['ip_address'] or ''
                cursor.execute("""
                    UPDATE tunnel_mali SET status = NULL, username = NULL, branch_name = NULL,
                    reservation_date = NULL, description = NULL, destination_ip = NULL
                    WHERE id = ?
                """, (record_id,))
                log_activity('warning', 'حذف سرویس Tunnel مالی', f'{branch}: {ip}', username)

            elif table == 'tunnel200_ips':
                branch = row['branch_name'] or ''
                ip = row['ip_address'] or ''
                cursor.execute("""
                    UPDATE tunnel200_ips SET status = NULL, username = NULL, branch_name = NULL,
                    reservation_date = NULL, description = NULL
                    WHERE id = ?
                """, (record_id,))
                log_activity('warning', 'حذف سرویس Tunnel200', f'{branch}: {ip}', username)

            elif table == 'ptmp_connections':
                branch = row['branch_name'] or row['branch_name_en'] or ''
                intf = row['interface_name'] or ''
                cursor.execute("DELETE FROM ptmp_connections WHERE id = ?", (record_id,))
                log_activity('warning', 'حذف سرویس PTMP', f'{branch}: {intf}', username)

            conn.commit()
        except Exception as inner_e:
            conn.rollback()
            raise inner_e
        finally:
            conn.close()

        # Clear stats cache
        _stats_cache['data'] = None
        _stats_cache['time'] = 0

        return jsonify({'status': 'ok', 'message': 'سرویس با موفقیت حذف شد'})
    except Exception as e:
        print(f"Delete service error: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/api/stats', methods=['GET'])
def get_stats():
    global _stats_cache

    # Return cached result if less than 60 seconds old
    if _stats_cache['data'] and (time.time() - _stats_cache['time']) < STATS_CACHE_SECONDS:
        return jsonify(_stats_cache['data'])
    
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # OPTIMIZED: Single query instead of many separate queries
        cursor.execute("""
            SELECT
                (SELECT COUNT(*) FROM lan_ips) as total_lan,
                (SELECT COUNT(*) FROM lan_ips WHERE (username IS NULL OR username = '') AND (branch_name IS NULL OR branch_name = '')) as free_lan,
                (SELECT COUNT(*) FROM intranet_tunnels) as total_tun,
                (SELECT COUNT(*) FROM intranet_tunnels WHERE LOWER(status) = 'free') as free_tun,
                (SELECT COUNT(*) FROM apn_ips) as total_apn,
                (SELECT COUNT(*) FROM apn_ips WHERE username IS NULL OR username = '') as free_apn,
                (SELECT COUNT(*) FROM apn_mali) as total_mali,
                (SELECT COUNT(*) FROM apn_mali WHERE username IS NULL OR username = '') as free_mali,
                (SELECT COUNT(*) FROM tunnel200_ips) as total_t200,
                (SELECT COUNT(*) FROM tunnel200_ips WHERE status IS NULL OR status = '' OR LOWER(status) = 'free') as free_t200,
                (SELECT COUNT(*) FROM tunnel_mali) as total_tmali,
                (SELECT COUNT(*) FROM tunnel_mali WHERE status IS NULL OR status = '' OR LOWER(status) = 'free') as free_tmali,
                (SELECT COUNT(*) FROM vpls_tunnels) as total_vpls,
                (SELECT COUNT(*) FROM vpls_tunnels WHERE LOWER(status) = 'free') as free_vpls,
                (SELECT COUNT(*) FROM ptmp_connections) as total_ptmp,
                (SELECT COUNT(*) FROM ptmp_connections WHERE branch_name IS NOT NULL) as matched_ptmp
        """)

        row = cursor.fetchone()
        conn.close()

        total_lan, free_lan = row[0], row[1]
        total_tun, free_tun = row[2], row[3]
        total_apn, free_apn = row[4], row[5]
        total_mali, free_mali = row[6], row[7]
        total_t200, free_t200 = row[8], row[9]
        total_tmali, free_tmali = row[10], row[11]
        total_vpls, free_vpls = row[12], row[13]
        total_ptmp, matched_ptmp = row[14], row[15]

        result = {
            'lan_ips': {'total': total_lan, 'free': free_lan, 'used': total_lan - free_lan},
            'tunnels': {'total': total_tun, 'free': free_tun, 'used': total_tun - free_tun},
            'apn': {'total': total_apn, 'free': free_apn, 'used': total_apn - free_apn},
            'apn_mali': {'total': total_mali, 'free': free_mali, 'used': total_mali - free_mali},
            'tunnel200': {'total': total_t200, 'free': free_t200, 'used': total_t200 - free_t200},
            'tunnel_mali': {'total': total_tmali, 'free': free_tmali, 'used': total_tmali - free_tmali},
            'vpls': {'total': total_vpls, 'free': free_vpls, 'used': total_vpls - free_vpls},
            'ptmp': {'total': total_ptmp, 'matched': matched_ptmp, 'used': total_ptmp}
        }
        
        # Cache the result
        _stats_cache['data'] = result
        _stats_cache['time'] = time.time()
        
        return jsonify(result)
    except Exception as e:
        print(f"❌ Stats error: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== EXPIRING RESERVATIONS ====================
@app.route('/api/expiring-reservations', methods=['GET'])
def get_expiring_reservations():
    """Get count of reservations expiring in next 7 days"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Calculate date 7 days from now
        future_date = (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')
        today = datetime.now().strftime('%Y-%m-%d')
        
        cursor.execute("""
            SELECT COUNT(*) FROM reserved_ips 
            WHERE expiry_date BETWEEN ? AND ?
            AND (status = 'reserved' OR status IS NULL)
        """, (today, future_date))
        
        count = cursor.fetchone()[0]
        conn.close()
        
        return jsonify({'count': count})
    except Exception as e:
        print(f"❌ Expiring reservations error: {e}")
        return jsonify({'count': 0})

# ==================== RECENT RESERVATIONS ====================
@app.route('/api/recent-reservations', methods=['GET'])
def get_recent_reservations():
    """Get recent IP reservations from all tables"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        reservations = []
        
        # From apn_ips (APN غیرمالی)
        try:
            cursor.execute("""
                SELECT branch_name, ip_wan_apn as ip, province, username, reservation_date, 'APN-INT' as type
                FROM apn_ips 
                WHERE username IS NOT NULL AND username != ''
                ORDER BY reservation_date DESC
                LIMIT 5
            """)
            for row in cursor.fetchall():
                reservations.append({
                    'branch_name': row['branch_name'] or '',
                    'ip': row['ip'] or '',
                    'province': row['province'] or '',
                    'username': row['username'] or '',
                    'date': row['reservation_date'] or '',
                    'type': 'APN-INT'
                })
        except Exception as e:
            print(f"⚠️ Recent reservations query: {e}")
        
        # From apn_mali (APN مالی)
        try:
            cursor.execute("""
                SELECT branch_name, ip_wan as ip, province, username, reservation_date, 'APN-MALI' as type
                FROM apn_mali 
                WHERE username IS NOT NULL AND username != ''
                ORDER BY reservation_date DESC
                LIMIT 5
            """)
            for row in cursor.fetchall():
                reservations.append({
                    'branch_name': row['branch_name'] or '',
                    'ip': row['ip'] or '',
                    'province': row['province'] or '',
                    'username': row['username'] or '',
                    'date': row['reservation_date'] or '',
                    'type': 'APN-MALI'
                })
        except Exception as e:
            print(f"⚠️ Recent reservations query: {e}")
        
        # From reserved_ips (LAN IPs)
        try:
            cursor.execute("""
                SELECT branch_name, octet2, octet3, province, username, reservation_date, 'LAN' as type
                FROM reserved_ips 
                WHERE username IS NOT NULL AND username != ''
                ORDER BY reservation_date DESC
                LIMIT 5
            """)
            for row in cursor.fetchall():
                reservations.append({
                    'branch_name': row['branch_name'] or '',
                    'ip': f"10.{row['octet2']}.{row['octet3']}.0/24",
                    'province': row['province'] or '',
                    'username': row['username'] or '',
                    'date': row['reservation_date'] or '',
                    'type': 'LAN'
                })
        except Exception as e:
            print(f"⚠️ Recent reservations query: {e}")

        # From PTMP connections (manual saves)
        try:
            cursor.execute("""
                SELECT COALESCE(branch_name, branch_name_en) as branch_name,
                       interface_name as ip, province, username, reservation_date, 'PTMP' as type
                FROM ptmp_connections
                WHERE username IS NOT NULL AND username != ''
                AND reservation_date IS NOT NULL
                ORDER BY reservation_date DESC
                LIMIT 5
            """)
            for row in cursor.fetchall():
                reservations.append({
                    'branch_name': row['branch_name'] or '',
                    'ip': row['ip'] or '',
                    'province': row['province'] or '',
                    'username': row['username'] or '',
                    'date': row['reservation_date'] or '',
                    'type': 'PTMP'
                })
        except Exception as e:
            print(f"⚠️ Recent PTMP reservations query: {e}")

        conn.close()

        # Sort by date descending and return top 5
        reservations.sort(key=lambda x: x['date'] or '', reverse=True)
        return jsonify(reservations[:5])
        
    except Exception as e:
        print(f"❌ Recent reservations error: {e}")
        return jsonify([])

# ==================== TOP PROVINCES ====================
@app.route('/api/top-provinces', methods=['GET'])
def get_top_provinces():
    """Get top provinces by active IP count"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT province, COUNT(*) as count
            FROM lan_ips 
            WHERE province IS NOT NULL AND province != ''
            AND branch_name IS NOT NULL AND branch_name != ''
            GROUP BY province
            ORDER BY count DESC
            LIMIT 10
        """)
        
        provinces = []
        for row in cursor.fetchall():
            provinces.append({
                'province': row['province'],
                'count': row['count']
            })
        
        conn.close()
        return jsonify(provinces)
        
    except Exception as e:
        print(f"❌ Top provinces error: {e}")
        return jsonify([])

# ==================== TODAY ACTIVITY ====================
@app.route('/api/today-activity', methods=['GET'])
def get_today_activity():
    """Get today's activity count"""
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        
        # Read activity log
        if os.path.exists(ACTIVITY_LOG):
            with open(ACTIVITY_LOG, 'r', encoding='utf-8') as f:
                activities = json.load(f)
                
            today_count = sum(1 for a in activities if a.get('time', '').startswith(today))
            return jsonify({'count': today_count})
        
        return jsonify({'count': 0})
        
    except Exception as e:
        print(f"❌ Today activity error: {e}")
        return jsonify({'count': 0})

# ==================== PROVINCES ====================
@app.route('/api/provinces', methods=['GET'])
def get_provinces():
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Get provinces from all tables
        provinces = set()
        
        # Garbage values to filter out
        garbage = {'SW-Roof-To-Site', 'hgfvc', '؟؟؟', 'رزرو', 'سیار', 'لوازم یدکی شاهان'}
        
        # From lan_ips (main source)
        cursor.execute("""
            SELECT DISTINCT province FROM lan_ips 
            WHERE province IS NOT NULL AND province != ''
        """)
        for row in cursor.fetchall():
            if row[0] not in garbage:
                provinces.add(row[0])
        
        # From apn_mali (for APN مالی compatibility)
        cursor.execute("""
            SELECT DISTINCT province FROM apn_mali 
            WHERE province IS NOT NULL AND province != ''
        """)
        for row in cursor.fetchall():
            if row[0] and len(row[0]) > 2 and row[0] not in garbage:
                provinces.add(row[0])
        
        # From apn_ips
        cursor.execute("""
            SELECT DISTINCT province FROM apn_ips 
            WHERE province IS NOT NULL AND province != ''
        """)
        for row in cursor.fetchall():
            if row[0] and len(row[0]) > 2 and row[0] not in garbage:
                provinces.add(row[0])
        
        conn.close()
        result = sorted(list(provinces))
        print(f"✓ Provinces: {len(result)}")
        return jsonify(result)
    except Exception as e:
        print(f"❌ Provinces error: {e}")
        return jsonify([])

# ==================== BRANCHES ====================
@app.route('/api/branches', methods=['GET'])
def get_branches():
    """Get branches for APN-INT - from lan_ips table + reserved IPs"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, branch_name, province, octet2, octet3, wan_ip
            FROM lan_ips 
            WHERE branch_name IS NOT NULL AND branch_name != ''
            ORDER BY province, branch_name
        """)
        
        branches = []
        for row in cursor.fetchall():
            branches.append({
                'name': row['branch_name'],
                'province': row['province'] or '',
                'lan_ip': f"10.{row['octet2']}.{row['octet3']}.0/24",
                'x': row['octet2'],
                'y': row['octet3'],
                'type': 'active'
            })
        
        # Add reserved IPs (not yet activated)
        try:
            # Try with status column first
            cursor.execute("""
                SELECT id, branch_name, province, octet2, octet3
                FROM reserved_ips 
                WHERE status = 'reserved' OR status IS NULL
            """)
            for row in cursor.fetchall():
                branches.append({
                    'name': f"🔖 {row['branch_name']} (رزرو شده)",
                    'province': row['province'] or '',
                    'lan_ip': f"10.{row['octet2']}.{row['octet3']}.0/24",
                    'x': row['octet2'],
                    'y': row['octet3'],
                    'type': 'reserved'
                })
        except sqlite3.OperationalError:
            # status column doesn't exist, get all reservations
            try:
                cursor.execute("SELECT id, branch_name, province, octet2, octet3 FROM reserved_ips")
                for row in cursor.fetchall():
                    branches.append({
                        'name': f"🔖 {row['branch_name']} (رزرو شده)",
                        'province': row['province'] or '',
                        'lan_ip': f"10.{row['octet2']}.{row['octet3']}.0/24",
                        'x': row['octet2'],
                        'y': row['octet3'],
                        'type': 'reserved'
                    })
            except Exception:
                pass
        except Exception:
            pass
        
        conn.close()
        print(f"✓ Branches: {len(branches)}")
        return jsonify(branches)
    except Exception as e:
        print(f"❌ Branches error: {e}")
        return jsonify([])

@app.route('/api/mali-branches', methods=['GET'])
def get_mali_branches():
    """Get branches for APN-Mali - from lan_ips table (same as APN-INT) + reserved IPs"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Use lan_ips table (same as /api/branches for consistency)
        cursor.execute("""
            SELECT id, branch_name, province, octet2, octet3, wan_ip
            FROM lan_ips 
            WHERE branch_name IS NOT NULL AND branch_name != ''
            ORDER BY province, branch_name
        """)
        
        branches = []
        for row in cursor.fetchall():
            branches.append({
                'name': row['branch_name'],
                'province': row['province'] or '',
                'lan_ip': f"10.{row['octet2']}.{row['octet3']}.0/24",
                'x': row['octet2'],
                'y': row['octet3'],
                'type': 'active'
            })
        
        # Add reserved IPs (not yet activated)
        try:
            cursor.execute("""
                SELECT id, branch_name, province, octet2, octet3
                FROM reserved_ips 
                WHERE status = 'reserved' OR status IS NULL
            """)
            for row in cursor.fetchall():
                branches.append({
                    'name': f"🔖 {row['branch_name']} (رزرو شده)",
                    'province': row['province'] or '',
                    'lan_ip': f"10.{row['octet2']}.{row['octet3']}.0/24",
                    'x': row['octet2'],
                    'y': row['octet3'],
                    'type': 'reserved'
                })
        except sqlite3.OperationalError:
            try:
                cursor.execute("SELECT id, branch_name, province, octet2, octet3 FROM reserved_ips")
                for row in cursor.fetchall():
                    branches.append({
                        'name': f"🔖 {row['branch_name']} (رزرو شده)",
                        'province': row['province'] or '',
                        'lan_ip': f"10.{row['octet2']}.{row['octet3']}.0/24",
                        'x': row['octet2'],
                        'y': row['octet3'],
                        'type': 'reserved'
                    })
            except Exception:
                pass
        except Exception:
            pass
        
        conn.close()
        print(f"✓ Mali Branches (from lan_ips): {len(branches)}")
        return jsonify(branches)
    except Exception as e:
        print(f"❌ Mali Branches error: {e}")
        return jsonify([])

# ==================== INTRANET TUNNELS ====================
@app.route('/tunnels', methods=['GET'])
@app.route('/api/tunnels', methods=['GET'])
def get_tunnels():
    try:
        conn = get_db()
        cursor = conn.cursor()
        # Only return FREE tunnels (status = 'Free')
        cursor.execute("""
            SELECT * FROM intranet_tunnels 
            WHERE LOWER(status) = 'free'
            ORDER BY province, tunnel_name
        """)
        
        # Map to expected field names for HTML
        tunnels = []
        for row in cursor.fetchall():
            tunnels.append({
                'IP Address': row['ip_address'],
                'Tunnel Name': row['tunnel_name'] or '',
                'IP LAN': row['ip_lan'] or '',
                'IP Intranet': row['ip_intranet'] or '',
                'Description': row['description'] or '',
                'Province': row['province'] or '',
                'Status': row['status'] or ''
            })
        
        conn.close()
        print(f"✓ Free Tunnels: {len(tunnels)}")
        return jsonify(tunnels)
    except Exception as e:
        print(f"❌ Tunnels error: {e}")
        return jsonify([])

@app.route('/reserve', methods=['POST'])
def reserve_tunnel():
    try:
        data = request.json
        ip_address = data.get('IP Address')
        username = data.get('by')
        tunnel_name = data.get('Tunnel Name')
        ip_lan = data.get('IP LAN')
        ip_intranet = data.get('IP Intranet')
        description = data.get('Description')
        province = data.get('Province')
        branch_name = data.get('Branch Name') or None

        conn = get_db()
        cursor = conn.cursor()
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Update ALL fields, not just status
        cursor.execute("""
            UPDATE intranet_tunnels
            SET status = 'Reserved',
                reserved_by = ?,
                reserved_at = ?,
                tunnel_name = COALESCE(?, tunnel_name),
                ip_lan = COALESCE(?, ip_lan),
                ip_intranet = COALESCE(?, ip_intranet),
                description = COALESCE(?, description),
                province = COALESCE(?, province),
                branch_name = COALESCE(?, branch_name)
            WHERE ip_address = ?
        """, (username, now, tunnel_name, ip_lan, ip_intranet, description, province, branch_name, ip_address))
        
        conn.commit()
        conn.close()
        
        log_activity('success', 'رزرو تونل', ip_address, username)
        return jsonify({'status': 'ok'})
    except Exception as e:
        print(f"❌ Reserve tunnel error: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/check-tunnel-name', methods=['POST'])
def check_tunnel_name():
    """Check if a tunnel name is already used (reserved) in intranet_tunnels"""
    try:
        data = request.json
        tunnel_name = data.get('tunnel_name', '').strip()
        if not tunnel_name:
            return jsonify({'status': 'error', 'error': 'Tunnel name is required'}), 400

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, ip_address, description, reserved_by, status
            FROM intranet_tunnels
            WHERE tunnel_name = ? AND LOWER(status) != 'free'
        """, (tunnel_name,))
        row = cursor.fetchone()
        conn.close()

        if row:
            return jsonify({
                'exists': True,
                'ip_address': row['ip_address'] or '',
                'description': row['description'] or '',
                'reserved_by': row['reserved_by'] or '',
                'status': row['status'] or ''
            })
        return jsonify({'exists': False})
    except Exception as e:
        print(f"Check tunnel name error: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/reserved-intranet', methods=['GET'])
def get_reserved_intranet():
    """Get all reserved intranet tunnels for re-config feature"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        q = request.args.get('q', '').strip()

        if q:
            cursor.execute("""
                SELECT id, ip_address, tunnel_name, ip_lan, ip_intranet,
                       description, province, reserved_by, reserved_at
                FROM intranet_tunnels
                WHERE LOWER(status) = 'reserved'
                AND (tunnel_name LIKE ? OR description LIKE ? OR province LIKE ? OR ip_address LIKE ? OR ip_lan LIKE ?)
                ORDER BY reserved_at DESC
            """, (f'%{q}%', f'%{q}%', f'%{q}%', f'%{q}%', f'%{q}%'))
        else:
            cursor.execute("""
                SELECT id, ip_address, tunnel_name, ip_lan, ip_intranet,
                       description, province, reserved_by, reserved_at
                FROM intranet_tunnels
                WHERE LOWER(status) = 'reserved'
                ORDER BY reserved_at DESC
            """)

        results = []
        for row in cursor.fetchall():
            results.append({
                'id': row[0],
                'ip_address': row[1] or '',
                'tunnel_name': row[2] or '',
                'ip_lan': row[3] or '',
                'ip_intranet': row[4] or '',
                'description': row[5] or '',
                'province': row[6] or '',
                'reserved_by': row[7] or '',
                'reserved_at': row[8] or ''
            })
        conn.close()
        return jsonify(results)
    except Exception as e:
        print(f"Reserved intranet error: {e}")
        return jsonify([])

# ==================== VPLS/MPLS TUNNEL IPs ====================
@app.route('/api/vpls-tunnels', methods=['GET'])
def get_vpls_tunnels():
    """Get free VPLS/MPLS tunnel IPs, optionally filtered by province"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        province = request.args.get('province', '').strip()

        if province:
            # Map supports both English and Persian province names so freed tunnels
            # (stored with either form) are returned correctly after delete+re-search
            FA_TO_EN = {
                'اردبیل': 'Ardabil', 'آذربایجان شرقی': 'East Azerbaijan',
                'آذربایجان غربی': 'West Azerbaijan', 'البرز': 'Alborz',
                'اصفهان': 'Isfahan', 'ایلام': 'Ilam', 'بوشهر': 'Bushehr',
                'تهران': 'Tehran', 'چهارمحال و بختیاری': 'Chaharmahal and Bakhtiari',
                'خراسان جنوبی': 'South Khorasan', 'خراسان رضوی': 'Razavi Khorasan',
                'خراسان شمالی': 'North Khorasan', 'خوزستان': 'Khuzestan',
                'زنجان': 'Zanjan', 'سمنان': 'Semnan',
                'سیستان و بلوچستان': 'Sistan and Baluchestan',
                'فارس': 'Fars', 'قزوین': 'Qazvin', 'قم': 'Qom',
                'کرمان': 'Kerman', 'کرمانشاه': 'Kermanshah',
                'کهگیلویه و بویراحمد': 'Kohgiluyeh and Boyer-Ahmad',
                'گلستان': 'Golestan', 'گیلان': 'Gilan', 'لرستان': 'Lorestan',
                'مازندران': 'Mazandaran', 'مرکزی': 'Markazi',
                'هرمزگان': 'Hormozgan', 'همدان': 'Hamadan', 'یزد': 'Yazd',
            }
            EN_TO_FA = {v: k for k, v in FA_TO_EN.items()}
            alt = FA_TO_EN.get(province) or EN_TO_FA.get(province) or province
            province_variants = list(set([province, alt]))
            placeholders = ','.join(['?'] * len(province_variants))
            cursor.execute(f"""
                SELECT id, ip_address, hub_ip, branch_ip, tunnel_name, description,
                       province, status
                FROM vpls_tunnels
                WHERE LOWER(status) = 'free' AND province IN ({placeholders})
                ORDER BY id
            """, province_variants)
        else:
            cursor.execute("""
                SELECT id, ip_address, hub_ip, branch_ip, tunnel_name, description,
                       province, status
                FROM vpls_tunnels
                WHERE LOWER(status) = 'free'
                ORDER BY id
            """)
        tunnels = []
        for row in cursor.fetchall():
            tunnels.append({
                'id': row['id'],
                'ip_address': row['ip_address'],
                'hub_ip': row['hub_ip'],
                'branch_ip': row['branch_ip'],
                'tunnel_name': row['tunnel_name'] or '',
                'description': row['description'] or '',
                'province': row['province'] or '',
                'status': row['status']
            })
        conn.close()
        print(f"✓ Free VPLS tunnels: {len(tunnels)}")
        return jsonify(tunnels)
    except Exception as e:
        print(f"❌ VPLS tunnels error: {e}")
        return jsonify([])

@app.route('/api/reserve-vpls-tunnel', methods=['POST'])
def reserve_vpls_tunnel():
    """Reserve a VPLS/MPLS tunnel IP"""
    try:
        data = request.json
        tunnel_id = data.get('id')
        tunnel_name = data.get('tunnel_name', '')
        description = data.get('description', '')
        province = data.get('province', '')
        branch_name = data.get('branch_name', '')
        username = data.get('username', '')
        wan_ip = data.get('wan_ip', '')
        tunnel_dest = data.get('tunnel_dest', '')
        lan_ip = data.get('lan_ip', '')

        conn = get_db()
        cursor = conn.cursor()

        # Ensure lan_ip column exists (idempotent migration)
        try:
            cursor.execute("ALTER TABLE vpls_tunnels ADD COLUMN lan_ip TEXT")
            conn.commit()
        except Exception:
            pass  # Column already exists

        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        cursor.execute("""
            UPDATE vpls_tunnels
            SET status = 'Reserved',
                tunnel_name = ?,
                description = ?,
                province = ?,
                branch_name = ?,
                wan_ip = ?,
                tunnel_dest = ?,
                lan_ip = ?,
                username = ?,
                reservation_date = ?
            WHERE id = ? AND LOWER(status) = 'free'
        """, (tunnel_name, description, province, branch_name,
              wan_ip, tunnel_dest, lan_ip, username, now, tunnel_id))

        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'status': 'error', 'error': 'Tunnel IP already reserved or not found'}), 400

        conn.commit()
        conn.close()
        log_activity('success', 'رزرو تونل VPLS', tunnel_name, username)
        return jsonify({'status': 'ok'})
    except Exception as e:
        print(f"❌ Reserve VPLS tunnel error: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500

# ==================== PROVINCE TUNNEL TEMPLATES ====================
# Auto-discovered from analysis of VPLS_MPLS_Tunnel_IPs.xlsx
# Maps province abbreviation → hub IPs and branch WAN subnets
# hub_ip = tunnel destination for branches / tunnel source for hub
# subnet = first 3 octets of the WAN subnet used for branch tunnel source IPs
PROVINCE_TUNNEL_TEMPLATES = {
    'ARD':   {'x': 23, 'vpls': {'hub': '10.23.251.1',  'subnet': '10.23.251'},  'mpls': {'hub': '10.23.251.1',  'subnet': '10.23.251'}},
    'AZGH':  {'x': 33, 'vpls': {'hub': '10.33.251.1',  'subnet': '10.33.251'},  'mpls': {'hub': '10.33.251.1',  'subnet': '10.33.251'}},
    'AZSH':  {'x': 3,  'vpls': {'hub': '10.3.251.1',   'subnet': '10.3.251'},   'mpls': {'hub': '10.3.251.1',   'subnet': '10.3.251'}},
    'BSH':   {'x': 18, 'vpls': {'hub': '10.18.251.2',  'subnet': '10.18.251'},  'mpls': {'hub': '10.18.251.2',  'subnet': '10.18.251'}},
    'CHB':   {'x': 16, 'vpls': {'hub': '10.16.251.1',  'subnet': '10.16.251'},  'mpls': {'hub': '10.16.251.1',  'subnet': '10.16.251'}},
    'ESF':   {'x': 10, 'vpls': {'hub': '10.10.251.1',  'subnet': '10.10.251'},  'mpls': {'hub': '10.10.251.1',  'subnet': '10.10.251'}},
    'FRS':   {'x': 7,  'vpls': {'hub': '10.7.251.2',   'subnet': '10.7.251'},   'mpls': {'hub': '10.7.252.1',   'subnet': '10.7.252'}},
    'GIL':   {'x': 21, 'vpls': {'hub': '10.21.251.1',  'subnet': '10.21.251'},  'mpls': {'hub': '10.21.251.1',  'subnet': '10.21.251'}},
    'GLS':   {'x': 22, 'vpls': {'hub': '10.22.251.1',  'subnet': '10.22.251'},  'mpls': {'hub': '10.22.251.1',  'subnet': '10.22.251'}},
    'HMD':   {'x': 15, 'vpls': {'hub': '10.15.251.1',  'subnet': '10.15.251'},  'mpls': {'hub': '10.15.251.1',  'subnet': '10.15.251'}},
    'HMZ':   {'x': 17, 'vpls': {'hub': '10.17.251.1',  'subnet': '10.17.251'},  'mpls': {'hub': '10.17.251.1',  'subnet': '10.17.251'}},
    'ILM':   {'x': 25, 'vpls': {'hub': '10.25.251.1',  'subnet': '10.25.251'},  'mpls': {'hub': '10.25.251.1',  'subnet': '10.25.251'}},
    'KHB':   {'x': 26, 'vpls': {'hub': '10.26.251.2',  'subnet': '10.26.251'},  'mpls': {'hub': '10.26.251.2',  'subnet': '10.26.251'}},
    'KHR':   {'x': 9,  'vpls': {'hub': '10.9.252.1',   'subnet': '10.9.252'},   'mpls': {'hub': '10.9.252.1',   'subnet': '10.9.252'}},
    'KHRJ':  {'x': 29, 'vpls': {'hub': '10.29.250.1',  'subnet': '10.29.250'},  'mpls': {'hub': '10.29.250.1',  'subnet': '10.29.250'}},
    'KHSH':  {'x': 30, 'vpls': {'hub': '10.30.251.1',  'subnet': '10.30.251'},  'mpls': {'hub': '10.30.251.1',  'subnet': '10.30.251'}},
    'KHZ':   {'x': 6,  'vpls': {'hub': '10.6.253.1',   'subnet': '10.6.253'},   'mpls': {'hub': '10.6.249.1',   'subnet': '10.6.249'}},
    'KRD':   {'x': 12, 'vpls': {'hub': '10.12.251.1',  'subnet': '10.12.251'},  'mpls': {'hub': '10.12.251.1',  'subnet': '10.12.251'}},
    'KRMSH': {'x': 5,  'vpls': {'hub': '10.5.251.1',   'subnet': '10.5.251'},   'mpls': {'hub': '10.5.251.1',   'subnet': '10.5.251'}},
    'LOR':   {'x': 14, 'vpls': {'hub': '10.14.251.1',  'subnet': '10.14.251'},  'mpls': {'hub': '10.14.251.1',  'subnet': '10.14.251'}},
    'MAZ':   {'x': 32, 'vpls': {'hub': '10.32.251.2',  'subnet': '10.32.251'},  'mpls': {'hub': '10.32.251.2',  'subnet': '10.32.251'}},
    'MRZ':   {'x': 24, 'vpls': {'hub': '10.24.251.1',  'subnet': '10.24.251'},  'mpls': {'hub': '10.24.251.1',  'subnet': '10.24.251'}},
    'QOM':   {'x': 28, 'vpls': {'hub': '10.28.251.1',  'subnet': '10.28.251'},  'mpls': {'hub': '10.28.251.1',  'subnet': '10.28.251'}},
    'QZV':   {'x': 27, 'vpls': {'hub': '10.27.251.1',  'subnet': '10.27.251'},  'mpls': {'hub': '10.27.251.1',  'subnet': '10.27.251'}},
    'SMN':   {'x': 13, 'vpls': {'hub': '10.13.251.1',  'subnet': '10.13.251'},  'mpls': {'hub': '10.13.240.1',  'subnet': '10.13.240'}},
    'SNB':   {'x': 11, 'vpls': {'hub': '10.11.251.1',  'subnet': '10.11.251'},  'mpls': {'hub': '10.11.251.1',  'subnet': '10.11.251'}},
    'YZD':   {'x': 20, 'vpls': {'hub': '10.20.251.1',  'subnet': '10.20.251'},  'mpls': {'hub': '10.20.251.1',  'subnet': '10.20.251'}},
}

# Cache for Excel tunnel data (refreshes every 30 minutes)
_excel_tunnel_cache = {'data': None, 'loaded': False, 'time': 0}
EXCEL_CACHE_TTL = 1800  # 30 minutes

def _load_excel_tunnel_data():
    """Load and cache tunnel source/destination data from Excel"""
    if _excel_tunnel_cache['loaded'] and (time.time() - _excel_tunnel_cache['time']) < EXCEL_CACHE_TTL:
        return _excel_tunnel_cache['data']
    try:
        excel_path = os.path.join(os.path.dirname(__file__), 'data', 'VPLS_MPLS_Tunnel_IPs.xlsx')
        if os.path.exists(excel_path):
            df = pd.read_excel(excel_path, sheet_name='All_Tunnels')
            # Collect all IPs from tunnel_source and tunnel_destination
            all_ips = set()
            for col in ['tunnel_source', 'tunnel_destination']:
                for ip in df[col].dropna():
                    ip_str = str(ip).strip()
                    if ip_str and ip_str[0].isdigit():
                        all_ips.add(ip_str)
            _excel_tunnel_cache['data'] = all_ips
            _excel_tunnel_cache['loaded'] = True
            _excel_tunnel_cache['time'] = time.time()
            print(f"✓ Loaded {len(all_ips)} tunnel IPs from Excel cache")
            return all_ips
    except Exception as e:
        print(f"⚠️ Excel tunnel cache load error: {e}")
    _excel_tunnel_cache['data'] = set()
    _excel_tunnel_cache['loaded'] = True
    _excel_tunnel_cache['time'] = time.time()
    return set()

@app.route('/api/tunnel-template', methods=['GET'])
def get_tunnel_template():
    """Get tunnel source/destination template for a province.
    Returns hub IP, branch subnet, used IPs, and next available IP.
    """
    province_abbr = request.args.get('province_abbr', '').strip()
    service_type = request.args.get('service_type', 'VPLS').strip().upper()

    if province_abbr not in PROVINCE_TUNNEL_TEMPLATES:
        return jsonify({
            'available': False,
            'message': 'No auto-fill template available for this province. Please enter IPs manually.'
        })

    template = PROVINCE_TUNNEL_TEMPLATES[province_abbr]
    svc_key = 'vpls' if service_type == 'VPLS' else 'mpls'
    hub_ip = template[svc_key]['hub']
    subnet_prefix = template[svc_key]['subnet']  # e.g. '10.13.251'

    # Collect all used IPs in this subnet
    used_ips = set()

    # 1. From Excel data (cached)
    excel_ips = _load_excel_tunnel_data()
    for ip in excel_ips:
        if ip.startswith(subnet_prefix + '.'):
            used_ips.add(ip)

    # 2. From database (vpls_tunnels table - wan_ip and tunnel_dest columns)
    try:
        conn = get_db()
        cursor = conn.cursor()
        like_pattern = subnet_prefix + '.%'
        cursor.execute(
            "SELECT wan_ip, tunnel_dest FROM vpls_tunnels WHERE wan_ip LIKE ? OR tunnel_dest LIKE ?",
            (like_pattern, like_pattern)
        )
        for row in cursor.fetchall():
            if row['wan_ip'] and str(row['wan_ip']).startswith(subnet_prefix + '.'):
                used_ips.add(str(row['wan_ip']))
            if row['tunnel_dest'] and str(row['tunnel_dest']).startswith(subnet_prefix + '.'):
                used_ips.add(str(row['tunnel_dest']))
        conn.close()
    except Exception as e:
        print(f"⚠️ DB query for tunnel template: {e}")

    # Parse last octets of all used IPs
    used_last_octets = set()
    for ip in used_ips:
        parts = ip.split('.')
        if len(parts) == 4:
            try:
                used_last_octets.add(int(parts[3]))
            except ValueError:
                pass

    # Always reserve network (.0), broadcast (.255), and hub IP
    used_last_octets.add(0)
    used_last_octets.add(255)
    hub_parts = hub_ip.split('.')
    if len(hub_parts) == 4:
        try:
            used_last_octets.add(int(hub_parts[3]))
        except ValueError:
            pass

    # Find next available IP (start from 3 to avoid common gateway IPs .1/.2)
    next_free_ip = None
    start_search = 3
    for i in range(start_search, 255):
        if i not in used_last_octets:
            next_free_ip = f'{subnet_prefix}.{i}'
            break

    return jsonify({
        'available': True,
        'hub_ip': hub_ip,
        'branch_subnet': subnet_prefix + '.0/24',
        'next_free_ip': next_free_ip,
        'used_ips': sorted(list(used_ips)),
        'used_count': len(used_ips),
        'total_capacity': 252,
        'remaining': max(0, 253 - len(used_last_octets)),  # 253 usable (.1-.254 minus hub)
        'service_type': service_type,
        'province_abbr': province_abbr
    })

# ==================== PTMP MANAGEMENT ====================
@app.route('/api/save-ptmp', methods=['POST'])
def save_ptmp():
    """Save a PTMP serial configuration to the database (from ptmp.html)"""
    try:
        data = request.json
        branch_name = data.get('branchName', '').strip()
        hostname = data.get('hostname', '').strip()
        province = data.get('province', '').strip()
        lan_ip = data.get('lanIp', '').strip()
        serial_port = data.get('serialPort', '').strip() or 'Serial0/0/0'
        username = data.get('username', '').strip()

        if not hostname or not lan_ip:
            return jsonify({'status': 'error', 'error': 'Hostname and LAN IP are required'}), 400

        conn = get_db()
        cursor = conn.cursor()
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        desc = f'** {branch_name} - PTMP **' if branch_name else '** PTMP **'

        cursor.execute("""
            INSERT INTO ptmp_connections
            (interface_name, description, branch_name, branch_name_en,
             bandwidth, ip_type, encapsulation,
             province, province_abbr, router_hostname, router_file,
             status, username, reservation_date, lan_ip)
            VALUES (?, ?, ?, ?, '64', 'unnumbered', 'ppp',
                    ?, '', ?, 'manual', 'Manual', ?, ?, ?)
        """, (
            serial_port, desc,
            branch_name or None, branch_name or None,
            province, hostname, username, now, lan_ip
        ))

        conn.commit()
        conn.close()

        log_activity('success', 'ذخیره PTMP', f'{branch_name}: {hostname} ({serial_port})', username)
        return jsonify({'status': 'ok', 'message': f'PTMP configuration saved for {branch_name or hostname}'})
    except Exception as e:
        print(f"Save PTMP error: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/api/import-ptmp', methods=['POST'])
def import_ptmp_from_configs():
    """Parse router configs and import/refresh PTMP Serial interfaces"""
    try:
        username = request.json.get('username', 'system') if request.json else 'system'
        from parse_router_configs import import_serial_to_db
        count = import_serial_to_db()
        log_activity('info', 'وارد‌سازی PTMP', f'{count} Serial interface imported', username)
        return jsonify({'status': 'ok', 'count': count})
    except Exception as e:
        print(f"Import PTMP error: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500


@app.route('/api/ptmp-stats', methods=['GET'])
def ptmp_stats():
    """Get PTMP statistics per province"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT province, province_abbr,
                   COUNT(*) as total,
                   SUM(CASE WHEN branch_name IS NOT NULL THEN 1 ELSE 0 END) as matched,
                   SUM(CASE WHEN branch_name_en IS NOT NULL THEN 1 ELSE 0 END) as with_branch
            FROM ptmp_connections
            GROUP BY province
            ORDER BY total DESC
        """)
        stats = []
        for r in cursor.fetchall():
            stats.append({
                'province': r['province'] or '',
                'province_abbr': r['province_abbr'] or '',
                'total': r['total'],
                'matched': r['matched'],
                'with_branch': r['with_branch'],
            })
        conn.close()
        return jsonify(stats)
    except Exception as e:
        return jsonify([])

# ==================== TUNNEL200 IPs ====================
@app.route('/api/tunnel200-ips', methods=['GET'])
def get_tunnel200_ips():
    """Get free Tunnel200 IPs for APN غیرمالی"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        try:
            cursor.execute("""
                SELECT * FROM tunnel200_ips 
                WHERE status IS NULL OR status = '' OR LOWER(status) = 'free'
                ORDER BY id
                LIMIT 100
            """)
            
            ips = []
            for row in cursor.fetchall():
                try:
                    ips.append({
                        'id': row['id'],
                        'hub_ip': row['hub_ip'] if 'hub_ip' in row.keys() else '',
                        'branch_ip': row['branch_ip'] if 'branch_ip' in row.keys() else '',
                        'pair': row['pair_notation'] if 'pair_notation' in row.keys() else '',
                        'pair_notation': row['pair_notation'] if 'pair_notation' in row.keys() else '',
                        'tunnel_number': row['tunnel_number'] if 'tunnel_number' in row.keys() else '',
                        'interface_name': row['interface_name'] if 'interface_name' in row.keys() else '',
                        'description': row['description'] if 'description' in row.keys() else '',
                        'status': row['status'] if 'status' in row.keys() else ''
                    })
                except Exception as row_err:
                    print(f"⚠️ Row error: {row_err}")
                    continue
            
            conn.close()
            print(f"✓ Free Tunnel200 IPs: {len(ips)}")
            return jsonify(ips)
            
        except sqlite3.OperationalError as e:
            print(f"⚠️ tunnel200_ips table error: {e}")
            conn.close()
            return jsonify([])
            
    except Exception as e:
        print(f"❌ Tunnel200 error: {e}")
        return jsonify([])

@app.route('/api/reserve-tunnel200', methods=['POST'])
def reserve_tunnel200():
    try:
        data = request.json
        hub_ip = data.get('hub_ip')
        branch_ip = data.get('branch_ip')
        username = data.get('username')
        branch_name = data.get('branch_name', '')
        tunnel_number = data.get('tunnel_number', '')
        interface_name = data.get('interface_name', f'Tunnel{tunnel_number}')
        description = data.get('description', f'APN-INT-{branch_name}')
        
        conn = get_db()
        cursor = conn.cursor()
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        cursor.execute("""
            UPDATE tunnel200_ips 
            SET status = 'Reserved', 
                username = ?, 
                branch_name = ?, 
                tunnel_number = ?, 
                interface_name = ?,
                description = ?,
                reservation_date = ?
            WHERE hub_ip = ? AND branch_ip = ?
        """, (username, branch_name, tunnel_number, interface_name, description, now, hub_ip, branch_ip))
        
        conn.commit()
        conn.close()
        
        log_activity('success', 'رزرو Tunnel200', f"{hub_ip}/{branch_ip} - {branch_name}", username)
        return jsonify({'status': 'ok'})
    except Exception as e:
        print(f"❌ Reserve tunnel200 error: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500

# ==================== TUNNEL MALI ====================
@app.route('/api/free-tunnel-pairs', methods=['GET'])
def get_free_tunnel_pairs():
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM tunnel_mali 
            WHERE status IS NULL OR status = '' OR LOWER(status) = 'free'
            ORDER BY id
        """)
        
        ips = []
        for row in cursor.fetchall():
            interface = row['interface_name'] or ''
            tunnel_num = ''.join(filter(str.isdigit, interface))
            ip_addr = row['ip_address'] or ''
            
            # Get hub_ip and branch_ip from database columns
            hub_ip = ''
            branch_ip = ''
            try:
                hub_ip = row['hub_ip'] or ''
                branch_ip = row['branch_ip'] or ''
            except Exception:
                pass
            
            # Fallback calculation if columns don't exist
            if not hub_ip or not branch_ip:
                base_ip = ip_addr.replace('/31', '').strip()
                parts = base_ip.split('.')
                if len(parts) == 4:
                    last_octet = int(parts[3])
                    if last_octet % 2 == 0:
                        hub_ip = base_ip
                        branch_ip = f"{parts[0]}.{parts[1]}.{parts[2]}.{last_octet + 1}"
                    else:
                        branch_ip = base_ip
                        hub_ip = f"{parts[0]}.{parts[1]}.{parts[2]}.{last_octet - 1}"
            
            ips.append({
                'id': row['id'],
                'tunnel_number': tunnel_num,
                'tunnel_ip_hub': hub_ip,
                'tunnel_ip_branch': branch_ip,
                'interface_name': interface,
                'description': row['description'] or '',
                'ip_address': ip_addr,
                'destination_ip': row['destination_ip'] or ''
            })
        
        conn.close()
        print(f"✓ Free Tunnel Mali pairs: {len(ips)}")
        return jsonify(ips)
    except Exception as e:
        print(f"❌ Free Tunnel Pairs error: {e}")
        return jsonify([])

@app.route('/api/reserve-tunnel', methods=['POST'])
def reserve_tunnel_mali():
    try:
        data = request.json
        tunnel_id = data.get('tunnel_id') or data.get('id')
        tunnel_number = data.get('tunnel_number') or data.get('tunnelNumber')
        username = data.get('username')
        branch_name = data.get('branch_name') or data.get('branchName', '')
        interface_name = data.get('interface_name') or data.get('interfaceName', '')
        description = data.get('description', '')
        ip_address = data.get('ip_address') or data.get('ipAddress', '')
        hub_ip = data.get('hub_ip') or data.get('hubIp', '')
        branch_ip = data.get('branch_ip') or data.get('branchIp', '')
        destination_ip = data.get('destination_ip') or data.get('destinationIp', '')
        
        print(f"📥 Reserve tunnel: number={tunnel_number}, branch={branch_name}, interface={interface_name}")
        
        conn = get_db()
        cursor = conn.cursor()
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        rows_updated = 0
        if tunnel_id:
            cursor.execute("""
                UPDATE tunnel_mali 
                SET status = 'Reserved', 
                    username = ?, 
                    branch_name = ?, 
                    reservation_date = ?,
                    interface_name = COALESCE(NULLIF(?, ''), interface_name),
                    description = COALESCE(NULLIF(?, ''), description),
                    hub_ip = COALESCE(NULLIF(?, ''), hub_ip),
                    branch_ip = COALESCE(NULLIF(?, ''), branch_ip),
                    destination_ip = COALESCE(NULLIF(?, ''), destination_ip)
                WHERE id = ?
            """, (username, branch_name, now, interface_name, description, hub_ip, branch_ip, destination_ip, tunnel_id))
            rows_updated = cursor.rowcount
        elif tunnel_number:
            # Try multiple matching strategies
            # 1. Try exact match on interface_name like "Tunnel12345"
            cursor.execute("""
                UPDATE tunnel_mali 
                SET status = 'Reserved', 
                    username = ?, 
                    branch_name = ?, 
                    reservation_date = ?,
                    interface_name = COALESCE(NULLIF(?, ''), interface_name),
                    description = COALESCE(NULLIF(?, ''), description),
                    hub_ip = COALESCE(NULLIF(?, ''), hub_ip),
                    branch_ip = COALESCE(NULLIF(?, ''), branch_ip),
                    destination_ip = COALESCE(NULLIF(?, ''), destination_ip)
                WHERE (interface_name = ? OR interface_name LIKE ? OR interface_name LIKE ?)
                  AND (status IS NULL OR status = '' OR LOWER(status) = 'free')
                LIMIT 1
            """, (username, branch_name, now, interface_name, description, hub_ip, branch_ip, destination_ip, 
                  f'Tunnel{tunnel_number}', f'%{tunnel_number}%', f'Tunnel{tunnel_number}%'))
            rows_updated = cursor.rowcount
            
            # If no rows updated, try finding any free tunnel
            if rows_updated == 0:
                print(f"⚠️ No matching tunnel found for {tunnel_number}, trying first free tunnel...")
                cursor.execute("""
                    UPDATE tunnel_mali 
                    SET status = 'Reserved', 
                        username = ?, 
                        branch_name = ?, 
                        reservation_date = ?,
                        interface_name = COALESCE(NULLIF(?, ''), interface_name),
                        description = COALESCE(NULLIF(?, ''), description),
                        hub_ip = COALESCE(NULLIF(?, ''), hub_ip),
                        branch_ip = COALESCE(NULLIF(?, ''), branch_ip),
                        destination_ip = COALESCE(NULLIF(?, ''), destination_ip)
                    WHERE (status IS NULL OR status = '' OR LOWER(status) = 'free')
                    ORDER BY id
                    LIMIT 1
                """, (username, branch_name, now, interface_name, description, hub_ip, branch_ip, destination_ip))
                rows_updated = cursor.rowcount
        
        conn.commit()
        conn.close()
        
        print(f"✓ Tunnel reserved: {rows_updated} rows updated")
        log_activity('success', 'رزرو تونل مالی', f'Tunnel {tunnel_number} - {branch_name}', username)
        return jsonify({'status': 'ok', 'rows_updated': rows_updated})
    except Exception as e:
        print(f"❌ Reserve tunnel mali error: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500

# ==================== APN IPs ====================
@app.route('/api/apn-ips', methods=['GET'])
def get_apn_ips():
    """Get free APN IPs for APN غیرمالی (10.250.66.x)"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # First check if table exists and has correct structure
        try:
            cursor.execute("""
                SELECT * FROM apn_ips 
                WHERE (username IS NULL OR username = '')
                ORDER BY id
                LIMIT 100
            """)
            
            ips = []
            for row in cursor.fetchall():
                # Try different field names
                ip_value = None
                for field in ['ip_wan_apn', 'ip', 'ip_wan', 'ip_address']:
                    try:
                        ip_value = row[field]
                        if ip_value:
                            break
                    except Exception:
                        continue
                
                if ip_value:
                    ips.append({
                        'id': row['id'],
                        'ip': ip_value,
                        'province': row['province'] if 'province' in row.keys() else '',
                        'branch_name': row['branch_name'] if 'branch_name' in row.keys() else ''
                    })
            
            conn.close()
            print(f"✓ Free APN IPs (غیرمالی): {len(ips)}")
            return jsonify(ips)
            
        except sqlite3.OperationalError as e:
            print(f"⚠️ apn_ips table error: {e}")
            conn.close()
            
            # Fallback: Generate IPs from 10.250.66.x range
            print("📋 Generating APN IPs from 10.250.66.x range...")
            ips = []
            for i in range(2, 255):  # 10.250.66.2 to 10.250.66.254 (skip .1 for HUB)
                ips.append({
                    'id': i,
                    'ip': f'10.250.66.{i}',
                    'province': '',
                    'branch_name': ''
                })
            print(f"✓ Generated {len(ips)} APN IPs")
            return jsonify(ips)
            
    except Exception as e:
        print(f"❌ APN IPs error: {e}")
        return jsonify([])

@app.route('/api/mali-free-ips', methods=['GET'])
def get_mali_free_ips():
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Try to get free IPs from apn_mali table
        try:
            cursor.execute("""
                SELECT id, ip_wan, province, branch_name FROM apn_mali 
                WHERE (username IS NULL OR username = '')
                AND ip_wan IS NOT NULL AND ip_wan != ''
                ORDER BY id
            """)
            
            ips = []
            for row in cursor.fetchall():
                if row['ip_wan']:
                    ips.append({
                        'id': row['id'],
                        'ip': row['ip_wan'],
                        'province': row['province'] or '',
                        'branch_name': row['branch_name'] or ''
                    })
            
            conn.close()
            print(f"✓ Mali Free IPs: {len(ips)}")
            return jsonify(ips)
        except sqlite3.OperationalError as e:
            # Table structure might be different
            print(f"⚠️ Mali Free IPs query error: {e}")
            conn.close()
            return jsonify([])
            
    except Exception as e:
        print(f"❌ Mali Free IPs error: {e}")
        return jsonify([])

# ==================== LAN IPs ====================
@app.route('/api/free-lan-ips', methods=['GET'])
def get_free_lan_ips():
    """Get free LAN IPs from lan_ips table (rows without branch name = Free IPs)"""
    province = request.args.get('province')
    
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        if province:
            cursor.execute("""
                SELECT id, branch_name, province, octet2, octet3, wan_ip, status
                FROM lan_ips 
                WHERE (branch_name IS NULL OR branch_name = '' OR status = 'Free')
                AND (username IS NULL OR username = '')
                AND province = ?
                ORDER BY octet2, octet3
            """, (province,))
        else:
            cursor.execute("""
                SELECT id, branch_name, province, octet2, octet3, wan_ip, status
                FROM lan_ips 
                WHERE (branch_name IS NULL OR branch_name = '' OR status = 'Free')
                AND (username IS NULL OR username = '')
                ORDER BY province, octet2, octet3
                LIMIT 500
            """)
        
        ips = []
        for row in cursor.fetchall():
            ips.append({
                'id': row['id'],
                'ip': f"10.{row['octet2']}.{row['octet3']}.0/24",
                'octet2': row['octet2'],
                'octet3': row['octet3'],
                'branch_name': row['branch_name'] or '',
                'province': row['province'] or '',
                'status': row['status'] or 'Free'
            })
        
        conn.close()
        print(f"✓ Free LAN IPs: {len(ips)} (province={province})")
        return jsonify(ips)
    except Exception as e:
        print(f"❌ Free LAN error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify([])

@app.route('/api/used-lan-ips', methods=['GET'])
def get_used_lan_ips():
    """Get active branches from lan_ips (branches with name = Active/In Use)"""
    province = request.args.get('province')
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Active branches have branch_name
        if province:
            cursor.execute("""
                SELECT id, branch_name, province, octet2, octet3, wan_ip, username, reservation_date, status
                FROM lan_ips 
                WHERE branch_name IS NOT NULL AND branch_name != ''
                AND province = ?
                ORDER BY branch_name
            """, (province,))
        else:
            cursor.execute("""
                SELECT id, branch_name, province, octet2, octet3, wan_ip, username, reservation_date, status
                FROM lan_ips 
                WHERE branch_name IS NOT NULL AND branch_name != ''
                ORDER BY province, branch_name
            """)
        
        ips = []
        for row in cursor.fetchall():
            ips.append({
                'id': row['id'],
                'ip': f"10.{row['octet2']}.{row['octet3']}.0/24",
                'octet2': row['octet2'],
                'octet3': row['octet3'],
                'branch_name': row['branch_name'] or '',
                'province': row['province'] or '',
                'username': row['username'] or '',
                'reservation_date': row['reservation_date'] or '',
                'status': row['status'] or 'Active'
            })
        
        conn.close()
        print(f"✓ Active branches: {len(ips)}")
        return jsonify(ips)
    except Exception as e:
        print(f"❌ Used LAN error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify([])

# ==================== LAN IPs API ====================
@app.route('/api/lan-ips', methods=['GET'])
def get_lan_ips():
    """Get LAN IPs from lan_ips table with optional pagination"""
    try:
        conn = get_db()
        cursor = conn.cursor()

        # Pagination support
        page = request.args.get('page', type=int)
        per_page = request.args.get('per_page', 200, type=int)
        per_page = min(per_page, 500)  # Cap at 500

        province_filter = request.args.get('province', '').strip()

        query = """
            SELECT id, branch_name, province, octet2, octet3, wan_ip, status, username
            FROM lan_ips
            WHERE branch_name IS NOT NULL AND branch_name != ''
            AND octet2 IS NOT NULL AND octet3 IS NOT NULL
        """
        params = []
        if province_filter:
            query += " AND province = ?"
            params.append(province_filter)
        query += " ORDER BY province, branch_name"

        if page is not None:
            offset = (page - 1) * per_page
            query += " LIMIT ? OFFSET ?"
            params.extend([per_page, offset])

        cursor.execute(query, params)

        ips = []
        for row in cursor.fetchall():
            ips.append({
                'id': row['id'],
                'branch_name': row['branch_name'] or '',
                'province': row['province'] or '',
                'octet2': row['octet2'],
                'octet3': row['octet3'],
                'wan_ip': row['wan_ip'] or '',
                'status': row['status'] or 'Active',
                'username': row['username'] or ''
            })

        conn.close()
        print(f"✓ LAN IPs for monitoring: {len(ips)}")
        return jsonify({'success': True, 'data': ips})
    except Exception as e:
        print(f"❌ LAN IPs error: {e}")
        return jsonify({'success': False, 'error': str(e), 'data': []})

# ==================== CHECK REQUEST NUMBER ====================
@app.route('/api/check-request-number', methods=['GET'])
def check_request_number():
    request_number = request.args.get('request_number')
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM reserved_ips WHERE request_number = ?", (request_number,))
        row = cursor.fetchone()
        conn.close()
        return jsonify({'exists': row is not None, 'data': dict(row) if row else None})
    except Exception as e:
        return jsonify({'exists': False, 'error': str(e)})

# ==================== PROBLEMATIC NODES (for Dashboard) ====================
@app.route('/api/problematic-nodes', methods=['GET'])
def get_problematic_nodes():
    """Get a sample of nodes for monitoring preview (random sample for demo)"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Get a sample of branch nodes for preview
        cursor.execute("""
            SELECT branch_name, province, octet2, octet3
            FROM lan_ips 
            WHERE branch_name IS NOT NULL AND branch_name != ''
            AND octet2 IS NOT NULL AND octet3 IS NOT NULL
            LIMIT 50
        """)
        
        nodes = []
        for row in cursor.fetchall():
            nodes.append({
                'ip': f"10.{row['octet2']}.254.{row['octet3']}",
                'branchName': row['branch_name'],
                'province': row['province'] or '',
                'status': 'unknown',
                'lastCheck': None
            })
        
        conn.close()
        
        # Note: Actual status would need real-time ping data
        # This endpoint returns nodes that can be monitored
        return jsonify({
            'success': True, 
            'nodes': [],  # Empty - actual data comes from monitoring scan
            'message': 'Run monitoring scan to get real data'
        })
    except Exception as e:
        print(f"❌ Problematic nodes error: {e}")
        return jsonify({'success': False, 'nodes': [], 'error': str(e)})

# ==================== ACTIVATE RESERVATION ====================
@app.route('/api/activate-reservation', methods=['POST'])
def activate_reservation():
    """Move IP from reserved to used when config is generated"""
    try:
        data = request.json
        lan_ip = data.get('lan_ip', '')
        config_type = data.get('config_type', 'unknown')
        username = data.get('username', '')
        
        parts = lan_ip.replace('/24', '').split('.')
        if len(parts) < 3:
            return jsonify({'status': 'ok', 'was_reserved': False})
        
        try:
            octet2 = int(parts[1])
            octet3 = int(parts[2])
        except (ValueError, TypeError):
            return jsonify({'status': 'ok', 'was_reserved': False})
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Check if this IP is reserved (try with status column, fall back without)
        reservation = None
        try:
            cursor.execute("""
                SELECT id, branch_name FROM reserved_ips 
                WHERE octet2 = ? AND octet3 = ? AND (status = 'reserved' OR status IS NULL)
            """, (octet2, octet3))
            reservation = cursor.fetchone()
        except sqlite3.OperationalError:
            # status column doesn't exist, just check if row exists
            cursor.execute("""
                SELECT id, branch_name FROM reserved_ips 
                WHERE octet2 = ? AND octet3 = ?
            """, (octet2, octet3))
            reservation = cursor.fetchone()
        
        if reservation:
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Try to update with status column
            try:
                cursor.execute("""
                    UPDATE reserved_ips SET status = 'activated', activated_at = ?, config_type = ?
                    WHERE id = ?
                """, (now, config_type, reservation['id']))
            except sqlite3.OperationalError:
                pass  # status column doesn't exist, skip
            
            # Update lan_ips to mark as used
            cursor.execute("""
                UPDATE lan_ips SET status = 'Used', username = ?
                WHERE octet2 = ? AND octet3 = ?
            """, (username, octet2, octet3))
            
            conn.commit()
            conn.close()
            
            log_activity('success', 'فعال‌سازی رزرو', f'10.{octet2}.{octet3}.0/24 - {reservation["branch_name"]}', username)
            return jsonify({'status': 'ok', 'was_reserved': True, 'message': 'رزرو فعال شد'})
        
        conn.close()
        return jsonify({'status': 'ok', 'was_reserved': False})
    except Exception as e:
        print(f"❌ Activate reservation error: {e}")
        return jsonify({'status': 'error', 'was_reserved': False, 'message': str(e)})

# ==================== NEXT FREE LAN IP ====================
@app.route('/api/next-free-lan-ip', methods=['GET'])
def get_next_free_lan_ip():
    """Get the next available free LAN IP for a given province (octet2).
    Returns the first free 10.X.Y.0/24 for the selected province.
    """
    try:
        province = request.args.get('province', '').strip()
        if not province:
            return jsonify({'available': False, 'message': 'Province is required'})

        conn = get_db()
        cursor = conn.cursor()

        # Find octet2 values for this province
        cursor.execute("""
            SELECT DISTINCT octet2 FROM lan_ips
            WHERE province = ? AND octet2 IS NOT NULL
            ORDER BY octet2
        """, (province,))
        octet2_list = [row['octet2'] for row in cursor.fetchall()]

        if not octet2_list:
            conn.close()
            return jsonify({'available': False, 'message': 'Province not found in database'})

        # For each octet2, find free IPs
        free_ips = []
        for o2 in octet2_list:
            cursor.execute("""
                SELECT octet2, octet3, branch_name, status
                FROM lan_ips
                WHERE octet2 = ?
                AND (username IS NULL OR username = '')
                AND (branch_name IS NULL OR branch_name = '')
                AND (status IS NULL OR status = '' OR LOWER(status) = 'free')
                ORDER BY octet3
                LIMIT 10
            """, (o2,))
            for row in cursor.fetchall():
                free_ips.append({
                    'octet2': row['octet2'],
                    'octet3': row['octet3'],
                    'lan_ip': f"10.{row['octet2']}.{row['octet3']}.0/24",
                    'status': 'Free'
                })

        conn.close()

        if not free_ips:
            return jsonify({'available': False, 'message': 'No free IPs available for this province'})

        return jsonify({
            'available': True,
            'next_free': free_ips[0],
            'free_list': free_ips[:10],
            'total_free': len(free_ips),
            'province': province,
            'octet2_values': octet2_list
        })
    except Exception as e:
        print(f"Next free LAN IP error: {e}")
        return jsonify({'available': False, 'message': str(e)})

# ==================== RESERVE LAN IP ====================
@app.route('/api/reserve-lan', methods=['POST'])
def reserve_lan_ip():
    try:
        if is_api_rate_limited(request.remote_addr, 'reserve-lan'):
            return jsonify({'status': 'error', 'message': 'Too many requests. Please wait.'}), 429
        data = request.json
        
        # Handle both old and new parameter names
        lan_ip = data.get('lan_ip', '')
        province = data.get('province', '')
        branch_name = data.get('point_name_persian') or data.get('branch_name', '')
        username = data.get('reserved_by') or data.get('username', '')
        request_number = data.get('request_number', '')
        point_type = data.get('point_type', '')
        mehregostar_code = data.get('mehrgestar_code') or data.get('mehregostar_code', '')
        
        # Parse lan_ip to get octet2 and octet3
        octet2 = data.get('octet2')
        octet3 = data.get('octet3')
        
        if lan_ip and not octet2:
            # Parse from lan_ip like "10.1.20.0/24"
            parts = lan_ip.replace('/24', '').split('.')
            if len(parts) >= 3:
                octet2 = int(parts[1])
                octet3 = int(parts[2])
        
        if not octet2 or not octet3 or not username:
            return jsonify({'status': 'error', 'message': 'اطلاعات ناقص است'}), 400

        if not validate_octet(octet2) or not validate_octet(octet3):
            return jsonify({'status': 'error', 'message': 'فرمت IP نامعتبر است (مقادیر باید بین 0 تا 255 باشند)'}), 400
        
        conn = get_db()
        cursor = conn.cursor()

        now = datetime.now()
        expiry = now + timedelta(days=60)

        # Use BEGIN IMMEDIATE to prevent race conditions on concurrent reservations
        cursor.execute("BEGIN IMMEDIATE")
        try:
            # Check if IP is still free before reserving
            cursor.execute("""
                SELECT status FROM lan_ips WHERE octet2 = ? AND octet3 = ?
            """, (octet2, octet3))
            row = cursor.fetchone()
            if row and row['status'] and row['status'].lower() in ('reserved', 'used', 'activated'):
                conn.rollback()
                conn.close()
                return jsonify({'status': 'error', 'message': 'این IP قبلاً رزرو شده است'}), 409

            # Update lan_ips table
            cursor.execute("""
                UPDATE lan_ips SET username = ?, reservation_date = ?, branch_name = ?, status = 'Reserved'
                WHERE octet2 = ? AND octet3 = ?
            """, (username, now.strftime('%Y-%m-%d'), branch_name, octet2, octet3))

            # Insert into reserved_ips table
            cursor.execute("""
                INSERT INTO reserved_ips (province, octet2, octet3, branch_name, username, reservation_date, expiry_date, request_number, point_type, mehregostar_code, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'reserved')
            """, (province, octet2, octet3, branch_name, username, now.strftime('%Y-%m-%d'), expiry.strftime('%Y-%m-%d'), request_number, point_type, mehregostar_code))

            conn.commit()
        except Exception as inner_e:
            conn.rollback()
            raise inner_e
        finally:
            conn.close()
        
        log_activity('success', 'رزرو IP LAN', f'10.{octet2}.{octet3}.0 برای {branch_name}', username)
        return jsonify({
            'status': 'ok',
            'success': True, 
            'message': f'IP با موفقیت رزرو شد: 10.{octet2}.{octet3}.0/24',
            'ip': f'10.{octet2}.{octet3}.0/24',
            'reservation_date': now.strftime('%Y-%m-%d'), 
            'expiry_date': expiry.strftime('%Y-%m-%d')
        })
    except Exception as e:
        print(f"❌ Reserve LAN error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'success': False, 'message': str(e)}), 500

# ==================== RELEASE USED LAN ====================
@app.route('/api/release-used-lan', methods=['POST'])
def release_used_lan():
    try:
        data = request.json
        octet2 = data.get('octet2')
        octet3 = data.get('octet3')
        lan_id = data.get('id')
        lan_ip = data.get('lan_ip')
        
        # Parse lan_ip if provided (e.g., "10.3.25.0")
        if lan_ip and not octet2:
            parts = lan_ip.replace('/24', '').split('.')
            if len(parts) >= 3:
                try:
                    octet2 = int(parts[1])
                    octet3 = int(parts[2])
                except Exception:
                    pass
        
        if not octet2 or not octet3:
            return jsonify({'status': 'error', 'message': 'IP نامعتبر است'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE")
        try:
            if lan_id:
                cursor.execute("UPDATE lan_ips SET username = NULL, reservation_date = NULL, branch_name = NULL, status = 'Free' WHERE id = ?", (lan_id,))
            elif octet2 and octet3:
                cursor.execute("UPDATE lan_ips SET username = NULL, reservation_date = NULL, branch_name = NULL, status = 'Free' WHERE octet2 = ? AND octet3 = ?", (octet2, octet3))

            # Also remove from reserved_ips if exists
            cursor.execute("DELETE FROM reserved_ips WHERE octet2 = ? AND octet3 = ?", (octet2, octet3))
            conn.commit()
        except Exception as inner_e:
            conn.rollback()
            raise inner_e
        finally:
            conn.close()

        log_activity('warning', 'آزادسازی IP', f'10.{octet2}.{octet3}.0')
        return jsonify({'status': 'ok', 'success': True, 'message': f'IP 10.{octet2}.{octet3}.0 آزاد شد'})
    except Exception as e:
        print(f"❌ Release used LAN error: {e}")
        return jsonify({'status': 'error', 'success': False, 'message': str(e)}), 500

# ==================== RESERVED IPs ====================
@app.route('/api/reserved-ips', methods=['GET'])
def get_reserved_ips():
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM reserved_ips ORDER BY reservation_date DESC")
        
        reserved = []
        for row in cursor.fetchall():
            octet2 = row['octet2']
            octet3 = row['octet3']
            reserved.append({
                'id': row['id'],
                'lan_ip': f"10.{octet2}.{octet3}.0/24",
                'province': row['province'] or '',
                'point_name_persian': row['branch_name'] or '',
                'point_type': row['point_type'] or '',
                'request_number': row['request_number'] or '',
                'reserved_by': row['username'] or '',
                'reserved_date': row['reservation_date'] or '',
                'expiry_date': row['expiry_date'] or '',
                'status': 'RESERVED',
                'octet2': octet2,
                'octet3': octet3
            })
        
        conn.close()
        return jsonify(reserved)
    except Exception as e:
        print(f"❌ Reserved IPs error: {e}")
        return jsonify([])

@app.route('/api/release-lan', methods=['POST'])
def release_lan():
    """Release a reserved LAN IP"""
    try:
        data = request.json
        lan_ip = data.get('lan_ip', '')
        
        # Parse lan_ip to get octet2 and octet3
        parts = lan_ip.replace('/24', '').split('.')
        if len(parts) < 3:
            return jsonify({'status': 'error', 'message': 'فرمت IP نامعتبر است'})
        
        octet2 = int(parts[1])
        octet3 = int(parts[2])
        
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE")
        try:
            # Update lan_ips - clear reservation
            cursor.execute("""
                UPDATE lan_ips
                SET username = NULL, reservation_date = NULL, status = 'Free'
                WHERE octet2 = ? AND octet3 = ?
            """, (octet2, octet3))

            # Delete from reserved_ips
            cursor.execute("""
                DELETE FROM reserved_ips
                WHERE octet2 = ? AND octet3 = ?
            """, (octet2, octet3))
            conn.commit()
        except Exception as inner_e:
            conn.rollback()
            raise inner_e
        finally:
            conn.close()

        log_activity('success', 'آزادسازی IP', f'{lan_ip}', data.get('username', 'unknown'))

        return jsonify({
            'status': 'ok',
            'success': True,
            'message': f'IP {lan_ip} با موفقیت آزاد شد'
        })
    except Exception as e:
        print(f"❌ Release LAN error: {e}")
        return jsonify({'status': 'error', 'success': False, 'message': str(e)})

@app.route('/api/release-reservation', methods=['POST'])
def release_reservation():
    try:
        data = request.json
        rid = data.get('id')
        
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE")
        try:
            cursor.execute("SELECT octet2, octet3 FROM reserved_ips WHERE id = ?", (rid,))
            res = cursor.fetchone()

            if res:
                cursor.execute("UPDATE lan_ips SET username = NULL, reservation_date = NULL WHERE octet2 = ? AND octet3 = ?",
                               (res['octet2'], res['octet3']))
                cursor.execute("DELETE FROM reserved_ips WHERE id = ?", (rid,))
            conn.commit()
        except Exception as inner_e:
            conn.rollback()
            raise inner_e
        finally:
            conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== RESERVE IPs ====================
@app.route('/api/reserve-ips', methods=['POST'])
def reserve_ips():
    try:
        data = request.json
        username = data.get('username')
        branch_name = data.get('branchName')
        lan_ip = data.get('lanIp')
        apn_ip = data.get('apnIp')
        province = data.get('province', '')
        ip_type = data.get('type', 'APN-INT')
        
        updates = []
        conn = get_db()
        cursor = conn.cursor()
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute("BEGIN IMMEDIATE")
        try:
            if lan_ip:
                parts = lan_ip.replace('/24', '').split('.')
                if len(parts) >= 3:
                    octet2, octet3 = int(parts[1]), int(parts[2])

                    # Update lan_ips - mark as Used if it was Reserved
                    cursor.execute("""
                        UPDATE lan_ips SET username = ?, reservation_date = ?, status = 'Used'
                        WHERE octet2 = ? AND octet3 = ?
                    """, (username, now, octet2, octet3))
                    updates.append(f"LAN IP {lan_ip} فعال شد")

                    # Also update reserved_ips status to 'activated'
                    cursor.execute("""
                        UPDATE reserved_ips
                        SET status = 'activated', activated_at = ?, config_type = 'APN_INT'
                        WHERE octet2 = ? AND octet3 = ? AND (status = 'reserved' OR status IS NULL)
                    """, (now, octet2, octet3))

            if apn_ip:
                # Update apn_ips with ALL fields
                cursor.execute("""
                    UPDATE apn_ips
                    SET username = ?,
                        reservation_date = ?,
                        branch_name = ?,
                        province = COALESCE(?, province),
                        type = COALESCE(?, type),
                        lan_ip = COALESCE(?, lan_ip)
                    WHERE ip_wan_apn = ?
                """, (username, now, branch_name, province, ip_type, lan_ip, apn_ip))
                updates.append(f"APN IP {apn_ip} رزرو شد")

            # *** IMPORTANT: Also mark LAN IP as Used by branch_name ***
            if branch_name and not lan_ip:
                cursor.execute("""
                    UPDATE lan_ips
                    SET status = 'Used'
                    WHERE branch_name = ? AND status = 'Reserved'
                """, (branch_name,))

                if cursor.rowcount > 0:
                    updates.append(f"LAN IP برای {branch_name} فعال شد")

                    cursor.execute("""
                        UPDATE reserved_ips
                        SET status = 'activated', activated_at = ?, config_type = 'APN_INT'
                        WHERE branch_name = ? AND (status = 'reserved' OR status IS NULL)
                    """, (now, branch_name))

            conn.commit()
        except Exception as inner_e:
            conn.rollback()
            raise inner_e
        finally:
            conn.close()

        log_activity('success', 'رزرو IP', f'{branch_name}: {lan_ip}, {apn_ip}', username)
        return jsonify({'status': 'ok', 'updates': updates})
    except Exception as e:
        print(f"❌ Reserve IPs error: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/reserve-mali-ips', methods=['POST'])
def reserve_mali_ips():
    try:
        data = request.json
        username = data.get('username')
        branch_name = data.get('branchName')
        apn_ip = data.get('apnIp')
        tunnel_id = data.get('tunnelId')
        tunnel_ip_branch = data.get('tunnelIpBranch')
        tunnel_ip_hub = data.get('tunnelIpHub')
        tunnel_number = data.get('tunnelNumber')
        province = data.get('province', '')
        lan_ip = data.get('lanIp', '')
        interface_name = data.get('interfaceName', f'Tunnel{tunnel_number}')
        description = data.get('description', f'Gilanet-{branch_name}')
        destination_ip = data.get('destinationIp', '')
        node_type = data.get('type', 'APN-MALI')  # کیوسک، شعبه، ATM
        
        updates = []
        conn = get_db()
        cursor = conn.cursor()
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute("BEGIN IMMEDIATE")
        try:
            if apn_ip:
                cursor.execute("""
                    UPDATE apn_mali
                    SET username = ?,
                        reservation_date = ?,
                        branch_name = ?,
                        province = COALESCE(?, province),
                        type = ?,
                        lan_ip = COALESCE(?, lan_ip)
                    WHERE ip_wan = ?
                """, (username, now, branch_name, province, node_type, lan_ip, apn_ip))
                updates.append(f"APN Mali IP {apn_ip} رزرو شد")

            if tunnel_id:
                cursor.execute("""
                    UPDATE tunnel_mali
                    SET status = 'Reserved',
                        username = ?,
                        branch_name = ?,
                        reservation_date = ?,
                        interface_name = COALESCE(?, interface_name),
                        description = COALESCE(?, description),
                        ip_address = COALESCE(?, ip_address),
                        hub_ip = COALESCE(?, hub_ip),
                        branch_ip = COALESCE(?, branch_ip),
                        destination_ip = COALESCE(?, destination_ip)
                    WHERE id = ?
                """, (username, branch_name, now, interface_name, description,
                      tunnel_ip_branch, tunnel_ip_hub, tunnel_ip_branch, destination_ip, tunnel_id))
                updates.append(f"Tunnel Mali رزرو شد")

            # *** IMPORTANT: Mark LAN IP as Used (Active) ***
            if branch_name:
                cursor.execute("""
                    UPDATE lan_ips
                    SET status = 'Used'
                    WHERE branch_name = ? AND status = 'Reserved'
                """, (branch_name,))

                if cursor.rowcount > 0:
                    updates.append(f"LAN IP برای {branch_name} فعال شد")

                    cursor.execute("""
                        UPDATE reserved_ips
                        SET status = 'activated', activated_at = ?, config_type = 'APN_MALI'
                        WHERE branch_name = ? AND (status = 'reserved' OR status IS NULL)
                    """, (now, branch_name))

            conn.commit()
        except Exception as inner_e:
            conn.rollback()
            raise inner_e
        finally:
            conn.close()

        log_activity('success', 'رزرو IP مالی', f'{branch_name}: {apn_ip}, Tunnel: {tunnel_number}', username)
        return jsonify({'status': 'ok', 'updates': updates, 'message': f'تمام فیلدها برای {branch_name} ذخیره شد'})
    except Exception as e:
        print(f"❌ Reserve Mali IPs error: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500

# ==================== EXISTING RESERVED POINTS ====================
@app.route('/api/mali-reserved-points')
def mali_reserved_points():
    """Get all reserved APN Mali points from database"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, branch_name, province, type, lan_ip, ip_wan, username, reservation_date
            FROM apn_mali
            WHERE username IS NOT NULL AND username != ''
            ORDER BY reservation_date DESC
        """)
        rows = cursor.fetchall()
        points = []
        for r in rows:
            points.append({
                'id': r[0], 'branch_name': r[1], 'province': r[2],
                'type': r[3], 'lan_ip': r[4], 'ip_wan': r[5],
                'username': r[6], 'reservation_date': r[7]
            })
        conn.close()
        return jsonify(points)
    except Exception as e:
        print(f"❌ Mali reserved points error: {e}")
        return jsonify([])

@app.route('/api/int-reserved-points')
def int_reserved_points():
    """Get all reserved APN INT points from database"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, branch_name, province, type, lan_ip, ip_wan_apn, username, reservation_date
            FROM apn_ips
            WHERE username IS NOT NULL AND username != ''
            ORDER BY reservation_date DESC
        """)
        rows = cursor.fetchall()
        points = []
        for r in rows:
            points.append({
                'id': r[0], 'branch_name': r[1], 'province': r[2],
                'type': r[3], 'lan_ip': r[4], 'ip_wan_apn': r[5],
                'username': r[6], 'reservation_date': r[7]
            })
        conn.close()
        return jsonify(points)
    except Exception as e:
        print(f"❌ INT reserved points error: {e}")
        return jsonify([])

@app.route('/api/free-mali-point', methods=['POST'])
def free_mali_point():
    """Free a reserved APN Mali point - release IP and tunnel"""
    try:
        data = request.json
        point_id = data.get('id')
        username = data.get('username', '')

        if not point_id:
            return jsonify({'status': 'error', 'error': 'شناسه نقطه مشخص نشده'}), 400

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE")
        try:
            # Get point info before freeing
            cursor.execute("SELECT branch_name, ip_wan, lan_ip FROM apn_mali WHERE id = ?", (point_id,))
            point = cursor.fetchone()
            if not point:
                conn.rollback()
                conn.close()
                return jsonify({'status': 'error', 'error': 'نقطه پیدا نشد'}), 404

            branch_name = point[0]
            ip_wan = point[1]
            lan_ip = point[2]

            updates = []

            # Free APN Mali IP
            cursor.execute("""
                UPDATE apn_mali SET username = NULL, branch_name = NULL, province = NULL,
                type = NULL, lan_ip = NULL, reservation_date = NULL WHERE id = ?
            """, (point_id,))
            updates.append(f'IP APN مالی آزاد شد: {ip_wan}')

            # Free associated tunnel (by destination_ip matching ip_wan)
            cursor.execute("""
                UPDATE tunnel_mali SET status = NULL, username = NULL, branch_name = NULL,
                reservation_date = NULL, description = NULL, destination_ip = NULL
                WHERE destination_ip = ?
            """, (ip_wan,))
            if cursor.rowcount > 0:
                updates.append(f'Tunnel مالی مرتبط آزاد شد')

            conn.commit()
        except Exception as inner_e:
            conn.rollback()
            raise inner_e
        finally:
            conn.close()

        log_activity('warning', 'آزادسازی IP مالی', f'{branch_name}: {ip_wan}', username)
        return jsonify({'status': 'ok', 'updates': updates, 'message': f'نقطه {branch_name} آزاد شد'})
    except Exception as e:
        print(f"❌ Free Mali point error: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/free-int-point', methods=['POST'])
def free_int_point():
    """Free a reserved APN INT point - release IP and tunnel200"""
    try:
        data = request.json
        point_id = data.get('id')
        username = data.get('username', '')

        if not point_id:
            return jsonify({'status': 'error', 'error': 'شناسه نقطه مشخص نشده'}), 400

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("BEGIN IMMEDIATE")
        try:
            # Get point info before freeing
            cursor.execute("SELECT branch_name, ip_wan_apn, lan_ip FROM apn_ips WHERE id = ?", (point_id,))
            point = cursor.fetchone()
            if not point:
                conn.rollback()
                conn.close()
                return jsonify({'status': 'error', 'error': 'نقطه پیدا نشد'}), 404

            branch_name = point[0]
            ip_wan_apn = point[1]
            lan_ip = point[2]

            updates = []

            # Free APN INT IP
            cursor.execute("""
                UPDATE apn_ips SET username = NULL, branch_name = NULL, province = NULL,
                type = NULL, lan_ip = NULL, reservation_date = NULL WHERE id = ?
            """, (point_id,))
            updates.append(f'IP APN غیرمالی آزاد شد: {ip_wan_apn}')

            # Free associated tunnel200 (by branch_name)
            cursor.execute("""
                UPDATE tunnel200_ips SET status = NULL, username = NULL, branch_name = NULL,
                reservation_date = NULL, description = NULL
                WHERE branch_name = ?
            """, (branch_name,))
            if cursor.rowcount > 0:
                updates.append(f'Tunnel200 مرتبط آزاد شد')

            conn.commit()
        except Exception as inner_e:
            conn.rollback()
            raise inner_e
        finally:
            conn.close()

        log_activity('warning', 'آزادسازی IP غیرمالی', f'{branch_name}: {ip_wan_apn}', username)
        return jsonify({'status': 'ok', 'updates': updates, 'message': f'نقطه {branch_name} آزاد شد'})
    except Exception as e:
        print(f"❌ Free INT point error: {e}")
        return jsonify({'status': 'error', 'error': str(e)}), 500

# ==================== PING ====================
@app.route('/api/ping-lan-ip', methods=['POST'])
def ping_lan_ip():
    try:
        data = request.json
        lan_ip = data.get('lan_ip', '')
        octet2 = data.get('octet2')
        octet3 = data.get('octet3')
        
        # Parse lan_ip if provided
        if lan_ip and not octet2:
            parts = lan_ip.replace('/24', '').split('.')
            if len(parts) >= 3:
                octet2 = int(parts[1])
                octet3 = int(parts[2])
        
        if not octet2 or not octet3:
            return jsonify({'reachable': False, 'message': 'پارامترها ناقص است'})
        
        # Ping format: 10.{octet2}.254.{octet3}
        ping_ip = f"10.{octet2}.254.{octet3}"
        
        # Cross-platform ping command
        if platform.system().lower() == 'windows':
            cmd = ['ping', '-n', '2', '-w', '2000', ping_ip]
        else:
            cmd = ['ping', '-c', '2', '-W', '2', ping_ip]

        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=10
        )
        
        reachable = result.returncode == 0
        
        if reachable:
            return jsonify({
                'reachable': True,
                'pinged_ip': ping_ip,
                'message': f'⚠️ IP پاسخ می‌دهد! ممکن است در استفاده باشد.'
            })
        else:
            return jsonify({
                'reachable': False,
                'pinged_ip': ping_ip,
                'message': f'✅ IP آزاد است ({ping_ip} پاسخ نداد)'
            })
    except subprocess.TimeoutExpired:
        return jsonify({
            'reachable': False,
            'pinged_ip': ping_ip if 'ping_ip' in locals() else '',
            'message': '✅ IP آزاد است (Timeout)'
        })
    except Exception as e:
        print(f"❌ Ping error: {e}")
        return jsonify({'reachable': False, 'message': f'خطا: {str(e)}'})

@app.route('/api/check-expired-reservations', methods=['GET', 'POST'])
def check_expired_reservations():
    """
    Manually check and optionally release expired reservations.
    GET: Just check how many are expired
    POST: Actually release them
    """
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        today = datetime.now().strftime('%Y-%m-%d')
        
        # Find expired reservations that are still 'reserved'
        cursor.execute("""
            SELECT r.id, r.octet2, r.octet3, r.branch_name, r.username, 
                   r.reservation_date, r.expiry_date, r.status
            FROM reserved_ips r
            WHERE r.expiry_date < ? 
            AND (r.status = 'reserved' OR r.status IS NULL)
        """, (today,))
        
        expired = cursor.fetchall()
        expired_list = []
        
        for row in expired:
            expired_list.append({
                'id': row['id'],
                'ip': f"10.{row['octet2']}.{row['octet3']}.0/24",
                'branch_name': row['branch_name'],
                'username': row['username'],
                'reservation_date': row['reservation_date'],
                'expiry_date': row['expiry_date'],
                'status': row['status']
            })
        
        if request.method == 'POST' and expired_list:
            # Actually release them
            released = 0
            for row in expired:
                cursor.execute("""
                    UPDATE lan_ips 
                    SET username = NULL, reservation_date = NULL, branch_name = NULL, status = 'Free'
                    WHERE octet2 = ? AND octet3 = ? AND status = 'Reserved'
                """, (row['octet2'], row['octet3']))
                
                cursor.execute("DELETE FROM reserved_ips WHERE id = ?", (row['id'],))
                released += 1
            
            conn.commit()
            conn.close()
            
            log_activity('info', 'آزادسازی دستی', f'{released} IP منقضی شده آزاد شد', 'Admin')
            return jsonify({
                'success': True,
                'released': released,
                'message': f'{released} IP منقضی شده آزاد شد'
            })
        
        conn.close()
        return jsonify({
            'success': True,
            'expired_count': len(expired_list),
            'expired': expired_list,
            'message': f'{len(expired_list)} IP منقضی شده پیدا شد' if expired_list else 'هیچ IP منقضی شده‌ای وجود ندارد'
        })
        
    except Exception as e:
        print(f"❌ Check expired error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/ping-loopback', methods=['POST'])
def ping_loopback():
    try:
        data = request.json
        
        # Accept either full IP or octet2/octet3
        loopback_ip = data.get('loopback_ip', '')
        
        if loopback_ip:
            ip = loopback_ip
        else:
            octet2 = data.get('octet2')
            octet3 = data.get('octet3')
            ip = f"10.{octet2}.254.{octet3}"
        
        # Cross-platform ping command
        if platform.system().lower() == 'windows':
            cmd = ['ping', '-n', '2', '-w', '2000', ip]
        else:
            cmd = ['ping', '-c', '2', '-W', '2', ip]
        
        start_time = time.time()
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
        response_time = (time.time() - start_time) * 1000 / 2  # Average per ping
        
        reachable = result.returncode == 0
        
        return jsonify({
            'success': reachable,
            'reachable': reachable,
            'ip': ip,
            'responseTime': round(response_time, 2) if reachable else None,
            'message': f'✅ {ip} پاسخ داد' if reachable else f'❌ {ip} پاسخ نداد'
        })
    except subprocess.TimeoutExpired:
        return jsonify({'success': False, 'reachable': False, 'message': 'Timeout'})
    except Exception as e:
        print(f"❌ Ping loopback error: {e}")
        return jsonify({'success': False, 'reachable': False, 'message': str(e)})

# ==================== DB MANAGEMENT ====================
# Activity log cache
_activity_cache = {'data': None, 'time': 0}

@app.route('/api/db/activity', methods=['GET'])
def get_activity():
    global _activity_cache
    
    # Return cached result if less than 30 seconds old
    if _activity_cache['data'] and (time.time() - _activity_cache['time']) < 30:
        return jsonify(_activity_cache['data'])
    
    try:
        if os.path.exists(ACTIVITY_LOG):
            with open(ACTIVITY_LOG, 'r', encoding='utf-8') as f:
                data = json.load(f)
                _activity_cache['data'] = data
                _activity_cache['time'] = time.time()
                return jsonify(data)
        return jsonify([])
    except Exception:
        return jsonify([])

@app.route('/api/db/preview-excel', methods=['POST'])
def preview_excel():
    try:
        username = request.form.get('username', '')
        if username != DB_ADMIN_USER:
            return jsonify({'error': 'فقط مدیر سیستم دسترسی دارد'}), 403
        file = request.files.get('file')
        if not file:
            return jsonify({'error': 'فایل انتخاب نشده'}), 400
        
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        return jsonify({
            'columns': list(df.columns), 
            'preview': df.head(20).fillna('').to_dict('records'), 
            'total_rows': len(df)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/db/import-excel', methods=['POST'])
def import_excel():
    try:
        file = request.files.get('file')
        table_name = request.form.get('table')
        username = request.form.get('username')
        
        if username != DB_ADMIN_USER:
            return jsonify({'error': 'فقط مدیر سیستم می‌تواند دیتابیس را تغییر دهد'}), 403
        
        if not file or not table_name:
            return jsonify({'error': 'فایل یا نام جدول مشخص نشده'}), 400
        
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        backup_name = f'backup_before_import_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
        shutil.copy2(DB_PATH, os.path.join(BACKUP_DIR, backup_name))
        
        conn = get_db()
        df.to_sql(table_name, conn, if_exists='replace', index=False)
        conn.close()
        
        log_activity('success', 'آپلود دیتا', f'{table_name}: {len(df)} ردیف', username)
        return jsonify({'success': True, 'rows': len(df), 'backup': backup_name})
    except Exception as e:
        print(f"❌ Import Excel error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/db/backup', methods=['POST'])
def create_backup():
    try:
        data = request.json or {}
        username = data.get('username', '')
        if username != DB_ADMIN_USER:
            return jsonify({'error': 'فقط مدیر سیستم می‌تواند بکاپ بگیرد'}), 403
        fname = f'backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
        shutil.copy2(DB_PATH, os.path.join(BACKUP_DIR, fname))
        log_activity('backup', 'Backup', fname, username)
        return jsonify({'success': True, 'filename': fname})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/db/backups', methods=['GET'])
def list_backups():
    try:
        backups = []
        if os.path.exists(BACKUP_DIR):
            for f in sorted(os.listdir(BACKUP_DIR), reverse=True):
                if f.endswith('.db'):
                    backups.append({
                        'filename': f, 
                        'size': f'{os.path.getsize(os.path.join(BACKUP_DIR, f))/1024:.1f} KB'
                    })
        return jsonify(backups)
    except Exception:
        return jsonify([])

@app.route('/api/db/restore', methods=['POST'])
def restore_backup():
    try:
        data = request.json
        fname = data.get('filename')
        username = data.get('username')
        
        if username != DB_ADMIN_USER:
            return jsonify({'error': 'فقط مدیر سیستم می‌تواند بازیابی کند'}), 403

        # Sanitize filename to prevent path traversal
        fname = os.path.basename(fname)
        if not fname.endswith('.db'):
            return jsonify({'error': 'فرمت فایل نامعتبر'}), 400

        src = os.path.join(BACKUP_DIR, fname)
        if os.path.exists(src):
            shutil.copy2(src, DB_PATH)
            log_activity('success', 'بازیابی', fname, username)
            return jsonify({'success': True})
        return jsonify({'error': 'فایل یافت نشد'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/db/reset-users', methods=['POST'])
def reset_users():
    try:
        data = request.json
        username = data.get('username')
        
        if username != DB_ADMIN_USER:
            return jsonify({'error': 'فقط مدیر سیستم می‌تواند این کار را انجام دهد'}), 403
        
        conn = get_db()
        conn.execute('DELETE FROM user_passwords')
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== EXPORT CSV (Admin Only) ====================
@app.route('/api/export/lan-ips', methods=['GET'])
def export_lan_ips():
    """Export all LAN IPs as CSV - Admin only"""
    username = request.args.get('username', '')
    if username != DB_ADMIN_USER:
        return jsonify({'error': 'دسترسی فقط برای ادمین مجاز است'}), 403
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT octet2, octet3, branch_name, province, wan_ip, username, reservation_date, status
            FROM lan_ips
            WHERE branch_name IS NOT NULL AND branch_name != ''
            ORDER BY province, branch_name
        """)
        rows = cursor.fetchall()
        conn.close()

        import io, csv
        output = io.StringIO()
        output.write('\ufeff')  # BOM for Excel Persian support
        writer = csv.writer(output)
        writer.writerow(['IP LAN', 'نام شعبه', 'استان', 'WAN IP', 'کاربر', 'تاریخ رزرو', 'وضعیت'])
        for r in rows:
            writer.writerow([f"10.{r['octet2']}.{r['octet3']}.0/24", r['branch_name'] or '', r['province'] or '',
                             r['wan_ip'] or '', r['username'] or '', r['reservation_date'] or '', r['status'] or 'Active'])

        return Response(output.getvalue(), mimetype='text/csv',
                        headers={'Content-Disposition': 'attachment; filename=lan_ips_export.csv'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/reservations', methods=['GET'])
def export_reservations():
    """Export all reservations as CSV - Admin only"""
    username = request.args.get('username', '')
    if username != DB_ADMIN_USER:
        return jsonify({'error': 'دسترسی فقط برای ادمین مجاز است'}), 403
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM reserved_ips ORDER BY reservation_date DESC")
        rows = cursor.fetchall()
        conn.close()

        import io, csv
        output = io.StringIO()
        output.write('\ufeff')
        writer = csv.writer(output)
        writer.writerow(['IP LAN', 'استان', 'نام شعبه', 'نوع نقطه', 'شماره درخواست', 'کاربر', 'تاریخ رزرو', 'تاریخ انقضا', 'وضعیت'])
        for r in rows:
            writer.writerow([f"10.{r['octet2']}.{r['octet3']}.0/24", r['province'] or '', r['branch_name'] or '',
                             r['point_type'] or '', r['request_number'] or '', r['username'] or '',
                             r['reservation_date'] or '', r['expiry_date'] or '', r['status'] or 'reserved'])

        return Response(output.getvalue(), mimetype='text/csv',
                        headers={'Content-Disposition': 'attachment; filename=reservations_export.csv'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== SMART SEARCH ====================
@app.route('/api/search', methods=['GET'])
def smart_search():
    """Search across all tables: IPs, tunnels, branches"""
    q = request.args.get('q', '').strip()
    if not q or len(q) < 2:
        return jsonify([])

    results = []
    conn = get_db()
    cursor = conn.cursor()
    like = f'%{q}%'

    # Search LAN IPs
    cursor.execute("""
        SELECT 'lan_ip' as type, branch_name, province, octet2, octet3, username, status
        FROM lan_ips WHERE branch_name LIKE ? OR province LIKE ? OR username LIKE ?
        OR (octet2||'.'||octet3) LIKE ? LIMIT 15
    """, (like, like, like, like))
    for r in cursor.fetchall():
        results.append({
            'type': 'lan_ip', 'icon': '📍',
            'title': f"10.{r['octet2']}.{r['octet3']}.0/24",
            'subtitle': f"{r['province']} - {r['branch_name'] or 'آزاد'}",
            'extra': r['username'] or 'بدون کاربر',
            'status': r['status'] or 'Free',
            'link': '/reserve-lan'
        })

    # Search Tunnels
    cursor.execute("""
        SELECT 'tunnel' as type, tunnel_name, ip_address, description, province, status, branch_name
        FROM intranet_tunnels WHERE branch_name LIKE ? OR tunnel_name LIKE ? OR ip_address LIKE ? OR description LIKE ? OR province LIKE ? LIMIT 10
    """, (like, like, like, like, like))
    for r in cursor.fetchall():
        display_name = r['branch_name'] or r['description'] or r['tunnel_name'] or r['ip_address']
        results.append({
            'type': 'tunnel', 'icon': '🔗',
            'title': display_name,
            'subtitle': r['province'] or '',
            'extra': r['ip_address'] or '',
            'status': r['status'] or 'Free',
            'link': '/intranet'
        })

    # Search APN IPs
    cursor.execute("""
        SELECT 'apn' as type, branch_name, province, lan_ip, ip_wan_apn, username
        FROM apn_ips WHERE branch_name LIKE ? OR province LIKE ? OR lan_ip LIKE ? OR ip_wan_apn LIKE ? LIMIT 10
    """, (like, like, like, like))
    for r in cursor.fetchall():
        results.append({
            'type': 'apn_int', 'icon': '🟣',
            'title': r['ip_wan_apn'] or r['lan_ip'] or '',
            'subtitle': f"{r['province']} - {r['branch_name'] or ''}",
            'extra': r['username'] or 'آزاد',
            'status': 'Used' if r['username'] else 'Free',
            'link': '/apn-int'
        })

    # Search APN Mali
    cursor.execute("""
        SELECT 'apn_mali' as type, branch_name, province, lan_ip, ip_wan, username
        FROM apn_mali WHERE branch_name LIKE ? OR province LIKE ? OR lan_ip LIKE ? OR ip_wan LIKE ? LIMIT 10
    """, (like, like, like, like))
    for r in cursor.fetchall():
        results.append({
            'type': 'apn_mali', 'icon': '🟢',
            'title': r['ip_wan'] or r['lan_ip'] or '',
            'subtitle': f"{r['province']} - {r['branch_name'] or ''}",
            'extra': r['username'] or 'آزاد',
            'status': 'Used' if r['username'] else 'Free',
            'link': '/apn-mali'
        })

    # Search PTMP Serial connections
    try:
        cursor.execute("""
            SELECT COALESCE(branch_name, branch_name_en) as bname, province, interface_name, lan_ip
            FROM ptmp_connections
            WHERE branch_name LIKE ? OR branch_name_en LIKE ? OR interface_name LIKE ? OR province LIKE ?
            LIMIT 10
        """, (like, like, like, like))
        for r in cursor.fetchall():
            results.append({
                'type': 'ptmp', 'icon': '📡',
                'title': r['bname'] or r['interface_name'] or '',
                'subtitle': f"{r['province'] or ''} - {r['interface_name'] or ''}",
                'extra': r['lan_ip'] or '',
                'status': 'Used',
                'link': '/service-management'
            })
    except Exception:
        pass

    conn.close()
    return jsonify(results[:30])


# ==================== PDF REPORT ====================
@app.route('/api/report/pdf', methods=['GET'])
def generate_pdf_report():
    """Generate network status PDF report"""
    try:
        conn = get_db()
        cursor = conn.cursor()

        # Gather all stats
        cursor.execute("""
            SELECT
                (SELECT COUNT(*) FROM lan_ips) as total_lan,
                (SELECT COUNT(*) FROM lan_ips WHERE (username IS NULL OR username = '') AND (branch_name IS NULL OR branch_name = '')) as free_lan,
                (SELECT COUNT(*) FROM intranet_tunnels) as total_tun,
                (SELECT COUNT(*) FROM intranet_tunnels WHERE LOWER(status) = 'free') as free_tun,
                (SELECT COUNT(*) FROM apn_ips) as total_apn,
                (SELECT COUNT(*) FROM apn_ips WHERE username IS NULL OR username = '') as free_apn,
                (SELECT COUNT(*) FROM apn_mali) as total_mali,
                (SELECT COUNT(*) FROM apn_mali WHERE username IS NULL OR username = '') as free_mali,
                (SELECT COUNT(*) FROM reserved_ips WHERE status = 'reserved') as active_reservations,
                (SELECT COUNT(*) FROM ptmp_connections) as total_ptmp,
                (SELECT COUNT(*) FROM ptmp_connections WHERE branch_name IS NOT NULL) as matched_ptmp
        """)
        s = cursor.fetchone()

        # Top provinces
        cursor.execute("""
            SELECT province, COUNT(*) as cnt FROM lan_ips
            WHERE province IS NOT NULL AND province != '' AND (username IS NOT NULL AND username != '')
            GROUP BY province ORDER BY cnt DESC LIMIT 10
        """)
        top_provinces = [{'province': r['province'], 'count': r['cnt']} for r in cursor.fetchall()]

        # Recent reservations
        cursor.execute("""
            SELECT province, branch_name, octet2, octet3, username, reservation_date, status
            FROM reserved_ips ORDER BY reservation_date DESC LIMIT 15
        """)
        recent_res = [dict(r) for r in cursor.fetchall()]

        # Expiring soon
        cursor.execute("""
            SELECT province, branch_name, octet2, octet3, expiry_date, username
            FROM reserved_ips WHERE status = 'reserved' AND expiry_date <= date('now', '+7 days')
            ORDER BY expiry_date ASC LIMIT 10
        """)
        expiring = [dict(r) for r in cursor.fetchall()]

        conn.close()

        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        username = request.args.get('user', 'System')

        # Build HTML for PDF (will be converted client-side via browser print)
        html = f"""<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
<meta charset="UTF-8">
<title>گزارش وضعیت شبکه - {now}</title>
<style>
    @page {{ size: A4; margin: 15mm; }}
    * {{ margin: 0; padding: 0; box-sizing: border-box; font-family: 'Segoe UI', Tahoma, sans-serif; }}
    body {{ padding: 20px; color: #1e293b; font-size: 12px; direction: rtl; }}
    .header {{ text-align: center; border-bottom: 3px solid #1e40af; padding-bottom: 15px; margin-bottom: 20px; }}
    .header h1 {{ color: #1e40af; font-size: 20px; margin-bottom: 4px; }}
    .header p {{ color: #64748b; font-size: 11px; }}
    .stats-grid {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 12px; margin-bottom: 20px; }}
    .stat-box {{ background: #f1f5f9; border: 1px solid #e2e8f0; border-radius: 8px; padding: 12px; text-align: center; }}
    .stat-box .val {{ font-size: 22px; font-weight: 700; color: #1e40af; }}
    .stat-box .lbl {{ font-size: 10px; color: #64748b; margin-top: 2px; }}
    .stat-box .sub {{ display: flex; justify-content: center; gap: 12px; margin-top: 6px; font-size: 10px; }}
    .stat-box .sub .free {{ color: #059669; }} .stat-box .sub .used {{ color: #dc2626; }}
    .section {{ margin-bottom: 18px; }}
    .section h2 {{ font-size: 14px; color: #1e40af; border-bottom: 2px solid #e2e8f0; padding-bottom: 6px; margin-bottom: 10px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 11px; }}
    th {{ background: #1e40af; color: white; padding: 6px 8px; text-align: right; }}
    td {{ padding: 5px 8px; border-bottom: 1px solid #e2e8f0; }}
    tr:nth-child(even) {{ background: #f8fafc; }}
    .bar-chart {{ display: flex; flex-direction: column; gap: 6px; }}
    .bar-row {{ display: flex; align-items: center; gap: 8px; }}
    .bar-name {{ width: 100px; font-size: 11px; text-align: left; }}
    .bar-track {{ flex: 1; height: 16px; background: #e2e8f0; border-radius: 4px; overflow: hidden; }}
    .bar-fill {{ height: 100%; border-radius: 4px; background: linear-gradient(90deg, #3b82f6, #1e40af); }}
    .bar-val {{ width: 40px; font-size: 10px; color: #64748b; }}
    .footer {{ text-align: center; color: #94a3b8; font-size: 10px; margin-top: 20px; border-top: 1px solid #e2e8f0; padding-top: 10px; }}
    .alert {{ background: #fef2f2; border: 1px solid #fecaca; border-radius: 6px; padding: 8px 12px; margin-bottom: 12px; color: #dc2626; font-size: 11px; }}
</style>
</head>
<body>
<div class="header">
    <h1>گزارش وضعیت شبکه - بانک کشاورزی</h1>
    <p>تاریخ تهیه: {now} | تهیه کننده: {username} | Network Configuration Portal</p>
</div>

<div class="stats-grid">
    <div class="stat-box">
        <div class="val">{s['total_lan']}</div>
        <div class="lbl">کل IP LAN</div>
        <div class="sub"><span class="free">آزاد: {s['free_lan']}</span> <span class="used">مصرفی: {s['total_lan'] - s['free_lan']}</span></div>
    </div>
    <div class="stat-box">
        <div class="val">{s['total_tun']}</div>
        <div class="lbl">Tunnel Intranet</div>
        <div class="sub"><span class="free">آزاد: {s['free_tun']}</span> <span class="used">مصرفی: {s['total_tun'] - s['free_tun']}</span></div>
    </div>
    <div class="stat-box">
        <div class="val">{s['total_apn']}</div>
        <div class="lbl">APN غیرمالی</div>
        <div class="sub"><span class="free">آزاد: {s['free_apn']}</span> <span class="used">مصرفی: {s['total_apn'] - s['free_apn']}</span></div>
    </div>
    <div class="stat-box">
        <div class="val">{s['total_mali']}</div>
        <div class="lbl">APN مالی</div>
        <div class="sub"><span class="free">آزاد: {s['free_mali']}</span> <span class="used">مصرفی: {s['total_mali'] - s['free_mali']}</span></div>
    </div>
    <div class="stat-box">
        <div class="val">{s['total_ptmp']}</div>
        <div class="lbl">PTMP سریال</div>
        <div class="sub"><span class="used">فعال: {s['total_ptmp']}</span> <span class="free">تطبیق نام: {s['matched_ptmp']}</span></div>
    </div>
</div>"""

        # Expiring warning
        if expiring:
            html += f'<div class="alert">⚠️ {len(expiring)} رزرو در ۷ روز آینده منقضی می‌شوند!</div>'

        # Top provinces bar chart
        html += '<div class="section"><h2>برترین استان‌ها (بر اساس IP فعال)</h2><div class="bar-chart">'
        if top_provinces:
            max_c = top_provinces[0]['count']
            for p in top_provinces:
                pct = int((p['count'] / max_c) * 100) if max_c else 0
                html += f'<div class="bar-row"><div class="bar-name">{p["province"]}</div><div class="bar-track"><div class="bar-fill" style="width:{pct}%"></div></div><div class="bar-val">{p["count"]}</div></div>'
        html += '</div></div>'

        # Recent reservations table
        html += '<div class="section"><h2>آخرین رزروها</h2><table><tr><th>IP</th><th>استان</th><th>شعبه</th><th>کاربر</th><th>تاریخ</th><th>وضعیت</th></tr>'
        for r in recent_res:
            ip = f"10.{r['octet2']}.{r['octet3']}.0/24"
            html += f"<tr><td>{ip}</td><td>{r['province'] or ''}</td><td>{r['branch_name'] or ''}</td><td>{r['username'] or ''}</td><td>{r['reservation_date'] or ''}</td><td>{r['status'] or ''}</td></tr>"
        html += '</table></div>'

        # Expiring table
        if expiring:
            html += '<div class="section"><h2>رزروهای در حال انقضا (۷ روز آینده)</h2><table><tr><th>IP</th><th>استان</th><th>شعبه</th><th>تاریخ انقضا</th><th>کاربر</th></tr>'
            for r in expiring:
                ip = f"10.{r['octet2']}.{r['octet3']}.0/24"
                html += f"<tr><td>{ip}</td><td>{r['province'] or ''}</td><td>{r['branch_name'] or ''}</td><td>{r['expiry_date'] or ''}</td><td>{r['username'] or ''}</td></tr>"
            html += '</table></div>'

        # Summary
        html += f"""
<div class="section"><h2>خلاصه</h2>
<table>
<tr><td><strong>رزروهای فعال</strong></td><td>{s['active_reservations']}</td></tr>
<tr><td><strong>درصد مصرف LAN</strong></td><td>{int(((s['total_lan']-s['free_lan'])/s['total_lan'])*100) if s['total_lan'] else 0}%</td></tr>
<tr><td><strong>درصد مصرف Tunnel</strong></td><td>{int(((s['total_tun']-s['free_tun'])/s['total_tun'])*100) if s['total_tun'] else 0}%</td></tr>
<tr><td><strong>PTMP سریال (کل)</strong></td><td>{s['total_ptmp']}</td></tr>
</table>
</div>

<div class="footer">
    Network Configuration Portal - Keshavarzi Bank - گزارش خودکار {now}
</div>
</body></html>"""

        log_activity('success', 'تهیه گزارش PDF', f'گزارش وضعیت شبکه تهیه شد', username)
        return Response(html, mimetype='text/html')

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== PING API ====================
@app.route('/api/ping', methods=['POST'])
def ping_host():
    """Ping a host and return reachability + latency"""
    data = request.json
    if not data or not data.get('host'):
        return jsonify({'error': 'Host is required'}), 400

    host = data['host'].strip()

    # Validate: only allow IPs and hostnames (prevent command injection)
    import re
    if not re.match(r'^[a-zA-Z0-9._:-]+$', host):
        return jsonify({'error': 'Invalid host format'}), 400

    try:
        count_flag = '-n' if platform.system().lower() == 'windows' else '-c'
        timeout_flag = '-w' if platform.system().lower() == 'windows' else '-W'
        timeout_val = '3000' if platform.system().lower() == 'windows' else '3'

        result = subprocess.run(
            ['ping', count_flag, '3', timeout_flag, timeout_val, host],
            capture_output=True, text=True, timeout=15
        )

        output = result.stdout + result.stderr
        reachable = result.returncode == 0

        # Extract average latency
        avg_ms = None
        if reachable:
            avg_match = re.search(r'Average\s*=\s*(\d+)', output)
            if not avg_match:
                avg_match = re.search(r'avg[^=]*=\s*[\d.]+/([\d.]+)', output)
            if avg_match:
                avg_ms = float(avg_match.group(1))

        return jsonify({
            'reachable': reachable,
            'host': host,
            'avg_ms': avg_ms,
            'output': output.strip()
        })
    except subprocess.TimeoutExpired:
        return jsonify({'reachable': False, 'host': host, 'avg_ms': None, 'output': 'Ping timed out'})
    except Exception as e:
        return jsonify({'reachable': False, 'host': host, 'avg_ms': None, 'output': str(e)})

# ==================== SHARED FILES API ====================
SHARED_FILES_DIR = os.path.join(os.path.dirname(__file__), 'data', 'shared_files')
os.makedirs(SHARED_FILES_DIR, exist_ok=True)
SHARED_MAX_SIZE = 100 * 1024 * 1024  # 100MB
SHARED_ALLOWED_EXT = {
    'pdf','doc','docx','xls','xlsx','ppt','pptx','txt','csv',
    'zip','rar','7z','tar','gz','png','jpg','jpeg','gif','bmp','svg',
    'bin','ios','img','conf','cfg','log','py','sh','bat','json','xml','yaml','yml'
}

@app.route('/api/shared-files', methods=['GET'])
def list_shared_files():
    try:
        files = []
        for f in os.listdir(SHARED_FILES_DIR):
            fpath = os.path.join(SHARED_FILES_DIR, f)
            if os.path.isfile(fpath):
                stat = os.stat(fpath)
                ext = f.rsplit('.', 1)[-1].lower() if '.' in f else ''
                files.append({'name': f, 'size': stat.st_size, 'ext': ext,
                    'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                    'uploaded': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S')})
        files.sort(key=lambda x: x['modified'], reverse=True)
        return jsonify({'status': 'ok', 'files': files})
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/shared-files/upload', methods=['POST'])
def upload_shared_file():
    try:
        from werkzeug.utils import secure_filename
        username = request.form.get('username', 'unknown')
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'error': 'No file selected'}), 400
        file = request.files['file']
        if not file.filename:
            return jsonify({'status': 'error', 'error': 'No file selected'}), 400
        ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        if ext not in SHARED_ALLOWED_EXT:
            return jsonify({'status': 'error', 'error': f'File type .{ext} not allowed'}), 400
        file.seek(0, 2)
        size = file.tell()
        file.seek(0)
        if size > SHARED_MAX_SIZE:
            return jsonify({'status': 'error', 'error': 'File too large (max 100MB)'}), 400
        filename = secure_filename(file.filename) or f'file_{int(time.time())}.{ext}'
        filepath = os.path.join(SHARED_FILES_DIR, filename)
        if os.path.exists(filepath):
            name_part = filename.rsplit('.', 1)[0] if '.' in filename else filename
            ext_part = filename.rsplit('.', 1)[1] if '.' in filename else ''
            filename = f"{name_part}_{int(time.time())}.{ext_part}" if ext_part else f"{name_part}_{int(time.time())}"
            filepath = os.path.join(SHARED_FILES_DIR, filename)
        file.save(filepath)
        stat = os.stat(filepath)
        return jsonify({'status': 'ok', 'message': f'File {filename} uploaded',
            'file': {'name': filename, 'size': stat.st_size, 'ext': ext,
                     'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')}})
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/shared-files/download/<filename>', methods=['GET'])
def download_shared_file(filename):
    try:
        from werkzeug.utils import secure_filename
        safe_name = secure_filename(filename)
        return send_from_directory(SHARED_FILES_DIR, safe_name, as_attachment=True)
    except FileNotFoundError:
        return jsonify({'status': 'error', 'error': 'File not found'}), 404

@app.route('/api/shared-files/delete', methods=['POST'])
def delete_shared_file():
    try:
        from werkzeug.utils import secure_filename
        data = request.json or {}
        filename = data.get('filename', '')
        if not filename:
            return jsonify({'status': 'error', 'error': 'No filename'}), 400
        safe_name = secure_filename(filename)
        filepath = os.path.join(SHARED_FILES_DIR, safe_name)
        if not os.path.exists(filepath):
            return jsonify({'status': 'error', 'error': 'File not found'}), 404
        os.remove(filepath)
        return jsonify({'status': 'ok', 'message': f'File {safe_name} deleted'})
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500


# ==================== REPORTS API ====================
PROVINCE_EN_TO_FA = {
    'East Azerbaijan': 'آذربایجان شرقی', 'West Azerbaijan': 'آذربایجان غربی',
    'Ardabil': 'اردبیل', 'Isfahan': 'اصفهان', 'Alborz': 'البرز', 'Ilam': 'ایلام',
    'Bushehr': 'بوشهر', 'Tehran': 'تهران', 'South Khorasan': 'خراسان جنوبی',
    'Razavi Khorasan': 'خراسان رضوی', 'North Khorasan': 'خراسان شمالی',
    'Khuzestan': 'خوزستان', 'Zanjan': 'زنجان', 'Semnan': 'سمنان',
    'Sistan and Baluchestan': 'سیستان و بلوچستان', 'Fars': 'فارس',
    'Qazvin': 'قزوین', 'Qom': 'قم', 'Lorestan': 'لرستان',
    'Mazandaran': 'مازندران', 'Markazi': 'مرکزی', 'Hormozgan': 'هرمزگان',
    'Hamadan': 'همدان', 'Chaharmahal and Bakhtiari': 'چهارمحال و بختیاری',
    'Kurdistan': 'کردستان', 'Kerman': 'کرمان', 'Kermanshah': 'کرمانشاه',
    'Kohgiluyeh and Boyer-Ahmad': 'کهگیلویه و بویراحمد', 'Golestan': 'گلستان',
    'Gilan': 'گیلان', 'Yazd': 'یزد',
}
PROVINCE_FA_TO_EN = {v: k for k, v in PROVINCE_EN_TO_FA.items()}

def _detect_point_type(name):
    if not name: return 'نامشخص'
    nl = name.lower()
    if 'atm' in nl or 'خودپرداز' in nl: return 'ATM'
    if 'kiosk' in nl or 'کیوسک' in nl or 'cashless' in nl: return 'کیوسک'
    if 'bj' in nl or 'bajeh' in nl or 'باجه' in nl: return 'باجه'
    if '24' in nl and ('ساعته' in nl or 'saate' in nl): return '24 ساعته'
    if 'vsat' in nl: return 'VSAT'
    return 'شعبه'

@app.route('/api/reports/provinces', methods=['GET'])
def report_provinces():
    provinces = set()
    try:
        conn = get_db()
        cursor = conn.cursor()
        for tbl, col in [('lan_ips','province'),('vpls_tunnels','province'),('apn_mali','province'),
                         ('apn_ips','province'),('intranet_tunnels','province'),('ptmp_connections','province')]:
            try:
                cursor.execute(f"SELECT DISTINCT {col} FROM {tbl} WHERE {col} IS NOT NULL AND {col} != ''")
                for r in cursor.fetchall():
                    provinces.add(PROVINCE_EN_TO_FA.get(r[0], r[0]))
            except: pass
        conn.close()
        return jsonify(sorted(provinces))
    except: return jsonify([])

@app.route('/api/reports/query', methods=['GET'])
def report_query():
    province = request.args.get('province', '').strip()
    service_type = request.args.get('service_type', '').strip()
    point_type = request.args.get('point_type', '').strip()
    province_en = PROVINCE_FA_TO_EN.get(province, province)
    results = []
    try:
        conn = get_db()
        cursor = conn.cursor()
        if service_type in ('', 'all', 'MPLS', 'VPLS', 'MPLS/VPLS'):
            sql = "SELECT branch_name, description, province, ip_address, wan_ip, tunnel_dest, tunnel_name, username, reservation_date, status FROM vpls_tunnels WHERE LOWER(status) IN ('reserved','used')"
            params = []
            if province_en:
                sql += " AND province = ?"
                params.append(province_en)
            cursor.execute(sql, params)
            for r in cursor.fetchall():
                name = r[0] or (r[1] or '').replace('** ','').replace(' **','').strip()
                pt = _detect_point_type(name)
                if point_type and pt != point_type: continue
                results.append({'service':'MPLS/VPLS','branch_name':name,'province':PROVINCE_EN_TO_FA.get(r[2],r[2] or ''),'point_type':pt,'ip':r[3] or '','wan_ip':r[4] or '','tunnel_dest':r[5] or '','tunnel_name':r[6] or '','username':r[7] or '','date':r[8] or '','status':r[9] or ''})

        if service_type in ('', 'all', 'Intranet'):
            sql = "SELECT tunnel_name, description, province, ip_address, ip_lan, reserved_by, reserved_at, status, branch_name FROM intranet_tunnels WHERE LOWER(status) = 'reserved'"
            params = []
            if province:
                sql += " AND (province = ? OR province = ?)"
                params.extend([province, province_en])
            cursor.execute(sql, params)
            for r in cursor.fetchall():
                name = (r[8] or r[1] or r[0] or '').replace('** ','').replace(' **','').strip()
                pt = _detect_point_type(name)
                if point_type and pt != point_type: continue
                results.append({'service':'Intranet','branch_name':name,'province':r[2] or '','point_type':pt,'ip':r[3] or '','wan_ip':r[4] or '','tunnel_dest':'','tunnel_name':r[0] or '','username':r[5] or '','date':r[6] or '','status':r[7] or ''})

        if service_type in ('', 'all', 'APN'):
            sql = "SELECT branch_name, province, ip_wan, lan_ip, username, reservation_date FROM apn_mali WHERE branch_name IS NOT NULL AND branch_name != ''"
            params = []
            if province:
                sql += " AND province = ?"
                params.append(province)
            cursor.execute(sql, params)
            for r in cursor.fetchall():
                pt = _detect_point_type(r[0])
                if point_type and pt != point_type: continue
                results.append({'service':'APN Mali','branch_name':r[0],'province':r[1] or '','point_type':pt,'ip':r[2] or '','wan_ip':r[3] or '','tunnel_dest':'','tunnel_name':'','username':r[4] or '','date':r[5] or '','status':'Active'})
            sql2 = "SELECT branch_name, province, ip_wan_apn, lan_ip, username, reservation_date FROM apn_ips WHERE branch_name IS NOT NULL AND branch_name != ''"
            params2 = []
            if province:
                sql2 += " AND province = ?"
                params2.append(province)
            cursor.execute(sql2, params2)
            for r in cursor.fetchall():
                pt = _detect_point_type(r[0])
                if point_type and pt != point_type: continue
                results.append({'service':'APN INT','branch_name':r[0],'province':r[1] or '','point_type':pt,'ip':r[2] or '','wan_ip':r[3] or '','tunnel_dest':'','tunnel_name':'','username':r[4] or '','date':r[5] or '','status':'Active'})

        if service_type in ('', 'all', 'PTMP'):
            sql = "SELECT branch_name, branch_name_en, province, interface_name, lan_ip, description, username, reservation_date, status FROM ptmp_connections WHERE (branch_name IS NOT NULL OR branch_name_en IS NOT NULL)"
            params = []
            if province:
                sql += " AND (province = ? OR province = ?)"
                params.extend([province, province_en])
            cursor.execute(sql, params)
            for r in cursor.fetchall():
                name = r[0] or r[1] or ''
                pt = _detect_point_type(name)
                if point_type and pt != point_type: continue
                results.append({'service':'PTMP','branch_name':name,'province':r[2] or '','point_type':pt,'ip':r[3] or '','wan_ip':r[4] or '','tunnel_dest':'','tunnel_name':r[5] or '','username':r[6] or '','date':r[7] or '','status':r[8] or ''})

        conn.close()
        return jsonify({'status':'ok','count':len(results),'results':results})
    except Exception as e:
        return jsonify({'status':'error','error':str(e),'results':[]}), 500

@app.route('/api/reports/export/excel', methods=['GET'])
def report_export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io
        resp = report_query()
        data = resp.get_json()
        rows = data.get('results', [])
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Report'
        ws.sheet_view.rightToLeft = True
        headers = ['ردیف','نوع سرویس','استان','نام نقطه','نوع نقطه','IP','WAN IP','Tunnel','کاربر','تاریخ']
        hfill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        hfont = Font(bold=True, color='FFFFFF', size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hfill; c.font = hfont; c.alignment = Alignment(horizontal='center'); c.border = border
        for idx, r in enumerate(rows, 1):
            for col, val in enumerate([idx, r['service'], r['province'], r['branch_name'], r['point_type'], r['ip'], r['wan_ip'], r['tunnel_name'], r['username'], r['date']], 1):
                c = ws.cell(row=idx+1, column=col, value=val)
                c.border = border; c.alignment = Alignment(horizontal='center' if col<=2 else 'right')
        for col in ws.columns:
            mx = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 40)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        province = request.args.get('province', 'all')
        svc = request.args.get('service_type', 'all')
        fname = f'report_{province}_{svc}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        return send_file(output, as_attachment=True, download_name=fname, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'status':'error','error':str(e)}), 500


# ==================== NETWORK MAP API ====================
import re as _re
import ipaddress as _ipaddr

PROVINCE_MAP_INFO = {
    'AZSH':{'fa':'آذربایجان شرقی','x':22,'y':10},'AZGH':{'fa':'آذربایجان غربی','x':14,'y':14},
    'ARD':{'fa':'اردبیل','x':28,'y':5},'ESF':{'fa':'اصفهان','x':48,'y':52},
    'ALZ':{'fa':'البرز','x':40,'y':22},'ILM':{'fa':'ایلام','x':16,'y':48},
    'BSH':{'fa':'بوشهر','x':42,'y':76},'M1-Tehran':{'fa':'تهران ۱','x':43,'y':28},
    'M2-Tehran':{'fa':'تهران ۲','x':47,'y':28},'OSTehran':{'fa':'استان تهران','x':45,'y':25},
    'KHRJ':{'fa':'خراسان جنوبی','x':78,'y':52},'KHR':{'fa':'خراسان رضوی','x':76,'y':34},
    'KhShomali':{'fa':'خراسان شمالی','x':72,'y':22},'KHZ':{'fa':'خوزستان','x':28,'y':60},
    'ZNJ':{'fa':'زنجان','x':28,'y':18},'SMN':{'fa':'سمنان','x':58,'y':24},
    'SNB':{'fa':'سیستان و بلوچستان','x':86,'y':70},'FRS':{'fa':'فارس','x':48,'y':70},
    'QZV':{'fa':'قزوین','x':34,'y':22},'QOM':{'fa':'قم','x':44,'y':38},
    'LOR':{'fa':'لرستان','x':26,'y':44},'MAZ':{'fa':'مازندران','x':50,'y':16},
    'MRZ':{'fa':'مرکزی','x':36,'y':40},'HMZ':{'fa':'هرمزگان','x':58,'y':82},
    'HMD':{'fa':'همدان','x':28,'y':34},'CHB':{'fa':'چهارمحال و بختیاری','x':38,'y':56},
    'KRD':{'fa':'کردستان','x':18,'y':28},'KRM':{'fa':'کرمان','x':66,'y':64},
    'KRMJ':{'fa':'کرمانشاه','x':18,'y':38},'KNB':{'fa':'کهگیلویه و بویراحمد','x':36,'y':64},
    'GLS':{'fa':'گلستان','x':62,'y':14},'GIL':{'fa':'گیلان','x':34,'y':12},
    'YZD':{'fa':'یزد','x':58,'y':54},'KRSH':{'fa':'خراسان رضوی','x':76,'y':34},
    'NIBN':{'fa':'مرکز داده','x':44,'y':28},
}

# Province abbreviation aliases for core switches
_SWITCH_PROV_MAP = {
    'ARD':'ARD','AZSH':'AZSH','AZSh':'AZSH','AZGH':'AZGH','AzGh':'AZGH',
    'ESF':'ESF','ALZ':'ALZ','BSH':'BSH','CHB':'CHB','FRS':'FRS',
    'GIL':'GIL','GLS':'GLS','HMD':'HMD','HMZ':'HMZ','ILM':'ILM',
    'KHR':'KHR','KHRJ':'KHRJ','KHSh':'KhShomali','KHSH':'KhShomali',
    'KHZ':'KHZ','KNB':'KNB','KRD':'KRD','KRM':'KRM','KRMJ':'KRMJ',
    'KRSH':'KRSH','LOR':'LOR','Maz':'MAZ','MAZ':'MAZ','MRZ':'MRZ',
    'QOM':'QOM','QZV':'QZV','SMN':'SMN','SNB':'SNB','YZD':'YZD','ZNJ':'ZNJ',
    'Teh':'OSTehran','TEHB':'OSTehran','TEH':'OSTehran',
}

def _parse_router_config_v2(filepath):
    """Enhanced parser - extracts full detail for schematic visualization"""
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
    except: return None

    info = {
        'hostname':'','interfaces':[],'tunnels':[],'nat_rules':[],'nat_interfaces':[],
        'ospf_processes':[],'static_routes':[],'access_lists':{},'crypto_maps':[]
    }

    m = _re.search(r'^hostname\s+(.+)', content, _re.MULTILINE)
    if m: info['hostname'] = m.group(1).strip()

    # Parse ALL interfaces
    for m in _re.finditer(r'^interface\s+(\S+)\s*\n((?:.*\n)*?)(?=^interface\s|^!\s*$|\Z)', content, _re.MULTILINE):
        iname, iblock = m.group(1), m.group(2)
        ips = _re.findall(r'ip address\s+(\S+)\s+(\S+)', iblock)
        desc = _re.search(r'description\s+(.+)', iblock)
        shut = bool(_re.search(r'^\s*shutdown', iblock, _re.MULTILINE))
        nat_side = ''
        if _re.search(r'ip nat inside', iblock): nat_side = 'inside'
        elif _re.search(r'ip nat outside', iblock): nat_side = 'outside'
        crypto = _re.search(r'crypto map\s+(\S+)', iblock)
        policy_route = _re.search(r'ip policy route-map\s+(\S+)', iblock)
        if nat_side:
            info['nat_interfaces'].append({'name': iname, 'side': nat_side})
        if ips:
            for ip, mask in ips:
                entry = {'name':iname,'ip':ip,'mask':mask,'shutdown':shut,'nat':nat_side}
                if desc: entry['description'] = desc.group(1).strip()
                if crypto: entry['crypto_map'] = crypto.group(1)
                if policy_route: entry['policy_route'] = policy_route.group(1)
                if 'Tunnel' in iname:
                    src = _re.search(r'tunnel source\s+(\S+)', iblock)
                    dst = _re.search(r'tunnel destination\s+(\S+)', iblock)
                    mode = _re.search(r'tunnel mode\s+(.+)', iblock)
                    cost = _re.search(r'ip ospf cost\s+(\d+)', iblock)
                    if src: entry['tunnel_src'] = src.group(1)
                    if dst: entry['tunnel_dst'] = dst.group(1)
                    if mode: entry['tunnel_mode'] = mode.group(1).strip()
                    if cost: entry['ospf_cost'] = int(cost.group(1))
                    info['tunnels'].append(entry)
                else:
                    info['interfaces'].append(entry)
        elif desc and not ips:
            entry = {'name':iname,'ip':'','mask':'','shutdown':shut,'nat':nat_side}
            entry['description'] = desc.group(1).strip()
            info['interfaces'].append(entry)

    # NAT rules
    for m in _re.finditer(r'^ip nat inside source\s+(.+)', content, _re.MULTILINE):
        rule_text = m.group(1).strip()
        rule = {'raw': 'ip nat inside source ' + rule_text, 'type': 'dynamic'}
        if rule_text.startswith('static'):
            rule['type'] = 'static'
            parts = rule_text.split()
            if len(parts) >= 3:
                rule['inside_ip'] = parts[1]
                rule['outside_ip'] = parts[2]
        else:
            lm = _re.search(r'list\s+(\S+)', rule_text)
            pm = _re.search(r'pool\s+(\S+)', rule_text)
            im = _re.search(r'interface\s+(\S+)', rule_text)
            if lm: rule['acl'] = lm.group(1)
            if pm: rule['pool'] = pm.group(1)
            if im: rule['interface'] = im.group(1)
            rule['overload'] = 'overload' in rule_text
        info['nat_rules'].append(rule)
    for m in _re.finditer(r'^ip nat pool\s+(\S+)\s+(\S+)\s+(\S+)\s+', content, _re.MULTILINE):
        info['nat_rules'].append({'type':'pool','name':m.group(1),'start':m.group(2),'end':m.group(3),'raw':'pool '+m.group(1)+' '+m.group(2)+'-'+m.group(3)})

    # OSPF
    for m in _re.finditer(r'^router ospf\s+(\d+)\s*\n((?:.*\n)*?)(?=^router\s|^!\s*$)', content, _re.MULTILINE):
        ospf_block = m.group(2)
        nets = _re.findall(r'network\s+(\S+)\s+(\S+)\s+area\s+(\S+)', ospf_block)
        rid = _re.search(r'router-id\s+(\S+)', ospf_block)
        redist = _re.findall(r'redistribute\s+(.+)', ospf_block)
        info['ospf_processes'].append({
            'process':m.group(1),
            'router_id': rid.group(1) if rid else '',
            'networks':[{'net':n,'wildcard':w,'area':a} for n,w,a in nets],
            'redistribute': [r.strip() for r in redist]
        })

    # Static routes - ALL
    for m in _re.finditer(r'^ip route\s+(\S+)\s+(\S+)\s+(.+)', content, _re.MULTILINE):
        dest, mask, rest = m.group(1), m.group(2), m.group(3).strip()
        name_m = _re.search(r'name\s+(\S+)', rest)
        nh_parts = rest.split()
        next_hop = nh_parts[0] if nh_parts else ''
        info['static_routes'].append({
            'dest': dest, 'mask': mask, 'next_hop': next_hop,
            'name': name_m.group(1) if name_m else ''
        })

    # Access lists
    for m in _re.finditer(r'^ip access-list (?:extended|standard)\s+(\S+)\s*\n((?:.*\n)*?)(?=^ip access-list|^!\s*$|\Z)', content, _re.MULTILINE):
        acl_name = m.group(1)
        entries = _re.findall(r'^\s+(?:\d+\s+)?(permit|deny)\s+(.+)', m.group(2), _re.MULTILINE)
        info['access_lists'][acl_name] = [{'action':a,'rule':r.strip()} for a,r in entries[:20]]
    for m in _re.finditer(r'^access-list\s+(\d+)\s+(permit|deny)\s+(.+)', content, _re.MULTILINE):
        num = m.group(1)
        if num not in info['access_lists']: info['access_lists'][num] = []
        info['access_lists'][num].append({'action':m.group(2),'rule':m.group(3).strip()})

    # Crypto maps
    for m in _re.finditer(r'^crypto map\s+(\S+)\s+(\d+)\s+', content, _re.MULTILINE):
        info['crypto_maps'].append(m.group(1))
    info['crypto_maps'] = list(set(info['crypto_maps']))

    return info

def _get_device_category(hostname, fname, subdir):
    hn = hostname.upper()
    fn = fname.upper()
    if subdir == 'Core Switches': return 'core-switch', 'Switch'
    if subdir == 'Core Routers':
        if 'ASR1006' in fn: return 'core-router', 'ASR1006'
        if 'SW' in fn[:3] or 'AGG' in fn[:3]: return 'core-router', 'Switch'
        return 'core-router', 'ISR4451'
    if 'ASR1002' in fn: return 'provincial-router', 'ASR1002X'
    if '3845' in fn: return 'provincial-router', '3845'
    if '3825' in fn: return 'provincial-router', '3825'
    if 'Mo-' in fn: return 'provincial-router', '3825'
    return 'other', fn.split('-')[0]

def _extract_switch_province(hostname):
    """Extract province from core switch hostname like SW3560X-ESF, SW3650-LOR etc."""
    m = _re.search(r'SW\w*-(\w+)', hostname)
    if m:
        raw = m.group(1)
        return _SWITCH_PROV_MAP.get(raw, raw)
    if 'Core-SW' in hostname or 'CORESW' in hostname:
        m = _re.search(r'(?:Core-?SW-?)(\w+)', hostname)
        if m: return _SWITCH_PROV_MAP.get(m.group(1), m.group(1))
    return None

def _extract_province_abbr(hostname):
    parts = hostname.split('-')
    if len(parts) < 2: return hostname
    abbr = parts[1]
    if len(parts) >= 3 and abbr in ('M1','M2','OS','Mo'):
        abbr = parts[1] + '-' + parts[2]
    return abbr

def _subnet_match(ip1, mask1, ip2, mask2):
    """Check if two IPs are on the same subnet"""
    try:
        if not ip1 or not ip2 or not mask1 or not mask2: return False
        net1 = _ipaddr.IPv4Network(f'{ip1}/{mask1}', strict=False)
        net2 = _ipaddr.IPv4Network(f'{ip2}/{mask2}', strict=False)
        return net1 == net2
    except: return False

# ── Branch / endpoint parsing constants ─────────────────────────────────────
_PROV_CANONICAL_BR = {
    'ISF': 'ESF', 'TEHB': 'TehB', 'KhRzv': 'KHR', 'Hmd': 'HMD', 'Fars': 'FRS',
    'M1': 'TehB', 'M2': 'TehB',
}
_PROV_CODES_BR = sorted([
    'AZSH', 'AZGH', 'ALZ', 'ARD', 'ESF', 'ISF', 'BSH', 'TEH', 'TehB', 'TEHB',
    'KHRJ', 'KHR', 'KhRzv', 'KHSH', 'KHZ', 'ZNJ', 'SMN', 'SNB', 'FRS', 'Fars',
    'QOM', 'KRD', 'KRM', 'JKRM', 'KRSH', 'CHB', 'KNB', 'GIL', 'LOR', 'MAZ',
    'MRZ', 'HMD', 'Hmd', 'HMZ', 'YZD', 'QZV', 'GLS', 'ILM',
], key=len, reverse=True)

def _get_branch_prov_br(desc):
    for code in _PROV_CODES_BR:
        if desc.startswith(code + '-'):
            return _PROV_CANONICAL_BR.get(code, code)
    return None

def _branch_type_br(desc):
    dl = desc.lower()
    if 'kiosk' in dl:
        return 'kiosk'
    if 'ATM' in desc or ('-atm-' in dl and 'atm' in dl):
        return 'atm'
    if '-Bj-' in desc or '-BJ-' in desc or '-BJH-' in desc or 'NBP' in desc:
        return 'baje'
    return 'branch'

def _parse_branches_by_province(router_dir):
    """
    Parse ALL router configs (WAN-INTR1/2, ISR-APN-RO, all provincial mgmt routers)
    to extract every branch, bajeh, ATM, kiosk grouped by province code.
    """
    import subprocess as _sp
    result = {}

    def _add(prov, name, ntype):
        if prov not in result:
            result[prov] = {'branches': [], 'bajes': [], 'atms': [], 'kiosks': []}
        key = {'branch': 'branches', 'baje': 'bajes', 'atm': 'atms', 'kiosk': 'kiosks'}.get(ntype, 'branches')
        lst = result[prov][key]
        if name not in lst:
            lst.append(name)

    def _read_strings(fpath):
        try:
            return _sp.run(['strings', fpath], capture_output=True, text=True, errors='replace').stdout
        except Exception:
            try:
                with open(fpath, 'r', encoding='utf-8', errors='ignore') as f:
                    return f.read()
            except Exception:
                return ''

    # Infrastructure keywords – descriptions containing these are skipped
    _infra_kw = frozenset([
        'MPLS', 'ShahinShahr', 'Intranet', 'Madar', 'PTMP', 'PDH', 'OStani',
        'AGG', 'Saatzani', 'EDGE', 'Extranet', 'FTP-MGRE',
        'GILANET', 'HIWEB', 'MOBINNET', 'LAN', 'APN', 'IranSolar',
        'Amuzesh', 'Amouzesh', 'Amouz', 'Jahad', 'Modiriyat',
        'Zirsakht', 'Backbone', 'SW-Agg', 'Iransolar', 'ETHERNET-Bashgah',
    ])
    _skip_names = frozenset(['MO', 'HQ', 'Jahad', 'Modiriyat', 'HQ-Tohid', 'Terminal',
                             'Gilanet', 'ShahinShahr', 'HIWEB', 'MOBINNET'])
    _provider_pfx = ('Gilanet-', 'MOBINNET-', 'HIWEB-', 'gilanet-', 'mobinnet-',
                     'hiweb-', 'Shutdown-', 'SHUTDOWN-')

    def _is_infra(name):
        if not name or len(name) < 2:
            return True
        if name in _skip_names:
            return True
        if any(kw in name for kw in _infra_kw):
            return True
        if '->' in name or '*E1' in name:
            return True
        if name.isdigit():
            return True
        return False

    def _process_desc(desc, default_prov=None):
        """Process one tunnel description; add to result if it's a real endpoint."""
        name = desc.strip()
        if not name:
            return
        # Strip ISP/provider prefix first
        for pfx in _provider_pfx:
            if name.startswith(pfx):
                name = name[len(pfx):]
                break
        if not name:
            return
        # Try to find a province code prefix in the (possibly stripped) name
        prov = _get_branch_prov_br(name)
        if prov:
            branch_name = name.split('-', 1)[1] if '-' in name else name
        elif default_prov:
            prov = default_prov
            # Strip matching province prefix if present (e.g. "KHR-Gonabad" in KHR router)
            for code in _PROV_CODES_BR:
                if name.startswith(code + '-'):
                    name = name[len(code) + 1:]
                    break
            branch_name = name
        else:
            return
        if not branch_name:
            return
        prov = _PROV_CANONICAL_BR.get(prov, prov)
        if branch_name.endswith('-MO') or branch_name.endswith('-HQ'):
            return
        if _is_infra(branch_name):
            return
        _add(prov, branch_name, _branch_type_br(name))

    def _parse_tunnel_descs(fpath, default_prov=None):
        out = _read_strings(fpath)
        cur = None
        for line in out.splitlines():
            line = line.strip()
            if line.startswith('interface Tunnel'):
                cur = line
            elif 'description **' in line and cur:
                m = _re.search(r'\*\*\s*(.*?)\s*\*\*', line)
                if m:
                    _process_desc(m.group(1).strip(), default_prov=default_prov)

    core_dir = os.path.join(router_dir, 'Core Routers')

    # ── 1) Core WAN/APN routers ───────────────────────────────────────────────
    for prefix in ('WAN-INTR1', 'WAN-INTR2', 'ISR-APN-RO'):
        if not os.path.exists(core_dir):
            break
        for fn in sorted(os.listdir(core_dir)):
            if fn.startswith(prefix):
                _parse_tunnel_descs(os.path.join(core_dir, fn))
                break

    # ── 2) ALL provincial management routers (auto-detected by filename) ─────
    # Filename → province code mapping overrides for special cases
    _fname_prov_override = {
        'M1-Tehran': 'TehB', 'M2-Tehran': 'TehB', 'OSTehran': 'TehB',
        'KhShomali': 'KHSH',
    }
    _valid_provs = set(_PROV_CODES_BR) | set(_PROV_CANONICAL_BR.values()) | {
        'TehB', 'KHSH', 'KRMJ', 'JKRM',
    }

    if os.path.exists(router_dir):
        for fname in sorted(os.listdir(router_dir)):
            fpath = os.path.join(router_dir, fname)
            if not os.path.isfile(fpath) or os.path.getsize(fpath) < 100:
                continue
            prov_code = None
            # 3825-PROV-N  /  3845-PROV-N  (also handles M1-Tehran, M2-Tehran)
            m = _re.match(r'^\d+-([\w]+-?(?:Tehran)?)-\d', fname)
            if m:
                raw = m.group(1).rstrip('-')
                prov_code = _fname_prov_override.get(raw, raw)
            # ASR1002X-PROV-date...
            m2 = _re.match(r'^ASR1002X-([A-Z]+)-', fname)
            if m2:
                raw = m2.group(1)
                prov_code = _fname_prov_override.get(raw, raw)
            # Mo-KhShomali-N
            m3 = _re.match(r'^Mo-([\w]+)-', fname)
            if m3:
                raw = m3.group(1)
                prov_code = _fname_prov_override.get(raw, raw)
            if not prov_code or prov_code not in _valid_provs:
                continue
            prov_code = _PROV_CANONICAL_BR.get(prov_code, prov_code)
            _parse_tunnel_descs(fpath, default_prov=prov_code)

    return result

# Province abbr code → router hostname abbr mapping
_PROV_TO_ROUTER_ABBR_BR = {
    'AZSH': 'AZSH', 'AZGH': 'AZGH', 'ALZ': 'ALZ', 'ARD': 'ARD',
    'ESF': 'ESF', 'BSH': 'BSH', 'TEH': 'OSTehran', 'TehB': 'OSTehran',
    'KHRJ': 'KHRJ', 'KHR': 'KHR', 'KHSH': 'KhShomali', 'KHZ': 'KHZ',
    'ZNJ': 'ZNJ', 'SMN': 'SMN', 'SNB': 'SNB', 'FRS': 'FRS',
    'QOM': 'QOM', 'KRD': 'KRD', 'KRM': 'KRM', 'JKRM': 'KRMJ',
    'KRSH': 'KRMJ', 'CHB': 'CHB', 'KNB': 'KNB', 'GIL': 'GIL',
    'LOR': 'LOR', 'MAZ': 'MAZ', 'MRZ': 'MRZ', 'HMD': 'HMD',
    'HMZ': 'HMZ', 'YZD': 'YZD', 'QZV': 'QZV', 'GLS': 'GLS', 'ILM': 'ILM',
}

@app.route('/api/network-map/topology', methods=['GET'])
def network_map_topology():
    router_dir = os.path.join(os.path.dirname(__file__), 'Router')
    nodes, links = [], []
    if not os.path.exists(router_dir):
        return jsonify({'nodes':[],'links':[],'error':'Router directory not found'})

    # ── Core router positions: 4 tiers above Iran map (wider spacing) ──
    CORE_POSITIONS = {
        # Tier 1 (y=-35): Primary WAN backbone hubs - prominent, wide apart
        'WAN-INTR1':       (22, -35, 'WAN INTR1'),
        'ASR1006-WAN-MB':  (50, -35, 'ASR1006 WAN'),
        'WAN-INTR2':       (78, -35, 'WAN INTR2'),
        # Tier 2 (y=-25): Major service routers
        'ISR-APN-RO':      (14, -25, 'APN Router'),
        'APN-INT-HUB':     (30, -25, 'APN HUB'),
        'INT-4451':         (50, -25, 'Intranet'),
        'EXT-Edge-4451':    (70, -25, 'EXT Edge'),
        'PSP-4451':         (86, -25, 'PSP'),
        # Tier 3 (y=-16): Infrastructure & aggregation
        'BKC-4451':         (10, -16, 'BKC'),
        'AGG-WAN-SW':       (24, -16, 'AGG WAN'),
        'EXT-AGG':          (38, -16, 'EXT AGG'),
        'BKI-MAGFA':        (52, -16, 'MAGFA'),
        '3825-NIBN':        (66, -16, 'NIBN'),
        'Router-HTSC':      (78, -16, 'HTSC'),
        'SW-Roof-To-Site':  (90, -16, 'Roof-Site'),
        # Tier 4 (y=-8): Legacy & secondary devices
        '7206-STM1':        (8,  -8, 'STM1'),
        'V-Jahad-3825':     (22, -8, 'Jahad'),
        '4451-PBN':         (36, -8, 'PBN'),
        '2821-Gostaresh':   (50, -8, 'Gostaresh'),
        '2821-Mizan':       (64, -8, 'Mizan MCI'),
        '3825-Sabt&Rotbe':  (76, -8, 'Sabt'),
        '4500-Site-To-Roof':(88, -8, 'Site-Roof'),
        'NIBN-Tarasht':     (68, -16, 'Tarasht'),
        '1841-ISC':         (92, -16, 'ISC Test'),
    }

    # Parse branch/endpoint data from router configs
    branches_by_prov = _parse_branches_by_province(router_dir)

    parsed = {}
    node_categories = {}
    all_files = []
    for subdir in ['', 'Core Routers', 'Core Switches']:
        scan_dir = os.path.join(router_dir, subdir) if subdir else router_dir
        if not os.path.exists(scan_dir): continue
        for fname in sorted(os.listdir(scan_dir)):
            fpath = os.path.join(scan_dir, fname)
            if not os.path.isfile(fpath) or os.path.getsize(fpath) < 100: continue
            if fname in ('new file',): continue
            all_files.append((fname, fpath, subdir))

    # Parse all configs
    for fname, fpath, subdir in all_files:
        info = _parse_router_config_v2(fpath)
        if not info or not info['hostname']: continue
        category, model = _get_device_category(info['hostname'], fname, subdir)
        abbr = _extract_province_abbr(info['hostname'])
        sw_prov = _extract_switch_province(info['hostname']) if category == 'core-switch' else None

        # Position logic
        if category == 'core-switch' and sw_prov and sw_prov in PROVINCE_MAP_INFO:
            pinfo = PROVINCE_MAP_INFO[sw_prov]
            x, y, label = pinfo['x'] + 8, pinfo['y'] + 6, pinfo['fa'] + ' SW'
        elif category == 'core-router' and info['hostname'] in CORE_POSITIONS:
            cp = CORE_POSITIONS[info['hostname']]
            x, y, label = cp[0], cp[1], cp[2]
        elif category == 'core-router':
            # Fallback for unknown core routers
            x, y, label = 45, -5, CORE_DEVICE_MAP.get(info['hostname'], {}).get('fa', info['hostname'])
        elif abbr in PROVINCE_MAP_INFO:
            pinfo = PROVINCE_MAP_INFO[abbr]
            x, y, label = pinfo['x'], pinfo['y'], pinfo['fa']
        else:
            x, y, label = 50, 50, abbr

        # Attach branch/bajeh/ATM/kiosk data for provincial routers
        # Merge from ALL matching province keys (e.g. TEH+TehB both go to OSTehran)
        prov_branches = {'branches': [], 'bajes': [], 'atms': [], 'kiosks': []}
        if category == 'provincial-router':
            for prov_key, bdata in branches_by_prov.items():
                mapped = _PROV_TO_ROUTER_ABBR_BR.get(prov_key, prov_key)
                if mapped == abbr or prov_key == abbr:
                    for k in ('branches', 'bajes', 'atms', 'kiosks'):
                        for item in bdata.get(k, []):
                            if item not in prov_branches[k]:
                                prov_branches[k].append(item)

        branches = prov_branches.get('branches', [])
        bajes = prov_branches.get('bajes', [])
        atms = prov_branches.get('atms', [])
        kiosks = prov_branches.get('kiosks', [])

        node = {
            'id': info['hostname'], 'abbr': abbr, 'label': label,
            'x': x, 'y': y, 'model': model, 'category': category,
            'subdir': subdir, 'province': sw_prov or abbr,
            'interfaces': info['interfaces'],
            'tunnels': info['tunnels'],
            'nat_rules': info['nat_rules'],
            'nat_interfaces': info['nat_interfaces'],
            'ospf': info['ospf_processes'],
            'static_routes': info['static_routes'],
            'access_lists': info['access_lists'],
            'crypto_maps': info['crypto_maps'],
            'interfaces_count': len(info['interfaces']),
            'tunnels_count': len(info['tunnels']),
            'nat_count': len(info['nat_rules']),
            'ospf_count': len(info['ospf_processes']),
            'static_routes_count': len(info['static_routes']),
            'acl_count': len(info['access_lists']),
            'branches': branches,
            'bajes': bajes,
            'atms': atms,
            'kiosks': kiosks,
            'branches_count': len(branches),
            'bajes_count': len(bajes),
            'atms_count': len(atms),
            'kiosks_count': len(kiosks),
            'endpoints_total': len(branches) + len(bajes) + len(atms) + len(kiosks),
        }
        nodes.append(node)
        parsed[info['hostname']] = info
        node_categories[info['hostname']] = category

    # ── Build links based on REAL tunnel analysis ──
    seen = set()
    def add_link(src, tgt, link_type, label=''):
        if src not in node_categories or tgt not in node_categories:
            return  # skip if device not found
        lk = tuple(sorted([src, tgt]))
        if lk not in seen and src != tgt:
            seen.add(lk)
            links.append({'source': src, 'target': tgt, 'type': link_type,
                          'tunnel': label, 'description': label,
                          'src_ip': '', 'dst_ip': ''})

    provincial_list = sorted([h for h, c in node_categories.items() if c == 'provincial-router'])
    core_set = {h for h, c in node_categories.items() if c == 'core-router'}

    # ── 1) ASR1006-WAN-MB → ALL provinces (MPLS backbone, Tunnel1-36) ──
    for prov in provincial_list:
        add_link('ASR1006-WAN-MB', prov, 'mpls', 'MPLS Backbone')

    # ── 2) WAN-INTR1 → ALL provinces (WAN tunnels via 10.30.42.200) ──
    for prov in provincial_list:
        add_link('WAN-INTR1', prov, 'wan', 'WAN Link')

    # ── 3) WAN-INTR2 → ALL provinces (WAN tunnels via 10.30.42.201) ──
    for prov in provincial_list:
        add_link('WAN-INTR2', prov, 'wan', 'WAN Link')

    # ── 4) ISR-APN-RO → ALL provinces (APN tunnels via 10.250.46.1) ──
    for prov in provincial_list:
        add_link('ISR-APN-RO', prov, 'apn', 'APN Link')

    # ── 5) Core-to-Core interconnections (based on actual config analysis) ──
    # WAN backbone triangle
    add_link('ASR1006-WAN-MB', 'WAN-INTR1', 'backbone', 'WAN Backbone')
    add_link('ASR1006-WAN-MB', 'WAN-INTR2', 'backbone', 'WAN Backbone')
    add_link('WAN-INTR1', 'WAN-INTR2', 'backbone', 'WAN Backbone')
    # APN network
    add_link('ISR-APN-RO', 'APN-INT-HUB', 'backbone', 'APN Core')
    add_link('ISR-APN-RO', 'ASR1006-WAN-MB', 'backbone', 'APN-WAN')
    # Intranet / services
    add_link('INT-4451', 'ASR1006-WAN-MB', 'core', 'Core Link')
    add_link('INT-4451', 'WAN-INTR1', 'core', 'Core Link')
    # External edge
    add_link('EXT-Edge-4451', 'ASR1006-WAN-MB', 'core', 'Core Link')
    add_link('EXT-Edge-4451', 'EXT-AGG', 'core', 'Aggregation')
    # PSP & BKC
    add_link('PSP-4451', 'ASR1006-WAN-MB', 'core', 'Core Link')
    add_link('BKC-4451', 'ASR1006-WAN-MB', 'core', 'Core Link')
    # Aggregation switches
    add_link('AGG-WAN-SW', 'ASR1006-WAN-MB', 'core', 'Aggregation')
    add_link('AGG-WAN-SW', 'WAN-INTR1', 'core', 'Aggregation')
    add_link('EXT-AGG', 'ASR1006-WAN-MB', 'core', 'Aggregation')
    # MAGFA, NIBN, HTSC, etc.
    add_link('BKI-MAGFA', 'ASR1006-WAN-MB', 'core', 'Core Link')
    add_link('3825-NIBN', 'ASR1006-WAN-MB', 'core', 'Core Link')
    add_link('Router-HTSC', 'ASR1006-WAN-MB', 'core', 'Core Link')
    add_link('SW-Roof-To-Site', '4500-Site-To-Roof', 'core', 'Site Link')
    add_link('4500-Site-To-Roof', 'ASR1006-WAN-MB', 'core', 'Core Link')
    # Legacy
    add_link('7206-STM1', 'ASR1006-WAN-MB', 'core', 'Legacy')
    add_link('V-Jahad-3825', 'ASR1006-WAN-MB', 'core', 'Legacy')
    add_link('4451-PBN', 'ASR1006-WAN-MB', 'core', 'Core Link')
    add_link('2821-Gostaresh', 'ASR1006-WAN-MB', 'core', 'Core Link')
    add_link('2821-Mizan', 'ASR1006-WAN-MB', 'core', 'Core Link')
    add_link('3825-Sabt&Rotbe', 'ASR1006-WAN-MB', 'core', 'Core Link')
    add_link('NIBN-Tarasht', '3825-NIBN', 'core', 'NIBN Link')
    add_link('1841-ISC', 'ASR1006-WAN-MB', 'core', 'Core Link')

    # ── 6) Core switches → their province's router ──
    prov_to_router = {}
    for n in nodes:
        if n['category'] == 'provincial-router':
            a = n.get('abbr', '')
            if a and a in PROVINCE_MAP_INFO:
                prov_to_router[a] = n['id']

    for n in nodes:
        if n['category'] == 'core-switch':
            prov = n.get('province', '')
            matched = prov_to_router.get(prov)
            if matched:
                add_link(matched, n['id'], 'lan', 'LAN')
            else:
                # Try connecting to nearest hub
                add_link('ASR1006-WAN-MB', n['id'], 'lan', 'LAN')

    # ── Count by type ──
    type_counts = {}
    for l in links:
        t = l['type']
        type_counts[t] = type_counts.get(t, 0) + 1

    core_count = sum(1 for n in nodes if n['category'] == 'core-router')
    switch_count = sum(1 for n in nodes if n['category'] == 'core-switch')
    provincial_count = sum(1 for n in nodes if n['category'] == 'provincial-router')

    return jsonify({
        'nodes': nodes, 'links': links,
        'total_routers': len(nodes), 'total_links': len(links),
        'core_count': core_count,
        'switch_count': switch_count,
        'provincial_count': provincial_count,
        'link_types': type_counts,
        '_version': 'v6-accurate-topology',
    })

# Core device positioning ring
CORE_DEVICE_MAP = {
    'ASR1006-WAN-MB':{'fa':'WAN Main','x':45,'y':30,'role':'core'},
    'INT-4451':{'fa':'Intranet','x':50,'y':27,'role':'core'},
    'WAN-INTR1':{'fa':'WAN INTR1','x':40,'y':27,'role':'core'},
    'WAN-INTR2':{'fa':'WAN INTR2','x':40,'y':33,'role':'core'},
    'EXT-Edge-4451':{'fa':'EXT Edge','x':50,'y':33,'role':'core'},
    'BKC-4451':{'fa':'BKC','x':43,'y':25,'role':'core'},
    'PSP-4451':{'fa':'PSP','x':47,'y':25,'role':'core'},
    'ISR-APN-RO':{'fa':'APN Router','x':50,'y':36,'role':'core'},
    'APN-INT-HUB':{'fa':'APN HUB','x':53,'y':30,'role':'core'},
    'AGG-WAN-SW':{'fa':'AGG WAN','x':40,'y':36,'role':'core'},
    'EXT-AGG':{'fa':'EXT AGG','x':53,'y':33,'role':'core'},
    'BKI-MAGFA':{'fa':'MAGFA','x':53,'y':27,'role':'core'},
    '7206-STM1':{'fa':'STM1','x':43,'y':36,'role':'core'},
    'V-Jahad-3825':{'fa':'Jahad','x':47,'y':36,'role':'core'},
    '4451-PBN':{'fa':'PBN','x':53,'y':36,'role':'core'},
    '2821-Gostaresh':{'fa':'Gostaresh','x':40,'y':30,'role':'core'},
    '2821-Mizan':{'fa':'Mizan MCI','x':37,'y':30,'role':'core'},
    '3825-NIBN':{'fa':'NIBN','x':47,'y':33,'role':'core'},
    '3825-Sabt&Rotbe':{'fa':'Sabt','x':37,'y':33,'role':'core'},
    '4500-Site-To-Roof':{'fa':'Site-Roof','x':37,'y':36,'role':'core'},
    'Router-HTSC':{'fa':'HTSC','x':43,'y':33,'role':'core'},
    'SW-Roof-To-Site':{'fa':'Roof-Site','x':50,'y':30,'role':'core'},
    'NIBN-Tarasht':{'fa':'Tarasht','x':37,'y':27,'role':'core'},
    '1841-ISC':{'fa':'ISC Test','x':53,'y':25,'role':'core'},
}


# ==================== NAT DIAGRAM PARSER ====================

def _sdb_scan_all_routers():
    """Return list of (filepath, fname, category) for every router/switch config."""
    base = os.path.dirname(__file__)
    router_dir   = os.path.join(base, 'Router')
    core_r_dir   = os.path.join(router_dir, 'Core Routers')
    core_sw_dir  = os.path.join(router_dir, 'Core Switches')
    results = []

    def _add_dir(directory, category):
        if not os.path.isdir(directory):
            return
        for fname in sorted(os.listdir(directory)):
            fpath = os.path.join(directory, fname)
            if os.path.isfile(fpath) and os.path.getsize(fpath) >= 100:
                results.append((fpath, fname, category))

    # Provincial routers sit directly in Router/ (not in subdirs)
    if os.path.isdir(router_dir):
        for fname in sorted(os.listdir(router_dir)):
            fpath = os.path.join(router_dir, fname)
            if os.path.isfile(fpath) and os.path.getsize(fpath) >= 100:
                results.append((fpath, fname, 'provincial-router'))

    _add_dir(core_r_dir,  'core-router')
    _add_dir(core_sw_dir, 'core-switch')
    return results


def _sdb_parse_nat_full(filepath):
    """Deep-parse one Cisco config file and return complete NAT flow data."""
    import re as _r
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as fh:
            content = fh.read()
    except Exception:
        return None

    result = {
        'hostname': '',
        'inside_interfaces': [],
        'outside_interfaces': [],
        'nat_rules': [],
        'nat_pools': [],
        'acl_content': {},
    }

    m = _r.search(r'^hostname\s+(.+)', content, _r.MULTILINE)
    if m:
        result['hostname'] = m.group(1).strip()

    # ── Interface blocks ────────────────────────────────────────────────────
    iface_re = _r.compile(
        r'^interface\s+(\S+)(.*?)(?=^interface\s|\Z)',
        _r.MULTILINE | _r.DOTALL
    )
    for im in iface_re.finditer(content):
        block = im.group(2)
        nat_side = _r.search(r'ip nat\s+(inside|outside)', block)
        if not nat_side:
            continue
        ip_m   = _r.search(r'ip address\s+(\d+\.\d+\.\d+\.\d+)\s+(\d+\.\d+\.\d+\.\d+)', block)
        desc_m = _r.search(r'description\s+(.+)', block)
        entry = {
            'name':        im.group(1),
            'ip':          ip_m.group(1)   if ip_m   else '',
            'mask':        ip_m.group(2)   if ip_m   else '',
            'description': desc_m.group(1).strip().strip('"') if desc_m else '',
        }
        if nat_side.group(1) == 'inside':
            result['inside_interfaces'].append(entry)
        else:
            result['outside_interfaces'].append(entry)

    # ── NAT rules ────────────────────────────────────────────────────────────
    for nm in _r.finditer(r'^ip nat\s+(?:inside|outside)\s+source\s+(.+)',
                           content, _r.MULTILINE):
        rt = nm.group(1).strip()
        if 'static' in rt:
            parts = rt.split()
            rule = {'type': 'static'}
            idx = 1  # skip 'static'
            if len(parts) > idx and parts[idx] in ('tcp', 'udp'):
                rule['protocol']    = parts[idx]
                rule['inside_ip']   = parts[idx+1] if len(parts) > idx+1 else ''
                rule['inside_port'] = parts[idx+2] if len(parts) > idx+2 else ''
                rule['outside_ip']  = parts[idx+3] if len(parts) > idx+3 else ''
                rule['outside_port']= parts[idx+4] if len(parts) > idx+4 else ''
            else:
                rule['inside_ip']  = parts[idx]   if len(parts) > idx   else ''
                rule['outside_ip'] = parts[idx+1] if len(parts) > idx+1 else ''
        else:
            acl_m   = _r.search(r'list\s+(\S+)',      rt)
            pool_m  = _r.search(r'pool\s+(\S+)',      rt)
            iface_m = _r.search(r'interface\s+(\S+)', rt)
            rule = {
                'type':          'dynamic',
                'acl':           acl_m.group(1)   if acl_m   else '',
                'pool':          pool_m.group(1)  if pool_m  else '',
                'overload_intf': iface_m.group(1) if iface_m else '',
                'overload':      'overload' in rt,
                'pat':           'overload' in rt,
            }
        result['nat_rules'].append(rule)

    # ── NAT pools ────────────────────────────────────────────────────────────
    for pm in _r.finditer(
        r'^ip nat pool\s+(\S+)\s+(\S+)\s+(\S+)\s+'
        r'(?:netmask\s+(\S+)|prefix-length\s+(\S+))',
        content, _r.MULTILINE
    ):
        result['nat_pools'].append({
            'name':     pm.group(1),
            'start_ip': pm.group(2),
            'end_ip':   pm.group(3),
            'netmask':  pm.group(4) or '',
            'prefix':   pm.group(5) or '',
        })

    # ── Extended ACLs ────────────────────────────────────────────────────────
    ext_acl_re = _r.compile(
        r'^ip access-list extended\s+(\S+)\s*\n(.*?)'
        r'(?=^ip access-list\s|^router\s|^!\s*\n!\s*\n|^ip nat\s|\Z)',
        _r.MULTILINE | _r.DOTALL
    )
    for am in ext_acl_re.finditer(content):
        entries = []
        for em in _r.finditer(
            r'^\s*(?:\d+\s+)?(permit|deny)\s+(\S+)\s+(\S+)(?:\s+(\S+))?'
            r'(?:\s+(\S+))?(?:\s+(\S+))?',
            am.group(2), _r.MULTILINE
        ):
            entries.append({
                'action':   em.group(1),
                'proto':    em.group(2),
                'src':      em.group(3),
                'src_wild': em.group(4) or '',
                'dst':      em.group(5) or 'any',
                'dst_wild': em.group(6) or '',
            })
        if entries:
            result['acl_content'][am.group(1)] = entries

    # ── Standard ACLs ────────────────────────────────────────────────────────
    std_acl_re = _r.compile(
        r'^ip access-list standard\s+(\S+)\s*\n(.*?)'
        r'(?=^ip access-list\s|^router\s|^!\s*\n!\s*\n|\Z)',
        _r.MULTILINE | _r.DOTALL
    )
    for am in std_acl_re.finditer(content):
        entries = []
        for em in _r.finditer(r'^\s*(?:\d+\s+)?(permit|deny)\s+(\S+)(?:\s+(\S+))?',
                               am.group(2), _r.MULTILINE):
            src = em.group(2)
            wild = em.group(3) or ''
            if src in ('any', 'host'):
                src = f"{src} {wild}".strip()
                wild = ''
            entries.append({'action': em.group(1), 'proto': 'ip',
                            'src': src, 'src_wild': wild, 'dst': 'any', 'dst_wild': ''})
        if entries:
            result['acl_content'][am.group(1)] = entries

    # ── Numbered ACLs ────────────────────────────────────────────────────────
    for am in _r.finditer(r'^access-list\s+(\S+)\s+(permit|deny)\s+(.+)',
                           content, _r.MULTILINE):
        acl_name = am.group(1)
        pts = am.group(3).strip().split()
        if acl_name not in result['acl_content']:
            result['acl_content'][acl_name] = []
        result['acl_content'][acl_name].append({
            'action':   am.group(2),
            'proto':    'ip',
            'src':      pts[0] if pts else '',
            'src_wild': pts[1] if len(pts) > 1 else '',
            'dst':      pts[2] if len(pts) > 2 else 'any',
            'dst_wild': pts[3] if len(pts) > 3 else '',
        })

    return result


# Province abbreviation map (hostname → abbreviation) re-used from topology
_PROV_ABBR_RE = None
def _sdb_abbr_from_hostname(hostname):
    """Extract province abbreviation from router hostname (best-effort)."""
    import re as _r
    if not hostname:
        return ''
    m = _r.search(r'(?:^|[-_])([A-Z]{2,6})(?:[-_\d]|$)', hostname)
    return m.group(1) if m else ''


@app.route('/api/network-map/nat-diagram', methods=['GET'])
def nat_diagram_api():
    """Return full NAT flow data for all routers/switches for visualization."""
    PROVINCE_FA = {
        'ALZ':'البرز','ARD':'اردبیل','AZGH':'آذربایجان غربی','AZSH':'آذربایجان شرقی',
        'BSH':'بوشهر','CHB':'چهارمحال','ESF':'اصفهان','FRS':'فارس','GIL':'گیلان',
        'GLS':'گلستان','HMD':'همدان','HMZ':'هرمزگان','ILM':'ایلام','KHR':'خراسان رضوی',
        'KHRJ':'خراسان جنوبی','KHZ':'خوزستان','KNB':'کرمانشاه','KRD':'کردستان',
        'KRM':'کرمان','KRMJ':'کرمانشاه','KRSH':'خراسان شمالی','LOR':'لرستان',
        'MAZ':'مازندران','MRZ':'مرکزی','QOM':'قم','QZV':'قزوین','SMN':'سمنان',
        'SNB':'سمنان','YZD':'یزد','ZNJ':'زنجان',
    }
    all_files = _sdb_scan_all_routers()
    devices = []

    for filepath, fname, category in all_files:
        data = _sdb_parse_nat_full(filepath)
        if not data or not data['nat_rules']:
            continue

        # Enrich dynamic rules with ACL entries and overload IP
        for rule in data['nat_rules']:
            if rule.get('type') == 'dynamic' and rule.get('acl'):
                rule['acl_entries'] = data['acl_content'].get(rule['acl'], [])
                rule['overload_ip'] = ''
                for oi in data['outside_interfaces']:
                    if oi['name'] == rule.get('overload_intf', ''):
                        rule['overload_ip'] = oi['ip']
                        break

        abbr = _sdb_abbr_from_hostname(data['hostname'])
        devices.append({
            'hostname':          data['hostname'],
            'label':             PROVINCE_FA.get(abbr, data['hostname']),
            'abbr':              abbr,
            'category':          category,
            'source_file':       fname,
            'inside_interfaces': data['inside_interfaces'],
            'outside_interfaces':data['outside_interfaces'],
            'nat_rules':         data['nat_rules'],
            'nat_pools':         data['nat_pools'],
        })

    order = {'core-router': 0, 'provincial-router': 1, 'core-switch': 2}
    devices.sort(key=lambda d: (order.get(d['category'], 3), d['hostname']))

    return jsonify({
        'devices':     devices,
        'total':       len(devices),
        'total_rules': sum(len(d['nat_rules']) for d in devices),
    })


# ==================== CUSTOM TRANSLATIONS API ====================
@app.route('/api/translations', methods=['GET'])
def get_translations():
    """Get all custom translations"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT id, name_en, name_fa, added_by, added_at FROM custom_translations ORDER BY added_at DESC")
        rows = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({'status': 'ok', 'translations': rows, 'builtin_count': len(FINGLISH_DICT)})
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/translations', methods=['POST'])
def add_translation():
    """Add a new custom translation (Finglish→Persian)"""
    try:
        data = request.json
        name_en = (data.get('name_en') or '').strip()
        name_fa = (data.get('name_fa') or '').strip()
        username = data.get('username', '')
        if not name_en or not name_fa:
            return jsonify({'status': 'error', 'error': 'نام انگلیسی و فارسی هر دو الزامی است'}), 400
        conn = get_db()
        cursor = conn.cursor()
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute("INSERT OR REPLACE INTO custom_translations (name_en, name_fa, added_by, added_at) VALUES (?,?,?,?)",
                       (name_en, name_fa, username, now))
        conn.commit()
        conn.close()
        # Update in-memory dicts and invalidate fuzzy cache (thread-safe)
        with _translation_lock:
            FINGLISH_DICT[name_en] = name_fa
            if name_fa not in PERSIAN_TO_FINGLISH:
                PERSIAN_TO_FINGLISH[name_fa] = []
            if name_en not in PERSIAN_TO_FINGLISH[name_fa]:
                PERSIAN_TO_FINGLISH[name_fa].append(name_en)
            _rebuild_fuzzy_cache()   # keep fuzzy index up-to-date
        log_activity('info', 'افزودن ترجمه', f'{name_en} → {name_fa}', username)
        return jsonify({'status': 'ok', 'message': f'ترجمه "{name_en}" → "{name_fa}" اضافه شد'})
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/api/translations/<int:tid>', methods=['DELETE'])
def delete_translation(tid):
    """Delete a custom translation"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT name_en, name_fa FROM custom_translations WHERE id = ?", (tid,))
        row = cursor.fetchone()
        if row:
            en = row['name_en']
            fa = row['name_fa']
            cursor.execute("DELETE FROM custom_translations WHERE id = ?", (tid,))
            conn.commit()
            # Remove from in-memory dicts and rebuild fuzzy cache (thread-safe)
            with _translation_lock:
                FINGLISH_DICT.pop(en, None)
                if fa in PERSIAN_TO_FINGLISH:
                    try:
                        PERSIAN_TO_FINGLISH[fa].remove(en)
                    except ValueError:
                        pass
                    if not PERSIAN_TO_FINGLISH[fa]:
                        del PERSIAN_TO_FINGLISH[fa]
                _rebuild_fuzzy_cache()
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

# ==================== AI ASSISTANT (Multi-Provider) ====================

def _get_ai_config():
    """Returns (provider, model, api_key, base_url).
    Env vars:
      AI_PROVIDER   = "anthropic" (default) | "openai"
      ANTHROPIC_API_KEY — used when provider=anthropic
      AI_API_KEY / OPENAI_API_KEY — used when provider=openai
      AI_BASE_URL   = "http://localhost:11434/v1"  (Ollama default)
      AI_MODEL      = model name override
    """
    provider = os.environ.get('AI_PROVIDER', 'anthropic').lower()
    if provider == 'anthropic':
        model = os.environ.get('AI_MODEL', 'claude-sonnet-4-6')
        api_key = os.environ.get('ANTHROPIC_API_KEY', '')
        base_url = None
    else:
        model = os.environ.get('AI_MODEL', 'llama3.2')
        api_key = os.environ.get('AI_API_KEY') or os.environ.get('OPENAI_API_KEY') or 'ollama'
        base_url = os.environ.get('AI_BASE_URL', 'http://localhost:11434/v1')
    return provider, model, api_key, base_url


def _tools_openai_fmt():
    """Convert _AI_TOOLS (Anthropic format) → OpenAI function calling format."""
    return [
        {"type": "function", "function": {
            "name": t["name"],
            "description": t["description"],
            "parameters": t["input_schema"]
        }}
        for t in _AI_TOOLS
    ]

_AI_SYSTEM_PROMPT = """شما دستیار هوشمند شبکه بانک کشاورزی ایران (BKI) هستید.
You are the AI Network Assistant for Bank Keshavarzi Iran (BKI) IPAM Portal.

Your capabilities:
- Answer questions about IP address management (LAN, APN INT/Mali, VPLS/MPLS tunnels, Tunnel200, Intranet)
- Search and analyze IP reservations, usage patterns, and capacity
- Identify expiring reservations before they cause issues
- Analyze province-level network usage and suggest optimizations
- Help users understand network topology and configurations
- Provide actionable recommendations based on real data

Key facts about the network:
- LAN IPs: 10.x.y.z/24 subnets for branch offices (each branch gets a /24)
- APN INT: Mobile network tunnels for ATMs and kiosks
- APN Mali: Financial mobile network endpoints
- VPLS/MPLS tunnels: Province-level L2/L3 VPN connectivity
- Tunnel200: Backup tunnel endpoints
- PTMP: Point-to-Multipoint serial connections for remote branches
- 33 provinces in Iran: KHZ=خوزستان, FRS=فارس, ESF=اصفهان, KHR=خراسان رضوی,
  TehB/TEH=تهران, AZSH=آذربایجان شرقی, AZGH=آذربایجان غربی, etc.

Rules:
- Always respond in the SAME language the user writes in (Persian or English)
- When showing IPs or technical data, use `monospace` format
- Be concise but complete - don't pad responses
- Always use the provided tools to get real data before answering data questions
- If a question is outside your scope, say so clearly"""

_AI_TOOLS = [
    {
        "name": "get_network_stats",
        "description": "Get comprehensive network statistics: total/free/used IPs across all service types, pending reservations, expired leases",
        "input_schema": {"type": "object", "properties": {}, "required": []}
    },
    {
        "name": "get_province_details",
        "description": "Get detailed IP usage statistics for a specific province including LAN, APN, VPLS data",
        "input_schema": {
            "type": "object",
            "properties": {
                "province": {"type": "string", "description": "Province abbreviation (e.g. KHZ, FRS, ESF, KHR, TehB, AZSH, AZGH, GIL, MAZ, etc.)"}
            },
            "required": ["province"]
        }
    },
    {
        "name": "get_free_ips",
        "description": "Get list of available/free IP subnets for a province or all provinces",
        "input_schema": {
            "type": "object",
            "properties": {
                "province": {"type": "string", "description": "Province abbreviation. Omit for all provinces"},
                "limit": {"type": "integer", "description": "Max results to return (default 20, max 50)"}
            },
            "required": []
        }
    },
    {
        "name": "get_expiring_reservations",
        "description": "Get IP reservations expiring within N days - critical for proactive management",
        "input_schema": {
            "type": "object",
            "properties": {
                "days": {"type": "integer", "description": "Days ahead to check (default 10)"}
            },
            "required": []
        }
    },
    {
        "name": "search_branches",
        "description": "Search for branch offices by name across all service types (LAN, APN, VPLS, PTMP)",
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Branch name or partial name"},
                "province": {"type": "string", "description": "Optional: filter by province abbreviation"}
            },
            "required": ["query"]
        }
    },
    {
        "name": "get_recent_activity",
        "description": "Get recent IP reservation and release activity log",
        "input_schema": {
            "type": "object",
            "properties": {
                "limit": {"type": "integer", "description": "Number of activities to return (default 15)"}
            },
            "required": []
        }
    },
    {
        "name": "analyze_network",
        "description": "Analyze network utilization: top provinces by usage, bottlenecks, high-demand areas, underutilized subnets",
        "input_schema": {"type": "object", "properties": {}, "required": []}
    },
    {
        "name": "get_tunnel_stats",
        "description": "Get tunnel IP statistics by type: APN INT, APN Mali, Tunnel200, VPLS/MPLS, Intranet tunnels",
        "input_schema": {"type": "object", "properties": {}, "required": []}
    }
]


def _ai_run_tool(name, inp):
    """Execute an AI tool call and return JSON string result."""
    try:
        conn = get_db()
        cursor = conn.cursor()

        if name == "get_network_stats":
            stats = {}
            for tbl, key in [('lan_ips', 'lan'), ('apn_ips', 'apn_int'), ('apn_mali', 'apn_mali'),
                              ('tunnel200_ips', 'tunnel200'), ('vpls_tunnels', 'vpls'),
                              ('intranet_tunnels', 'intranet')]:
                try:
                    cursor.execute(f"SELECT COUNT(*) t, "
                                   f"SUM(CASE WHEN LOWER(status)='free' OR status IS NULL THEN 1 ELSE 0 END) f, "
                                   f"SUM(CASE WHEN LOWER(status)!='free' AND status IS NOT NULL THEN 1 ELSE 0 END) u "
                                   f"FROM {tbl}")
                    r = cursor.fetchone()
                    stats[key] = {'total': r[0] or 0, 'free': r[1] or 0, 'used': r[2] or 0}
                except Exception:
                    stats[key] = {'total': 0, 'free': 0, 'used': 0}
            cursor.execute("SELECT COUNT(*) FROM reserved_ips WHERE LOWER(status)='reserved'")
            stats['pending_reservations'] = cursor.fetchone()[0]
            today = datetime.now().strftime('%Y-%m-%d')
            cursor.execute("SELECT COUNT(*) FROM reserved_ips WHERE expiry_date < ? AND LOWER(status)='reserved'", (today,))
            stats['expired_unreleased'] = cursor.fetchone()[0]
            conn.close()
            return json.dumps(stats, ensure_ascii=False)

        elif name == "get_province_details":
            prov = inp.get('province', '').upper()
            result = {'province': prov}
            cursor.execute("""SELECT COUNT(*) t,
                SUM(CASE WHEN LOWER(status)='free' THEN 1 ELSE 0 END) free,
                SUM(CASE WHEN LOWER(status)='reserved' THEN 1 ELSE 0 END) reserved,
                SUM(CASE WHEN LOWER(status) IN ('used','activated') THEN 1 ELSE 0 END) used
                FROM lan_ips WHERE UPPER(province)=?""", (prov,))
            r = cursor.fetchone()
            result['lan'] = {'total': r[0], 'free': r[1] or 0, 'reserved': r[2] or 0, 'used': r[3] or 0}
            cursor.execute("""SELECT branch_name, username, reservation_date, expiry_date, status
                FROM reserved_ips WHERE UPPER(province)=? ORDER BY expiry_date LIMIT 20""", (prov,))
            result['reservations'] = [dict(zip([c[0] for c in cursor.description], row)) for row in cursor.fetchall()]
            cursor.execute("SELECT COUNT(*) FROM apn_ips WHERE UPPER(province)=?", (prov,))
            result['apn_int_count'] = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM apn_mali WHERE UPPER(province)=?", (prov,))
            result['apn_mali_count'] = cursor.fetchone()[0]
            conn.close()
            return json.dumps(result, ensure_ascii=False)

        elif name == "get_free_ips":
            prov = inp.get('province', '').upper()
            limit = min(int(inp.get('limit', 20)), 50)
            if prov:
                cursor.execute("SELECT province, octet2, octet3 FROM lan_ips WHERE LOWER(status)='free' AND UPPER(province)=? LIMIT ?", (prov, limit))
            else:
                cursor.execute("SELECT province, octet2, octet3 FROM lan_ips WHERE LOWER(status)='free' ORDER BY province LIMIT ?", (limit,))
            rows = cursor.fetchall()
            ips = [{'subnet': f"10.{r[1]}.{r[2]}.0/24", 'province': r[0]} for r in rows]
            conn.close()
            return json.dumps({'count': len(ips), 'free_subnets': ips}, ensure_ascii=False)

        elif name == "get_expiring_reservations":
            days = int(inp.get('days', 10))
            today = datetime.now().strftime('%Y-%m-%d')
            target = (datetime.now() + timedelta(days=days)).strftime('%Y-%m-%d')
            cursor.execute("""SELECT province, branch_name, username, expiry_date,
                CAST(JULIANDAY(expiry_date) - JULIANDAY(?) AS INTEGER) days_left
                FROM reserved_ips WHERE expiry_date BETWEEN ? AND ? AND LOWER(status)='reserved'
                ORDER BY expiry_date""", (today, today, target))
            rows = [dict(zip([c[0] for c in cursor.description], r)) for r in cursor.fetchall()]
            conn.close()
            return json.dumps({'total': len(rows), 'check_days': days, 'expiring': rows}, ensure_ascii=False)

        elif name == "search_branches":
            q = f"%{inp.get('query', '')}%"
            pf = inp.get('province', '').upper()
            results = []
            for tbl, stype in [('lan_ips', 'LAN'), ('apn_ips', 'APN INT'), ('apn_mali', 'APN Mali'),
                                ('vpls_tunnels', 'VPLS'), ('ptmp_connections', 'PTMP')]:
                prov_q = f"AND UPPER(province)=?" if pf else ""
                params = (q, pf) if pf else (q,)
                try:
                    cursor.execute(f"SELECT branch_name, province, status, username FROM {tbl} "
                                   f"WHERE branch_name LIKE ? {prov_q} LIMIT 8", params)
                    for r in cursor.fetchall():
                        results.append({'type': stype, 'branch': r[0], 'province': r[1],
                                        'status': r[2], 'user': r[3]})
                except Exception:
                    pass
            conn.close()
            return json.dumps({'total': len(results), 'results': results}, ensure_ascii=False)

        elif name == "get_recent_activity":
            limit = min(int(inp.get('limit', 15)), 50)
            activities = []
            if os.path.exists(ACTIVITY_LOG):
                try:
                    with open(ACTIVITY_LOG, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    activities = list(reversed(data[-limit:]))
                except Exception:
                    pass
            conn.close()
            return json.dumps({'total': len(activities), 'activities': activities}, ensure_ascii=False)

        elif name == "analyze_network":
            analysis = {}
            cursor.execute("""SELECT province, COUNT(*) total,
                SUM(CASE WHEN LOWER(status)='free' THEN 1 ELSE 0 END) free,
                SUM(CASE WHEN LOWER(status)!='free' AND status IS NOT NULL THEN 1 ELSE 0 END) used
                FROM lan_ips GROUP BY province ORDER BY total DESC LIMIT 15""")
            analysis['provinces_by_size'] = [dict(zip([c[0] for c in cursor.description], r)) for r in cursor.fetchall()]
            cursor.execute("""SELECT province, COUNT(*) cnt FROM reserved_ips
                WHERE LOWER(status)='reserved' GROUP BY province ORDER BY cnt DESC LIMIT 8""")
            analysis['most_pending_by_province'] = [dict(zip([c[0] for c in cursor.description], r)) for r in cursor.fetchall()]
            cursor.execute("""SELECT username, COUNT(*) cnt FROM reserved_ips
                WHERE LOWER(status)='reserved' GROUP BY username ORDER BY cnt DESC LIMIT 5""")
            analysis['top_users_by_reservations'] = [dict(zip([c[0] for c in cursor.description], r)) for r in cursor.fetchall()]
            today = datetime.now().strftime('%Y-%m-%d')
            cursor.execute("SELECT COUNT(*) FROM reserved_ips WHERE expiry_date < ? AND LOWER(status)='reserved'", (today,))
            analysis['expired_not_released'] = cursor.fetchone()[0]
            conn.close()
            return json.dumps(analysis, ensure_ascii=False)

        elif name == "get_tunnel_stats":
            stats = {}
            for tbl, label in [('apn_ips', 'APN INT'), ('apn_mali', 'APN Mali'),
                                ('tunnel200_ips', 'Tunnel 200'), ('vpls_tunnels', 'VPLS/MPLS'),
                                ('intranet_tunnels', 'Intranet')]:
                try:
                    cursor.execute(f"SELECT COUNT(*), "
                                   f"SUM(CASE WHEN LOWER(status)='free' OR status IS NULL THEN 1 ELSE 0 END) "
                                   f"FROM {tbl}")
                    r = cursor.fetchone()
                    total = r[0] or 0
                    free = r[1] or 0
                    stats[label] = {'total': total, 'free': free, 'used': total - free,
                                    'utilization_pct': round((total - free) / total * 100, 1) if total else 0}
                except Exception:
                    stats[label] = {'total': 0, 'free': 0, 'used': 0, 'utilization_pct': 0}
            conn.close()
            return json.dumps(stats, ensure_ascii=False)

        conn.close()
        return json.dumps({'error': f'Unknown tool: {name}'})
    except Exception as e:
        return json.dumps({'error': str(e)})


@app.route('/api/ai/config', methods=['GET'])
def ai_config():
    """Return current AI provider config (no secrets)."""
    provider, model, api_key, base_url = _get_ai_config()
    configured = bool(api_key and api_key not in ('', 'ollama'))
    # For Ollama/local, just having a base_url means it may be configured
    if provider == 'openai':
        configured = True  # local models don't need a real key
    return jsonify({
        'provider': provider,
        'model': model,
        'base_url': base_url,
        'configured': configured,
        'has_key': bool(api_key)
    })


@app.route('/api/ai/chat', methods=['POST'])
def ai_chat():
    """Non-streaming AI chat with agentic tool-use loop. Supports Anthropic + OpenAI-compatible providers."""
    username = get_current_user()
    if not username:
        return jsonify({'error': 'Authentication required'}), 401

    data = request.json or {}
    messages = data.get('messages', [])
    if not messages:
        return jsonify({'error': 'No messages provided'}), 400

    provider, model, api_key, base_url = _get_ai_config()

    # ---- Anthropic provider ----
    if provider == 'anthropic':
        if not api_key:
            return jsonify({'error': 'Set ANTHROPIC_API_KEY to enable AI', 'demo': True}), 503
        try:
            import anthropic
            client = anthropic.Anthropic(api_key=api_key)
            cur = list(messages)
            for _ in range(6):
                resp = client.messages.create(
                    model=model, max_tokens=2048, system=_AI_SYSTEM_PROMPT,
                    tools=_AI_TOOLS, messages=cur)
                if resp.stop_reason == 'tool_use':
                    results = []
                    for blk in resp.content:
                        if blk.type == 'tool_use':
                            results.append({'type': 'tool_result', 'tool_use_id': blk.id,
                                            'content': _ai_run_tool(blk.name, blk.input)})
                    cur.append({'role': 'assistant', 'content': resp.content})
                    cur.append({'role': 'user', 'content': results})
                else:
                    text = ''.join(b.text for b in resp.content if hasattr(b, 'text'))
                    return jsonify({'response': text,
                                    'tokens': resp.usage.input_tokens + resp.usage.output_tokens,
                                    'model': model, 'provider': 'anthropic'})
            return jsonify({'error': 'Max iterations reached'}), 500
        except ImportError:
            return jsonify({'error': 'Run: pip install anthropic', 'demo': True}), 503
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    # ---- OpenAI-compatible provider (Ollama, LM Studio, OpenRouter, etc.) ----
    try:
        import openai
        client = openai.OpenAI(api_key=api_key, base_url=base_url)
        # Build messages in OpenAI format (system message first)
        oai_msgs = [{'role': 'system', 'content': _AI_SYSTEM_PROMPT}]
        for m in messages:
            oai_msgs.append({'role': m['role'], 'content': m['content']})

        oai_tools = _tools_openai_fmt()
        for _ in range(6):
            kwargs = {'model': model, 'max_tokens': 2048, 'messages': oai_msgs}
            try:
                kwargs['tools'] = oai_tools
                resp = client.chat.completions.create(**kwargs)
            except Exception:
                # Model may not support tools — retry without
                kwargs.pop('tools', None)
                resp = client.chat.completions.create(**kwargs)

            choice = resp.choices[0]
            if choice.finish_reason == 'tool_calls' and choice.message.tool_calls:
                # Execute all tool calls
                tool_calls_msg = {'role': 'assistant', 'content': choice.message.content or '',
                                  'tool_calls': [
                                      {'id': tc.id, 'type': 'function',
                                       'function': {'name': tc.function.name, 'arguments': tc.function.arguments}}
                                      for tc in choice.message.tool_calls
                                  ]}
                oai_msgs.append(tool_calls_msg)
                for tc in choice.message.tool_calls:
                    try:
                        inp = json.loads(tc.function.arguments or '{}')
                    except Exception:
                        inp = {}
                    result = _ai_run_tool(tc.function.name, inp)
                    oai_msgs.append({'role': 'tool', 'tool_call_id': tc.id, 'content': result})
            else:
                text = choice.message.content or ''
                tokens = getattr(resp.usage, 'total_tokens', 0) if resp.usage else 0
                return jsonify({'response': text, 'tokens': tokens, 'model': model, 'provider': provider})

        return jsonify({'error': 'Max iterations reached'}), 500
    except ImportError:
        return jsonify({'error': 'Run: pip install openai', 'demo': True}), 503
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/ai/stream', methods=['POST'])
def ai_stream():
    """Streaming AI chat via SSE. Supports Anthropic + OpenAI-compatible providers."""
    username = get_current_user()
    if not username:
        def _unauth():
            yield 'data: {"error":"Authentication required"}\n\n'
        return Response(stream_with_context(_unauth()), mimetype='text/event-stream', status=401)

    data = request.json or {}
    messages = data.get('messages', [])
    provider, model, api_key, base_url = _get_ai_config()

    _sse_headers = {'X-Accel-Buffering': 'no', 'Cache-Control': 'no-cache'}

    # ---- Demo mode: no key for Anthropic ----
    if provider == 'anthropic' and not api_key:
        def _demo():
            tips = (
                "🤖 **AI Assistant - Demo Mode**\n\n"
                "برای فعال‌سازی کامل، یکی از روش‌های زیر را انتخاب کنید:\n\n"
                "**گزینه ۱ — Anthropic Claude (ابری):**\n"
                "`export ANTHROPIC_API_KEY=\"sk-ant-...\"`\n\n"
                "**گزینه ۲ — Ollama (آفلاین/محلی):**\n"
                "`export AI_PROVIDER=openai`\n"
                "`export AI_BASE_URL=http://localhost:11434/v1`\n"
                "`export AI_MODEL=llama3.2`\n\n"
                "**گزینه ۳ — LM Studio:**\n"
                "`export AI_PROVIDER=openai`\n"
                "`export AI_BASE_URL=http://localhost:1234/v1`\n\n"
                "**گزینه ۴ — OpenRouter (ابری، چند مدل):**\n"
                "`export AI_PROVIDER=openai`\n"
                "`export AI_API_KEY=sk-or-...`\n"
                "`export AI_BASE_URL=https://openrouter.ai/api/v1`\n"
                "`export AI_MODEL=meta-llama/llama-3.3-70b-instruct`"
            )
            for char in tips:
                yield f'data: {json.dumps({"text": char})}\n\n'
                time.sleep(0.008)
            yield f'data: {json.dumps({"done": True})}\n\n'
        return Response(stream_with_context(_demo()), mimetype='text/event-stream', headers=_sse_headers)

    # ---- Anthropic streaming ----
    def _anthropic_gen():
        try:
            import anthropic
            client = anthropic.Anthropic(api_key=api_key)
            cur = list(messages)
            for _ in range(6):
                tool_blocks = []
                cur_tool = None
                with client.messages.stream(
                    model=model, max_tokens=2048, system=_AI_SYSTEM_PROMPT,
                    tools=_AI_TOOLS, messages=cur
                ) as stream:
                    for event in stream:
                        etype = type(event).__name__
                        if etype == 'ContentBlockStart':
                            cb = getattr(event, 'content_block', None)
                            if cb and getattr(cb, 'type', '') == 'tool_use':
                                cur_tool = {'id': cb.id, 'name': cb.name, 'raw_input': ''}
                                yield f'data: {json.dumps({"tool_call": cb.name})}\n\n'
                        elif etype == 'ContentBlockDelta':
                            d = getattr(event, 'delta', None)
                            if d:
                                if hasattr(d, 'text') and d.text:
                                    yield f'data: {json.dumps({"text": d.text})}\n\n'
                                elif hasattr(d, 'partial_json') and cur_tool:
                                    cur_tool['raw_input'] += d.partial_json
                        elif etype == 'ContentBlockStop':
                            if cur_tool:
                                try:
                                    cur_tool['input'] = json.loads(cur_tool['raw_input'] or '{}')
                                except Exception:
                                    cur_tool['input'] = {}
                                tool_blocks.append(cur_tool)
                                cur_tool = None
                    final = stream.get_final_message()

                if final.stop_reason == 'tool_use':
                    results = []
                    for t in tool_blocks:
                        res = _ai_run_tool(t['name'], t.get('input', {}))
                        results.append({'type': 'tool_result', 'tool_use_id': t['id'], 'content': res})
                        yield f'data: {json.dumps({"tool_result": t["name"], "data_preview": res[:100]})}\n\n'
                    cur.append({'role': 'assistant', 'content': final.content})
                    cur.append({'role': 'user', 'content': results})
                else:
                    tokens = final.usage.input_tokens + final.usage.output_tokens
                    yield f'data: {json.dumps({"done": True, "tokens": tokens, "model": model})}\n\n'
                    return
            yield f'data: {json.dumps({"done": True})}\n\n'
        except ImportError:
            yield f'data: {json.dumps({"error": "Run: pip install anthropic"})}\n\n'
        except Exception as e:
            yield f'data: {json.dumps({"error": str(e)})}\n\n'

    # ---- OpenAI-compatible streaming (Ollama, LM Studio, OpenRouter…) ----
    def _openai_gen():
        try:
            import openai
            client = openai.OpenAI(api_key=api_key, base_url=base_url)
            oai_msgs = [{'role': 'system', 'content': _AI_SYSTEM_PROMPT}]
            for m in messages:
                oai_msgs.append({'role': m['role'], 'content': m['content']})

            oai_tools = _tools_openai_fmt()
            for _ in range(6):
                # Accumulate tool calls across streamed chunks
                tool_calls_buf = {}   # index → {id, name, args}
                text_buf = ''
                finish_reason = None

                try:
                    stream_kwargs = {'model': model, 'max_tokens': 2048,
                                     'messages': oai_msgs, 'tools': oai_tools, 'stream': True}
                    stream = client.chat.completions.create(**stream_kwargs)
                except Exception:
                    # Retry without tools if model doesn't support function calling
                    stream_kwargs.pop('tools', None)
                    stream = client.chat.completions.create(**stream_kwargs)

                for chunk in stream:
                    if not chunk.choices:
                        continue
                    delta = chunk.choices[0].delta
                    finish_reason = chunk.choices[0].finish_reason or finish_reason

                    if delta.content:
                        text_buf += delta.content
                        yield f'data: {json.dumps({"text": delta.content})}\n\n'

                    if delta.tool_calls:
                        for tc in delta.tool_calls:
                            idx = tc.index
                            if idx not in tool_calls_buf:
                                tool_calls_buf[idx] = {'id': tc.id or '', 'name': '', 'args': ''}
                                if tc.function and tc.function.name:
                                    tool_calls_buf[idx]['name'] = tc.function.name
                                    yield f'data: {json.dumps({"tool_call": tc.function.name})}\n\n'
                            if tc.function:
                                if tc.function.name and not tool_calls_buf[idx]['name']:
                                    tool_calls_buf[idx]['name'] = tc.function.name
                                    yield f'data: {json.dumps({"tool_call": tc.function.name})}\n\n'
                                if tc.function.arguments:
                                    tool_calls_buf[idx]['args'] += tc.function.arguments

                if finish_reason == 'tool_calls' and tool_calls_buf:
                    # Build assistant message with tool_calls
                    tc_list = []
                    for idx in sorted(tool_calls_buf.keys()):
                        tc = tool_calls_buf[idx]
                        tc_list.append({'id': tc['id'], 'type': 'function',
                                        'function': {'name': tc['name'], 'arguments': tc['args']}})
                    oai_msgs.append({'role': 'assistant', 'content': text_buf or '', 'tool_calls': tc_list})

                    # Execute tools
                    for tc in tc_list:
                        try:
                            inp = json.loads(tc['function']['arguments'] or '{}')
                        except Exception:
                            inp = {}
                        res = _ai_run_tool(tc['function']['name'], inp)
                        yield f'data: {json.dumps({"tool_result": tc["function"]["name"], "data_preview": res[:100]})}\n\n'
                        oai_msgs.append({'role': 'tool', 'tool_call_id': tc['id'], 'content': res})
                else:
                    yield f'data: {json.dumps({"done": True, "model": model, "provider": provider})}\n\n'
                    return

            yield f'data: {json.dumps({"done": True})}\n\n'
        except ImportError:
            yield f'data: {json.dumps({"error": "Run: pip install openai"})}\n\n'
        except Exception as e:
            yield f'data: {json.dumps({"error": str(e)})}\n\n'

    gen = _anthropic_gen if provider == 'anthropic' else _openai_gen
    return Response(stream_with_context(gen()), mimetype='text/event-stream', headers=_sse_headers)


# ==================== MIRROR API ====================

@app.route('/api/db/mirror/status')
def mirror_status():
    """وضعیت live.db mirror را برمی‌گرداند."""
    try:
        _init_live_db()
        conn = sqlite3.connect(LIVE_DB_PATH)
        pending  = conn.execute("SELECT COUNT(*) FROM _change_log WHERE merged=0").fetchone()[0]
        total    = conn.execute("SELECT COUNT(*) FROM _change_log").fetchone()[0]
        last_mrg = conn.execute("SELECT MAX(merged_at) FROM _change_log WHERE merged=1").fetchone()[0]
        oldest   = conn.execute("SELECT MIN(ts) FROM _change_log WHERE merged=0").fetchone()[0]
        by_tbl   = conn.execute(
            "SELECT table_name, COUNT(*) AS c FROM _change_log "
            "WHERE merged=0 GROUP BY table_name ORDER BY c DESC"
        ).fetchall()
        last_run = None
        try:
            last_run = conn.execute("SELECT MAX(ts) FROM _merge_log WHERE dry_run=0").fetchone()[0]
        except Exception:
            pass
        conn.close()
        return jsonify({
            'live_db':         LIVE_DB_PATH,
            'main_db':         DB_PATH,
            'live_size_kb':    round(os.path.getsize(LIVE_DB_PATH) / 1024, 1) if os.path.exists(LIVE_DB_PATH) else 0,
            'main_size_kb':    round(os.path.getsize(DB_PATH) / 1024, 1) if os.path.exists(DB_PATH) else 0,
            'pending_changes': pending,
            'total_logged':    total,
            'last_merged_at':  last_mrg,
            'oldest_pending':  oldest,
            'last_merge_run':  last_run,
            'by_table':        {r[0]: r[1] for r in by_tbl},
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/db/mirror/merge', methods=['POST'])
def mirror_merge():
    """تغییرات live.db را به network_ipam.db اعمال می‌کند (دستی)."""
    try:
        data     = request.json or {}
        username = data.get('username', '')
        dry_run  = bool(data.get('dry_run', False))
        if username != DB_ADMIN_USER:
            return jsonify({'error': 'فقط ادمین مجاز است'}), 403
        stats = _mirror_merge_to_main(dry_run=dry_run)
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== MAIN ====================
if __name__ == '__main__':
    print("=" * 70)
    print("🚀 Network Config Portal - COMPLETE FIXED VERSION")
    print("=" * 70)
    print(f"📂 Database: {DB_PATH}")
    print(f"👥 Users: {', '.join(ALLOWED_USERS)}")
    print(f"🔐 DB Admin: {DB_ADMIN_USER}")
    print(f"🗑️ Auto-Release Check: Every {AUTO_RELEASE_INTERVAL // 3600} hours")
    print("=" * 70)
    
    # Initialize live.db mirror and start nightly merge thread
    _init_live_db()
    t_mirror = threading.Thread(target=_nightly_mirror_thread, daemon=True)
    t_mirror.start()
    print("✅ [Mirror] Nightly merge thread started (runs at 23:30)")

    # Start auto-release thread for expired reservations
    start_auto_release_thread()
    
    if socketio and REMOTE_ENABLED:
        print("🔌 WebSocket enabled (SSH/Telnet/RDP)")
        socketio.run(app, host='0.0.0.0', port=5000, debug=False, use_reloader=False)
    else:
        print("📡 Running without WebSocket (SSH/Telnet/RDP disabled)")
        app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)
